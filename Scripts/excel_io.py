# -*- coding: utf-8 -*-
"""
Excel 读写与行/Key 抽象。
只做一件事：从 openpyxl 的 Workbook/Worksheet 读取或写入“按行、按首列 Key”的数据结构，
以及 Key 规范化、行相等判断，不包含合并算法或对比逻辑。
"""

import openpyxl

from config import SKIP_SHEET_PREFIX


# -----------------------------------------------------------------------------
# 单元格与 Key
# -----------------------------------------------------------------------------

def cell_str(c):
    """将单元格值转为用于比较的字符串；None 或空视为 ''。"""
    if c is None:
        return ""
    return str(c).strip()


def key_str_normalized(c):
    """
    合并时统一 Key：数字 1 与 1.0 视为同一行，避免漏检冲突。
    用于冲突检测与合并时以首列作为行唯一标识。
    """
    if c is None:
        return ""
    s = str(c).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return str(f)
    except ValueError:
        return s


# -----------------------------------------------------------------------------
# Sheet 名过滤
# -----------------------------------------------------------------------------

def should_skip_sheet(name):
    """仅跳过以配置前缀（如 #）开头的表名；Sheet1/Sheet2 等会参与合并与对比。"""
    s = (name or "").strip()
    return s.startswith(SKIP_SHEET_PREFIX)


def get_sheet_names(wb):
    """返回需要参与合并/对比的 Sheet 名称列表（按需可保持顺序）。"""
    return [n for n in wb.sheetnames if not should_skip_sheet(n)]


# -----------------------------------------------------------------------------
# 合并单元格取值（openpyxl 仅左上格有值，其余为 None）
# -----------------------------------------------------------------------------

def get_merged_cell_value(ws, row, col):
    """
    若 (row, col) 落在合并区域内，返回该区域左上角的值；
    否则返回该格子的值。用于避免合并格导致首列“消失”。
    """
    try:
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(row=rng.min_row, column=rng.min_col).value
    except Exception:
        pass
    return ws.cell(row=row, column=col).value


# -----------------------------------------------------------------------------
# 按行加载：每行为 list，可选跳过首列为空的行
# -----------------------------------------------------------------------------

def load_sheet_rows(ws, max_col=None):
    """
    加载 Sheet 所有数据行，每行为 list；第一列为空则跳过该行。
    用于对比模式或简单合并（不保留表头行号）。
    """
    if ws is None:
        return []
    rows = []
    if max_col is None:
        max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_col=max_col, values_only=True):
        row_list = list(row) if row else []
        while len(row_list) < max_col:
            row_list.append("")
        key = cell_str(row_list[0]) if row_list else ""
        if not key:
            continue
        rows.append(row_list)
    return rows


def load_sheet_rows_full(ws, max_col=None):
    """
    加载 Sheet 所有行（不跳过首列为空的行，表头会保留）。
    合并单元格会用区域左上角的值填满整区域。
    返回 (rows, row_indices)，row_indices 为每行在 Sheet 中的 1-based 行号。
    """
    if ws is None:
        return [], []
    rows = []
    row_indices = []
    if max_col is None:
        max_col = ws.max_column or 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_col=max_col, values_only=True), start=1):
        row_list = list(row) if row else []
        while len(row_list) < max_col:
            row_list.append("")
        rows.append(row_list)
        row_indices.append(row_idx)
    # 用合并区域左上角补齐空单元格
    try:
        for i in range(len(rows)):
            for col in range(max_col):
                cur = rows[i][col]
                if cur is None or (isinstance(cur, str) and cur.strip() == ""):
                    v = get_merged_cell_value(ws, row_indices[i], col + 1)
                    if v is not None:
                        rows[i][col] = v
    except Exception:
        pass
    return rows, row_indices


# -----------------------------------------------------------------------------
# 行列表 ↔ 字典 / 顺序（用于合并与对比）
# -----------------------------------------------------------------------------

def rows_to_dict(rows):
    """按第一列 Key 转为 dict[key] = row_list；空 key 跳过。"""
    d = {}
    for r in rows:
        k = cell_str(r[0]) if r else ""
        if k:
            d[k] = r
    return d


def ordered_keys(rows):
    """从 rows 按行出现顺序提取 key 列表（去重）。"""
    keys = []
    seen = set()
    for r in rows:
        k = cell_str(r[0]) if r else ""
        if k and k not in seen:
            seen.add(k)
            keys.append(k)
    return keys


def rows_to_dict_with_indices(rows, row_indices):
    """从 (rows, row_indices) 建 dict 和 key->行号；首列空的行 key 为 __row_N。"""
    d = {}
    key_to_row_index = {}
    for i, r in enumerate(rows):
        if i >= len(row_indices):
            break
        k = cell_str(r[0]) if r else ""
        if not k:
            k = "__row_%d" % row_indices[i]
        d[k] = r
        key_to_row_index[k] = row_indices[i]
    return d, key_to_row_index


def ordered_keys_from_rows_and_indices(rows, row_indices):
    """从 (rows, row_indices) 按行顺序得到 key 列表（首列空用 __row_N）。"""
    keys = []
    seen = set()
    for i, r in enumerate(rows):
        if i >= len(row_indices):
            break
        k = cell_str(r[0]) if r else ""
        if not k:
            k = "__row_%d" % row_indices[i]
        if k not in seen:
            seen.add(k)
            keys.append(k)
    return keys


def rows_to_dict_normalized(rows):
    """按首列建 dict，key 用 key_str_normalized，供冲突检测用。"""
    d = {}
    for r in rows:
        k = key_str_normalized(r[0]) if r else ""
        if k:
            d[k] = r
    return d


def ordered_keys_normalized(rows):
    """按行顺序返回规范化 key 列表（去重）。"""
    keys, seen = [], set()
    for r in rows:
        k = key_str_normalized(r[0]) if r else ""
        if k and k not in seen:
            seen.add(k)
            keys.append(k)
    return keys


# -----------------------------------------------------------------------------
# 行相等判断（用于冲突检测与合并）
# -----------------------------------------------------------------------------

def row_equal(a, b):
    """两行（list 或 tuple）按单元格字符串逐格比较是否相等。"""
    if a is b:
        return True
    if a is None or b is None:
        return a == b
    return [cell_str(c) for c in a] == [cell_str(c) for c in b]


def col_equal(col_a, col_b):
    """两列（两个 list）按单元格字符串逐格比较是否相等。"""
    if col_a is col_b:
        return True
    if col_a is None or col_b is None:
        return col_a == col_b
    return [cell_str(c) for c in col_a] == [cell_str(c) for c in col_b]


# -----------------------------------------------------------------------------
# 前缀分组（用于新增行/列插入到同前缀组末尾）
# -----------------------------------------------------------------------------

def key_prefix(key_str):
    """
    从行 key 或列 key 字符串提取前缀，用于按前缀分组。
    规则：取第一个分隔符（-、_、空格、制表）前的部分；若无分隔符则返回整串。
    """
    s = (key_str or "").strip()
    if not s:
        return ""
    for sep in ("-", "_", " ", "\t"):
        if sep in s:
            return s.split(sep)[0].strip() or s
    return s


def row_keys_grouped_by_prefix(ordered_keys):
    """
    将有序行 key 列表按前缀分组，保持每组内顺序。
    返回 [(prefix, [key1, key2, ...]), ...]，顺序为 key 首次出现的顺序。
    """
    groups = {}  # prefix -> list of keys in order
    prefix_order = []  # order of first occurrence of each prefix
    for k in ordered_keys:
        p = key_prefix(k)
        if p not in groups:
            groups[p] = []
            prefix_order.append(p)
        groups[p].append(k)
    return [(p, groups[p]) for p in prefix_order]


def insertion_index_for_new_key(merged_ordered, new_key):
    """
    计算将 new_key 插入到 merged_ordered 的位置（同前缀组末尾）。
    返回 0-based 插入位置；若基准中无该前缀则插到末尾。
    """
    p = key_prefix(new_key)
    last_idx = -1
    for i, k in enumerate(merged_ordered):
        if key_prefix(k) == p:
            last_idx = i
    return last_idx + 1


def merge_ordered_with_new_rows(base_ordered, new_ordered_keys):
    """
    在 base_ordered 中按「同前缀组末尾」规则插入 new_ordered_keys，返回新顺序。
    不修改 base_ordered；new_ordered_keys 中若 key 已在 base 中则跳过（由调用方保证只传新增 key）。
    """
    base_set = set(base_ordered)
    merged = list(base_ordered)
    for k in new_ordered_keys:
        if k in base_set:
            continue
        idx = insertion_index_for_new_key(merged, k)
        merged.insert(idx, k)
        base_set.add(k)
    return merged


# -----------------------------------------------------------------------------
# 表头与列（用于新增列插入与冲突列）
# -----------------------------------------------------------------------------

def load_sheet_header(ws, max_col=None):
    """
    加载 Sheet 第一行作为表头，返回列 key 列表（cell_str）。
    ws 为 None 或空时返回 []。
    """
    if ws is None:
        return []
    if max_col is None:
        max_col = ws.max_column or 1
    row = next(ws.iter_rows(min_row=1, max_row=1, max_col=max_col, values_only=True), None)
    if not row:
        return []
    return [cell_str(c) for c in row]


def load_sheet_header_normalized(ws, max_col=None):
    """加载表头并做 key 规范化（与 key_str_normalized 一致），返回 list。"""
    raw = load_sheet_header(ws, max_col)
    return [key_str_normalized(c) for c in raw]


def column_keys_grouped_by_prefix(ordered_col_keys):
    """
    将有序列 key 列表按前缀分组，同 row_keys_grouped_by_prefix。
    返回 [(prefix, [col_key1, ...]), ...]。
    """
    groups = {}
    prefix_order = []
    for k in ordered_col_keys:
        p = key_prefix(k)
        if p not in groups:
            groups[p] = []
            prefix_order.append(p)
        groups[p].append(k)
    return [(p, groups[p]) for p in prefix_order]


def insertion_index_for_new_col(merged_ordered_cols, new_col_key):
    """计算将新列 key 插入到 merged 列顺序的位置（同前缀组末尾）。"""
    return insertion_index_for_new_key(merged_ordered_cols, new_col_key)


def merge_ordered_with_new_cols(base_ordered_cols, new_ordered_cols):
    """在 base 列顺序中按同前缀组末尾插入新列 key，返回新顺序。"""
    return merge_ordered_with_new_rows(base_ordered_cols, new_ordered_cols)


def get_column_values(ws, col_index_1based, max_row=None):
    """
    读取 Sheet 中某一列的所有单元格值（从第 1 行到 max_row）。
    合并单元格取区域左上角值。col_index_1based 为 1-based 列号。
    """
    if ws is None:
        return []
    if max_row is None:
        max_row = ws.max_row or 1
    return [
        cell_str(get_merged_cell_value(ws, r, col_index_1based))
        for r in range(1, max_row + 1)
    ]
