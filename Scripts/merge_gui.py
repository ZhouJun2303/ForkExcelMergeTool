# -*- coding: utf-8 -*-
"""
合并窗口：A 行不变 / B 列不变 / C 删除行 / D 删除列 / E 新增 Sheet / F 删除 Sheet / G 冲突行或列。
多选框选中项参与合并逻辑，选项持久化到本地 merge_options.json。
"""

import json
import os
import re
import sys
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl

from config import MAX_TREEVIEW_ROWS
from backup_util import (
    backup_project_parent,
    create_merge_backup,
    load_saved_backup_root,
    save_backup_root,
)
from excel_io import (
    cell_str,
    get_column_values,
    header_normalize_for_compare,
    key_str_normalized,
    load_sheet_header,
    load_sheet_rows_full,
    rows_to_dict,
)
from version import __version__ as APP_VERSION
from gui_common import (
    ToolTip,
    UI,
    UpdateButtonController,
    apply_app_icon,
    configure_button_icon,
    gui_log,
    make_header_icon,
    make_badge,
    make_color_legend,
    make_icon_button,
    make_separator,
    make_update_card,
    open_excel_file,
    open_containing_folder,
    setup_merge_styles,
)
from git_util import get_git_merge_info, stage_merged_and_cleanup
from log_util import merge_options_path, release_merge_lock
from merge_core import do_merge
from preview_core import build_merge_preview


# 选项默认值（勾选=参与逻辑）
DEFAULT_OPTIONS = {"A": False, "B": False, "C": False, "D": False, "E": True, "F": False, "G": True}
MERGE_OPTIONS_SCHEMA = 2


def _load_options_data():
    try:
        path = merge_options_path()
        if path and os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return {}


def _save_options_data(data):
    try:
        path = merge_options_path()
        if path:
            parent = os.path.dirname(path)
            if parent:
                os.makedirs(parent, exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _load_merge_options():
    """从本地文件加载合并选项，不存在或异常则返回默认。"""
    try:
        data = _load_options_data()
        if data:
            opts = dict(DEFAULT_OPTIONS)
            if data.get("merge_options_schema") == MERGE_OPTIONS_SCHEMA:
                return {k: bool(data.get(k, DEFAULT_OPTIONS.get(k, False))) for k in "ABCDEFG"}
            for k in "CDEFG":
                opts[k] = bool(data.get(k, DEFAULT_OPTIONS.get(k, False)))
            return opts
    except Exception:
        pass
    return dict(DEFAULT_OPTIONS)


def _save_merge_options(opts):
    """将合并选项写入本地文件，并保留其它偏好配置。"""
    try:
        data = _load_options_data()
        for k in "ABCDEFG":
            data[k] = bool(opts.get(k, DEFAULT_OPTIONS.get(k, False)))
        data["merge_options_schema"] = MERGE_OPTIONS_SCHEMA
        _save_options_data(data)
    except Exception:
        pass


def _load_auto_open_merged():
    """从 merge_options.json 读取「合并后自动打开」是否勾选，默认 True。"""
    try:
        data = _load_options_data()
        if data:
            return bool(data.get("auto_open_merged", True))
    except Exception:
        pass
    return True


def _save_auto_open_merged(value):
    """将「合并后自动打开」写入本地，与 merge_options.json 合并。"""
    try:
        data = _load_options_data()
        data["auto_open_merged"] = bool(value)
        _save_options_data(data)
    except Exception:
        pass


# 单实例：当前合并窗口引用，关闭时清空
_merge_instance = None


def get_existing_merge_window():
    """返回当前已存在的合并窗口（若存在且未关闭），否则返回 None。"""
    global _merge_instance
    if _merge_instance is not None and _merge_instance.root.winfo_exists():
        return _merge_instance
    _merge_instance = None
    return None


class MergeWindow:
    """
    合并 GUI：7 项多选（A～G）、基准选择、根据选项展示将删除/新增/冲突列表，生成合并结果。
    """

    OPTIONS = [
        ("A", "跳过新增行"),
        ("B", "跳过新增列"),
        ("C", "删除缺失行"),
        ("D", "删除缺失列"),
        ("E", "追加新 Sheet"),
        ("F", "删除缺失 Sheet"),
        ("G", "处理冲突"),
    ]
    BASE_SIDES = [("local", "本地 (Local)"), ("remote", "线上 (Remote)")]

    def __init__(self, path_local, path_base, path_remote, path_merged, completion_strategy=None):
        global _merge_instance
        self.path_local = path_local
        self.path_base = path_base
        self.path_remote = path_remote
        self.path_merged = path_merged
        self._merged_file_path = None
        self._backup_info = None
        self._backup_dir_path = None
        self._backup_merged_path = None
        self.merge_done = False
        self.base_side_var = None
        self.status_var = None
        self.backup_root_var = None
        self.tree = None
        self.option_vars = {}
        self.conflict_entries = []
        self.update_controller = None
        self.update_progress_var = None
        self._merge_items = []
        self._merge_page = 0
        self.merge_page_var = None
        self.summary_var = None
        self.preview_status_var = None
        self.target_path_var = None
        self._preview_cache = {}
        self._preview_request_after = None
        self._preview_request_id = 0
        self._preview_loading = False
        self._merge_generating = False
        self.completion_strategy = completion_strategy
        self.root = tk.Tk()
        self.root.title("Excel 多模式合并 v%s" % APP_VERSION)
        _merge_instance = self
        self.root.minsize(980, 680)
        self.root.geometry("1180x760")
        setup_merge_styles(self.root)
        apply_app_icon(self.root)
        self.root.protocol("WM_DELETE_WINDOW", self._on_cancel)

        self.local_info, self.remote_info = get_git_merge_info(path_merged)
        self._build_ui()
        if self.update_controller:
            self.update_controller.start_background_check()
        self._schedule_preview_refresh()

    def _build_ui(self):
        pad = 14
        self.status_var = tk.StringVar(self.root, value="")
        self.auto_open_var = tk.BooleanVar(self.root, value=_load_auto_open_merged())

        bottom_bar = ttk.Frame(self.root, padding=(pad, 10), style="BottomBar.TFrame")
        bottom_bar.pack(side=tk.BOTTOM, fill=tk.X)

        shell = ttk.Frame(self.root, style="App.TFrame")
        shell.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(pad, 0))

        title_row = ttk.Frame(shell, style="App.TFrame")
        title_row.pack(fill=tk.X)
        make_header_icon(title_row, self.root).pack(side=tk.LEFT, padx=(0, 10))
        title_text = ttk.Frame(title_row, style="App.TFrame")
        title_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        headline = ttk.Frame(title_text, style="App.TFrame")
        headline.pack(fill=tk.X)
        ttk.Label(headline, text="Excel 三向合并", style="Title.TLabel").pack(side=tk.LEFT)
        make_badge(headline, "v%s" % APP_VERSION, "primary").pack(side=tk.LEFT, padx=(10, 0))
        ttk.Label(
            title_text,
            text="预览 BASE / LOCAL / REMOTE 差异，选择规则后生成 Fork 可接收的合并结果。",
            style="Subtitle.TLabel",
        ).pack(anchor=tk.W, pady=(2, 0))
        self.preview_status_var = tk.StringVar(self.root, value="")
        ttk.Label(title_row, textvariable=self.preview_status_var, style="TLabel").pack(side=tk.RIGHT)

        target_frame = ttk.Frame(shell, padding=(12, 9), style="Panel.TFrame")
        target_frame.pack(fill=tk.X, pady=(10, 10))
        self.target_path_var = tk.StringVar(self.root, value=self._short_path(self.path_merged, 112))
        ttk.Label(target_frame, text="目标文件", style="Section.TLabel").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(target_frame, textvariable=self.target_path_var, style="Panel.TLabel").pack(side=tk.LEFT, fill=tk.X, expand=True)

        body = ttk.PanedWindow(shell, orient=tk.HORIZONTAL)
        body.pack(fill=tk.BOTH, expand=True)

        left_outer = ttk.Frame(body, style="Panel.TFrame")
        right_panel = ttk.Frame(body, padding=(12, 12), style="Panel.TFrame")
        body.add(left_outer, weight=0)
        body.add(right_panel, weight=1)

        left_canvas = tk.Canvas(left_outer, width=282, bg=UI["panel"], highlightthickness=0, borderwidth=0)
        left_scroll = ttk.Scrollbar(left_outer, orient=tk.VERTICAL, command=left_canvas.yview)
        left_panel = ttk.Frame(left_canvas, padding=(12, 12), style="Panel.TFrame")
        left_window = left_canvas.create_window((0, 0), window=left_panel, anchor=tk.NW)
        left_canvas.configure(yscrollcommand=left_scroll.set)
        left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        left_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        def _sync_left_scroll(_event=None):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))
            left_canvas.itemconfigure(left_window, width=left_canvas.winfo_width())

        def _left_scrollable():
            bbox = left_canvas.bbox("all")
            return bool(bbox and bbox[3] > left_canvas.winfo_height())

        def _pointer_over_left_canvas(event):
            try:
                x_root = getattr(event, "x_root", None)
                y_root = getattr(event, "y_root", None)
                if x_root is None or y_root is None:
                    x_root = self.root.winfo_pointerx()
                    y_root = self.root.winfo_pointery()
                left = left_canvas.winfo_rootx()
                top = left_canvas.winfo_rooty()
                return (
                    left <= x_root < left + left_canvas.winfo_width()
                    and top <= y_root < top + left_canvas.winfo_height()
                )
            except tk.TclError:
                return False

        def _on_left_mousewheel(event):
            if not _pointer_over_left_canvas(event) or not _left_scrollable():
                return None
            if getattr(event, "num", None) == 4:
                units = -3
            elif getattr(event, "num", None) == 5:
                units = 3
            else:
                delta = getattr(event, "delta", 0)
                units = -int(delta / 120) if delta else 0
                if units == 0 and delta:
                    units = -1 if delta > 0 else 1
            if units:
                left_canvas.yview_scroll(units, "units")
                return "break"
            return None

        left_panel.bind("<Configure>", _sync_left_scroll)
        left_canvas.bind("<Configure>", _sync_left_scroll)
        self.root.bind_all("<MouseWheel>", _on_left_mousewheel, add="+")
        self.root.bind_all("<Button-4>", _on_left_mousewheel, add="+")
        self.root.bind_all("<Button-5>", _on_left_mousewheel, add="+")

        rules_frame = ttk.Frame(left_panel, style="Panel.TFrame")
        rules_frame.pack(fill=tk.X)
        ttk.Label(rules_frame, text="合并规则", style="Section.TLabel").pack(anchor=tk.W)
        base_row = ttk.Frame(rules_frame, style="Panel.TFrame")
        base_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(base_row, text="基准侧", style="Panel.TLabel").pack(side=tk.LEFT, padx=(0, 8))
        self.base_side_var = tk.StringVar(self.root, value="local")
        base_cb = ttk.Combobox(base_row, textvariable=self.base_side_var, state="readonly", width=16)
        base_cb["values"] = [b[1] for b in self.BASE_SIDES]
        base_cb.current(0)
        base_cb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        base_cb.bind("<<ComboboxSelected>>", lambda e: self._schedule_preview_refresh())
        ToolTip(base_cb, "基准侧只决定格式来源和默认冲突选择；非冲突单侧修改会按 BASE 自动合并。")

        opts_row = ttk.Frame(rules_frame, style="Panel.TFrame")
        opts_row.pack(fill=tk.X, pady=(10, 0))
        loaded = _load_merge_options()
        self.option_vars = {}
        for key, label in self.OPTIONS:
            var = tk.BooleanVar(self.root, value=loaded.get(key, DEFAULT_OPTIONS.get(key, False)))
            self.option_vars[key] = var
            cb = ttk.Checkbutton(
                opts_row, text="%s %s" % (key, label), variable=var, command=self._on_option_click,
            )
            cb.pack(anchor=tk.W, pady=1)
        ToolTip(opts_row, "默认会合入另一侧新增行/列；A/B 可跳过新增结构，C/D/F 控制删除，G 控制人工冲突选择。")

        backup_frame = ttk.Frame(left_panel, style="Panel.TFrame")
        self.backup_root_var = tk.StringVar(self.root, value=load_saved_backup_root())

        make_separator(left_panel).pack(fill=tk.X, pady=12)

        info_frame = ttk.Frame(left_panel, padding=(8, 8), style="Card.TFrame")
        info_frame.pack(fill=tk.X)
        ttk.Label(info_frame, text="版本来源", style="CardSection.TLabel").pack(anchor=tk.W)
        left_box = ttk.Frame(info_frame, padding=(0, 6), style="Card.TFrame")
        left_box.pack(fill=tk.X, pady=(6, 4))
        make_badge(left_box, "本地 Local", "neutral").pack(anchor=tk.W)
        self._fill_commit_info(left_box, self.local_info)
        make_icon_button(left_box, self.root, "打开本地 Excel", "open", command=lambda: self._open_local(), style="Tiny.TButton").pack(anchor=tk.W, pady=(4, 0))
        make_separator(info_frame).pack(fill=tk.X, pady=(0, 4))
        right_box = ttk.Frame(info_frame, padding=(0, 6), style="Card.TFrame")
        right_box.pack(fill=tk.X)
        make_badge(right_box, "线上 Remote", "neutral").pack(anchor=tk.W)
        self._fill_commit_info(right_box, self.remote_info)
        make_icon_button(right_box, self.root, "打开线上 Excel", "open", command=lambda: self._open_remote(), style="Tiny.TButton").pack(anchor=tk.W, pady=(4, 0))

        make_separator(left_panel).pack(fill=tk.X, pady=12)

        backup_frame.pack(fill=tk.X)
        ttk.Label(backup_frame, text="备份", style="Section.TLabel").pack(anchor=tk.W)
        backup_row = ttk.Frame(backup_frame, style="Panel.TFrame")
        backup_row.pack(fill=tk.X, pady=(8, 0))
        backup_entry = ttk.Entry(backup_row, textvariable=self.backup_root_var)
        backup_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        btn_backup_pick = make_icon_button(backup_row, self.root, "选择", "folder", command=self._on_choose_backup_root, style="Secondary.TButton")
        btn_backup_pick.pack(side=tk.LEFT)
        make_icon_button(backup_frame, self.root, "保存备份设置", "backup", command=self._on_save_backup_root, style="Secondary.TButton").pack(anchor=tk.W, pady=(6, 0))
        ToolTip(backup_entry, "留空时使用合并文件同目录下的 MergeExcelBackup。")

        make_separator(left_panel).pack(fill=tk.X, pady=12)
        (
            update_card,
            self.btn_update,
            self.update_state_var,
            self.update_state_label,
            self.update_state_icon,
            update_progress_row,
            self.update_progress_var,
            self.update_progress_label,
            self.update_progress_bar,
        ) = make_update_card(left_panel, self.root)
        update_card.pack(fill=tk.X)
        ToolTip(self.btn_update, "检查 GitHub Release 是否有新版本；exe 运行模式支持自动下载替换。")

        preview_top = ttk.Frame(right_panel, style="Panel.TFrame")
        preview_top.pack(fill=tk.X)
        ttk.Label(preview_top, text="差异预览", style="Section.TLabel").pack(side=tk.LEFT)
        self.summary_var = tk.StringVar(self.root, value="")
        self.hint_label = ttk.Label(preview_top, textvariable=self.summary_var, style="Muted.TLabel")
        self.hint_label.pack(side=tk.LEFT, anchor=tk.W, padx=(12, 0))
        legend_merge = make_color_legend(preview_top, [
            (UI["success_bg"], "新增"),
            (UI["warning_bg"], "删除冲突"),
            (UI["danger_bg"], "修改冲突"),
            (UI["deleted_bg"], "将删除"),
        ])
        legend_merge.pack(side=tk.RIGHT)

        self.content_frame = ttk.Frame(right_panel, style="Panel.TFrame")
        self.content_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        table_frame = ttk.Frame(self.content_frame, style="Panel.TFrame")
        table_frame.pack(fill=tk.BOTH, expand=True)
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        cols = ("Sheet", "Key / 说明", "处理方式")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=14)
        for c, w, min_w, stretch in (
            ("Sheet", 150, 110, False),
            ("Key / 说明", 420, 260, True),
            ("处理方式", 190, 150, False),
        ):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, minwidth=min_w, stretch=stretch)
        self.tree.tag_configure("new", foreground=UI["success"], background=UI["success_bg"])
        self.tree.tag_configure("del", foreground=UI["deleted"], background=UI["deleted_bg"])
        self.tree.tag_configure("mod", foreground=UI["warning"], background=UI["warning_bg"])
        self.tree.tag_configure("del_conflict", foreground=UI["warning"], background=UI["warning_bg"])
        self.tree.tag_configure("conflict", foreground=UI["danger"], background=UI["danger_bg"])
        self.tree.configure(cursor="hand2")
        self.tree.bind("<Double-1>", self._on_merge_tree_double_click)
        self.tree.bind("<Return>", self._on_merge_tree_double_click)
        sb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        page_frame = ttk.Frame(self.content_frame, style="Panel.TFrame")
        page_frame.pack(fill=tk.X, pady=(8, 0))
        self.merge_page_var = tk.StringVar(self.root, value="")
        make_icon_button(page_frame, self.root, "上一页", "prev", command=lambda: self._change_merge_page(-1), style="Secondary.TButton").pack(side=tk.LEFT, padx=(0, 4))
        make_icon_button(page_frame, self.root, "下一页", "next", command=lambda: self._change_merge_page(1), style="Secondary.TButton").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(page_frame, textvariable=self.merge_page_var, style="Panel.TLabel").pack(side=tk.LEFT)
        sel_frame = ttk.Frame(self.content_frame, style="Panel.TFrame")
        sel_frame.pack(fill=tk.X, pady=(8, 0))
        ttk.Label(sel_frame, text="冲突选择", style="Panel.TLabel").pack(side=tk.LEFT, padx=(0, 8))
        make_icon_button(sel_frame, self.root, "取本地", "check", command=lambda: self._set_choice("本地"), style="Secondary.TButton").pack(side=tk.LEFT, padx=(0, 6))
        make_icon_button(sel_frame, self.root, "取线上", "check", command=lambda: self._set_choice("线上"), style="Secondary.TButton").pack(side=tk.LEFT)
        make_icon_button(sel_frame, self.root, "查看详情", "detail", command=self._show_selected_merge_detail, style="Secondary.TButton").pack(side=tk.RIGHT)
        ttk.Label(sel_frame, text="双击或回车查看完整内容", style="Muted.TLabel").pack(side=tk.RIGHT, padx=(0, 10))

        status_row = ttk.Frame(bottom_bar, style="BottomBar.TFrame")
        status_row.pack(fill=tk.X)
        ttk.Label(status_row, textvariable=self.status_var, style="Panel.TLabel").pack(side=tk.LEFT, anchor=tk.W)
        btn_row = ttk.Frame(bottom_bar, style="BottomBar.TFrame")
        btn_row.pack(fill=tk.X, pady=(8, 0))
        self.btn_merge = ttk.Button(btn_row, text="生成合并结果", command=self._on_generate_merge, style="Accent.TButton")
        configure_button_icon(self.root, self.btn_merge, "merge")
        self.btn_merge.pack(side=tk.LEFT, padx=(0, 8))
        self.btn_confirm = ttk.Button(btn_row, text="确认无误并解决冲突", command=self._on_confirm_done)
        configure_button_icon(self.root, self.btn_confirm, "check")
        self.btn_confirm.pack(side=tk.LEFT, padx=(0, 8))
        self.btn_confirm.config(state=tk.DISABLED)
        self.btn_open_merged = ttk.Button(btn_row, text="打开合并结果", command=self._on_open_merged)
        configure_button_icon(self.root, self.btn_open_merged, "open")
        self.btn_open_merged.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Checkbutton(
            btn_row, text="合并后自动打开", variable=self.auto_open_var,
            command=lambda: _save_auto_open_merged(self.auto_open_var.get()),
        ).pack(side=tk.LEFT, padx=(0, 12))
        make_icon_button(btn_row, self.root, "取消", "cancel", command=self._on_cancel, style="Secondary.TButton").pack(side=tk.LEFT)
        self.update_controller = UpdateButtonController(
            self.root, self.btn_update, status_var=self.status_var, on_quit=self._quit_for_update, compact=True,
        )
        self.update_controller.bind_state_widget(self.update_state_var, self.update_state_label, self.update_state_icon)

        backup_btn_row = ttk.Frame(bottom_bar, style="BottomBar.TFrame")
        backup_btn_row.pack(fill=tk.X, pady=(6, 0))
        self.btn_open_backup_merged = ttk.Button(backup_btn_row, text="打开备份文件", command=self._on_open_backup_merged)
        configure_button_icon(self.root, self.btn_open_backup_merged, "open")
        self.btn_open_backup_merged.pack(side=tk.LEFT, padx=(0, 8))
        self.btn_manual_backup = ttk.Button(backup_btn_row, text="手动保存备份", command=self._on_manual_save_backup)
        configure_button_icon(self.root, self.btn_manual_backup, "backup")
        self.btn_manual_backup.pack(side=tk.LEFT, padx=(0, 8))
        self.btn_open_backup_dir = ttk.Button(backup_btn_row, text="打开备份目录", command=self._on_open_backup_dir)
        configure_button_icon(self.root, self.btn_open_backup_dir, "folder")
        self.btn_open_backup_dir.pack(side=tk.LEFT)
        self.update_controller.bind_progress_widgets(
            self.update_progress_bar, self.update_progress_var, self.update_progress_label, update_progress_row,
        )

    def _get_options(self):
        """返回当前勾选的选项集合 {"A","B",...}。"""
        return {k for k, v in self.option_vars.items() if v.get()}

    def _get_base_side(self):
        raw = self.base_side_var.get()
        for sid, label in self.BASE_SIDES:
            if label == raw:
                return sid
        return "local"

    def _short_path(self, path, max_len=96):
        path = os.path.normpath(path or "")
        if len(path) <= max_len:
            return path
        keep = max(12, max_len - 1)
        return "…" + path[-keep:]

    def _on_option_click(self):
        opts = {k: self.option_vars[k].get() for k in self.option_vars}
        _save_merge_options(opts)
        self._schedule_preview_refresh()

    def _on_options_or_base_changed(self):
        self._schedule_preview_refresh()

    def _preview_cache_key(self, options, base_side):
        try:
            file_sig = tuple(
                (os.path.abspath(p), os.path.getmtime(p), os.path.getsize(p))
                for p in (self.path_local, self.path_base, self.path_remote)
            )
        except OSError:
            file_sig = tuple(os.path.abspath(p) for p in (self.path_local, self.path_base, self.path_remote))
        return (file_sig, base_side, tuple(sorted(options)))

    def _set_preview_busy(self, busy, text=None):
        self._preview_loading = busy
        state = tk.DISABLED if busy else tk.NORMAL
        if hasattr(self, "btn_merge") and self.btn_merge is not None and not self._merge_generating:
            self.btn_merge.config(state=state)
        if self.preview_status_var is not None:
            self.preview_status_var.set(text or ("正在计算预览..." if busy else ""))

    def _set_merge_generation_busy(self, busy, text=None):
        self._merge_generating = busy
        merge_state = tk.DISABLED if busy or self._preview_loading else tk.NORMAL
        if hasattr(self, "btn_merge") and self.btn_merge is not None:
            self.btn_merge.config(state=merge_state)
        if hasattr(self, "btn_confirm") and self.btn_confirm is not None:
            self.btn_confirm.config(state=tk.DISABLED if busy else (tk.NORMAL if self.merge_done else tk.DISABLED))
        if self.preview_status_var is not None:
            self.preview_status_var.set(text or ("正在生成合并结果..." if busy else ""))

    def _schedule_preview_refresh(self, delay_ms=160):
        if self._preview_request_after is not None:
            try:
                self.root.after_cancel(self._preview_request_after)
            except Exception:
                pass
        self._set_preview_busy(True, "预览待更新...")
        if self.summary_var is not None:
            self.summary_var.set("合并规则已变化，正在准备刷新预览...")
        self._preview_request_after = self.root.after(delay_ms, self._refresh_preview_async)

    def _refresh_preview_async(self):
        self._preview_request_after = None
        options = self._get_options()
        base_side = self._get_base_side()
        cache_key = self._preview_cache_key(options, base_side)
        cached = self._preview_cache.get(cache_key)
        if cached is not None:
            self._apply_preview_result(cached, from_cache=True)
            return

        self._preview_request_id += 1
        request_id = self._preview_request_id
        self._set_preview_busy(True)
        self.tree.delete(*self.tree.get_children())
        self.conflict_entries = []
        self._merge_items = []
        if self.summary_var is not None:
            self.summary_var.set("正在计算差异预览...")

        def worker():
            try:
                result = build_merge_preview(
                    self.path_local, self.path_base, self.path_remote, options, base_side
                )
            except Exception as e:
                result = {"error": e}
            self.root.after(0, lambda: self._finish_preview_refresh(request_id, cache_key, result))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_preview_refresh(self, request_id, cache_key, result):
        if request_id != self._preview_request_id:
            return
        if result.get("error"):
            err = result["error"]
            self._set_preview_busy(False, "预览加载失败")
            gui_log("加载数据失败: " + str(err), self.status_var, is_error=True)
            if self.summary_var is not None:
                self.summary_var.set("加载失败: " + str(err) + " 详见日志。")
            return
        self._preview_cache[cache_key] = result
        if len(self._preview_cache) > 6:
            for old_key in list(self._preview_cache.keys())[:-6]:
                self._preview_cache.pop(old_key, None)
        self._apply_preview_result(result, from_cache=False)

    def _apply_preview_result(self, result, from_cache=False):
        self.conflict_entries = list(result.get("conflict_entries") or [])
        self._merge_items = self._merge_items_with_current_choices(result.get("items") or [])
        self._merge_page = 0
        shown = self._refresh_merge_tree_page()
        total = len(self._merge_items)
        base_label = "本地" if result.get("base_side") == "local" else "线上"
        summary = result.get("summary") or {}
        pieces = [
            "基准=%s" % base_label,
            "新增 %d" % summary.get("new", 0),
            "删除 %d" % summary.get("delete", 0),
            "冲突 %d" % summary.get("conflict", 0),
            "信息 %d" % summary.get("info", 0),
            "合计 %d" % total,
        ]
        if total > MAX_TREEVIEW_ROWS:
            pieces.append("每页 %d 条" % MAX_TREEVIEW_ROWS)
        if self.summary_var is not None:
            self.summary_var.set("；".join(pieces))
        source = "缓存" if from_cache else "计算"
        self._set_preview_busy(False, "预览已更新")
        elapsed = result.get("elapsed_ms")
        elapsed_text = "，耗时 %dms" % elapsed if elapsed is not None else ""
        gui_log("已%s合并预览，共 %d 条，显示 %d 条%s" % (source, total, shown, elapsed_text), self.status_var)

    def _merge_items_with_current_choices(self, items):
        out = []
        for item in items:
            if len(item) < 4:
                out.append(item)
                continue
            tag = item[3]
            tags = tuple(tag) if isinstance(tag, (tuple, list)) else (tag,)
            if tags:
                try:
                    idx = int(tags[0])
                except (TypeError, ValueError):
                    idx = None
                if idx is not None and 0 <= idx < len(self.conflict_entries):
                    choice = self.conflict_entries[idx].get("choice", "本地")
                    display = self.conflict_entries[idx].get("display")
                    if not display:
                        display = "将保留本地" if choice == "本地" else "将保留线上"
                    out.append((item[0], item[1], display, item[3]))
                    continue
            out.append(item)
        return out

    def _refresh_merge_tree_page(self):
        """只渲染当前页，避免 Treeview 在大文件下卡顿。"""
        self.tree.delete(*self.tree.get_children())
        total = len(self._merge_items)
        if total <= 0:
            if self.merge_page_var is not None:
                self.merge_page_var.set("")
            return 0
        page_size = MAX_TREEVIEW_ROWS
        max_page = max((total - 1) // page_size, 0)
        if self._merge_page < 0:
            self._merge_page = 0
        if self._merge_page > max_page:
            self._merge_page = max_page
        start = self._merge_page * page_size
        end = min(start + page_size, total)
        for sheet_name, key, choice, tag in self._merge_items[start:end]:
            tags = tuple(tag) if isinstance(tag, (tuple, list)) else (tag,)
            self.tree.insert("", tk.END, values=(sheet_name, key, choice), tags=tags)
        if self.merge_page_var is not None:
            self.merge_page_var.set("第 %d/%d 页，显示 %d-%d / %d" % (
                self._merge_page + 1, max_page + 1, start + 1, end, total))
        return end - start

    def _change_merge_page(self, delta):
        """切换合并列表页。"""
        if not self._merge_items:
            return
        old_page = self._merge_page
        self._merge_page += delta
        shown = self._refresh_merge_tree_page()
        if self._merge_page != old_page:
            gui_log("合并列表翻页：第 %d 页，显示 %d 条" % (self._merge_page + 1, shown), self.status_var)

    def _fill_commit_info(self, parent, info):
        fields = []
        if info and isinstance(info, dict):
            if info.get("short_hash"):
                fields.append(("Hash", info["short_hash"]))
            if info.get("author"):
                fields.append(("提交人", info["author"]))
            if info.get("message"):
                msg = info["message"]
                fields.append(("事件", msg[:34] + "..." if len(msg) > 34 else msg))
            if info.get("date"):
                fields.append(("时间", info["date"][:16] if len(info["date"]) >= 16 else info["date"]))
        if not fields:
            ttk.Label(parent, text="无法获取 Git 信息", style="CardMuted.TLabel").pack(anchor=tk.W, pady=(6, 0))
        else:
            for label, value in fields:
                row = ttk.Frame(parent, style="Card.TFrame")
                row.pack(fill=tk.X, pady=(3, 0))
                ttk.Label(row, text=label, style="CardMuted.TLabel", width=6).pack(side=tk.LEFT)
                ttk.Label(row, text=value, style="Card.TLabel").pack(side=tk.LEFT, fill=tk.X, expand=True)

    def _show_selected_merge_detail(self):
        if not self.tree.selection():
            messagebox.showinfo("提示", "请先选择一条差异预览记录。")
            return
        self._on_merge_tree_double_click(None)

    def _open_local(self):
        if open_excel_file(self.path_local):
            gui_log("已打开本地 Excel", self.status_var)
        else:
            messagebox.showwarning("提示", "文件不存在或无法打开")

    def _open_remote(self):
        if open_excel_file(self.path_remote):
            gui_log("已打开线上 Excel", self.status_var)
        else:
            messagebox.showwarning("提示", "文件不存在或无法打开")

    def _current_backup_root(self):
        return (self.backup_root_var.get() if self.backup_root_var is not None else "").strip()

    def _on_choose_backup_root(self):
        initial_dir = self._current_backup_root()
        if not initial_dir or not os.path.isdir(initial_dir):
            initial_dir = os.path.dirname(os.path.abspath(self.path_merged)) or os.getcwd()
        folder = filedialog.askdirectory(
            parent=self.root,
            title="选择备份根目录",
            initialdir=initial_dir,
        )
        if folder:
            self.backup_root_var.set(os.path.normpath(folder))
            self._on_save_backup_root(show_message=False)

    def _on_save_backup_root(self, show_message=True):
        root_dir = self._current_backup_root()
        try:
            save_backup_root(root_dir)
            target = backup_project_parent(self.path_merged, root_dir)
            msg = "备份目录设置已保存：%s" % target
            gui_log(msg, self.status_var)
            if show_message:
                messagebox.showinfo("已保存", msg)
        except Exception as e:
            gui_log("保存备份目录设置失败: %s" % e, self.status_var, is_error=True)
            messagebox.showerror("错误", "保存备份目录设置失败：%s" % e)

    def _on_open_merged(self):
        if not self.merge_done:
            messagebox.showwarning("提示", "请先点击「生成合并结果」")
            return
        path = self._merged_file_path or os.path.normpath(os.path.abspath(self.path_merged))
        if not os.path.isfile(path):
            messagebox.showwarning("提示", "合并文件不存在：%s" % path)
            return
        if open_excel_file(path):
            gui_log("已打开合并结果：%s" % path, self.status_var)
        else:
            messagebox.showwarning("提示", "无法打开合并文件")

    def _on_open_backup_merged(self):
        if not self.merge_done or not self._backup_merged_path:
            messagebox.showwarning("提示", "请先点击「生成合并结果」")
            return
        path = os.path.normpath(os.path.abspath(self._backup_merged_path))
        if not os.path.isfile(path):
            messagebox.showwarning("提示", "备份的合并文件不存在：%s" % path)
            return
        if open_excel_file(path):
            gui_log("已打开备份的合并文件：%s" % path, self.status_var)
        else:
            messagebox.showwarning("提示", "无法打开备份文件")

    def _set_backup_info(self, backup_info):
        self._backup_info = backup_info or None
        self._backup_dir_path = (backup_info or {}).get("dir")
        self._backup_merged_path = (backup_info or {}).get("merged")

    def _on_manual_save_backup(self):
        if not self.merge_done:
            messagebox.showwarning("提示", "请先点击「生成合并结果」")
            return
        merged_path = self._merged_file_path or os.path.normpath(os.path.abspath(self.path_merged))
        if not os.path.isfile(merged_path):
            messagebox.showwarning("提示", "合并文件不存在：%s" % merged_path)
            return
        try:
            backup_info = create_merge_backup(
                self.path_local,
                self.path_remote,
                merged_path,
                backup_root=self._current_backup_root(),
            )
            self._set_backup_info(backup_info)
            gui_log("已手动保存备份：%s" % backup_info["dir"], self.status_var)
            messagebox.showinfo("已保存", "备份已保存到：\n%s" % backup_info["dir"])
        except Exception as e:
            gui_log("手动保存备份失败: %s" % e, self.status_var, is_error=True)
            messagebox.showerror("错误", "手动保存备份失败：%s" % e)

    def _on_open_backup_dir(self):
        path = self._backup_dir_path
        if not path:
            root_dir = self._current_backup_root()
            path = backup_project_parent(self.path_merged, root_dir)
        if not os.path.isdir(path):
            try:
                os.makedirs(path, exist_ok=True)
            except Exception:
                messagebox.showwarning("提示", "备份目录不存在：%s" % path)
                return
        if open_containing_folder(path, select_file=False):
            gui_log("已打开备份目录：%s" % path, self.status_var)
        else:
            messagebox.showwarning("提示", "无法打开备份目录")

    def _on_generate_merge(self):
        if self._merge_generating:
            return
        path_out = os.path.normpath(os.path.abspath(self.path_merged))
        if os.path.isfile(path_out):
            try:
                with open(path_out, "ab") as _:
                    pass
            except PermissionError:
                messagebox.showerror(
                    "无法写入合并结果",
                    "合并结果文件正在被占用（如 Excel 已打开），请先关闭后再点击「生成合并结果」。"
                )
                return
        options = self._get_options()
        gui_log("生成合并结果，勾选项: %s（非冲突单侧修改会自动合入）" % (", ".join(sorted(options)) or "无"), self.status_var)
        base_side = self._get_base_side()
        d_choices = []
        if "G" in options:
            for entry in self.conflict_entries:
                obj = entry["data"]
                kind = entry.get("kind", "row")
                choice = "local" if entry.get("choice") == "本地" else "remote"
                d_choices.append({
                    "sheet": obj["sheet"],
                    "key": obj["key"],
                    "choice": choice,
                    "kind": "row" if kind == "row" else "column",
                })
        try:
            self._on_save_backup_root(show_message=False)
        except Exception as e:
            gui_log("保存合并配置失败: " + str(e), self.status_var, is_error=True)
            messagebox.showerror("错误", str(e))
            return

        backup_root = self._current_backup_root()
        backup_context_path = getattr(self.completion_strategy, "context_path", None)
        self._set_merge_generation_busy(True, "正在后台生成合并结果...")
        gui_log("后台生成合并结果已开始。", self.status_var)

        def worker():
            try:
                code = do_merge(
                    self.path_local, self.path_base, self.path_remote, self.path_merged,
                    base_side=base_side, d_choices=d_choices if d_choices else None,
                    options=options, backup_root=backup_root,
                    backup_context_path=backup_context_path,
                )
                if code != 0:
                    raise RuntimeError("合并返回码 %d" % code)
                result = {
                    "merged_file": os.path.normpath(os.path.abspath(self.path_merged)),
                    "backup_info": getattr(do_merge, "last_backup_info", None),
                }
            except Exception as e:
                import traceback
                result = {"error": e, "traceback": traceback.format_exc()}
            try:
                self.root.after(0, lambda: self._finish_generate_merge(result))
            except Exception:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def _finish_generate_merge(self, result):
        self._set_merge_generation_busy(False, "合并结果已生成" if not result.get("error") else "合并失败")
        if result.get("error"):
            err = result["error"]
            tb = result.get("traceback", "")
            if isinstance(err, PermissionError):
                gui_log("合并失败（文件可能被占用）: " + str(err), self.status_var, is_error=True)
                messagebox.showerror(
                    "无法写入合并结果",
                    "合并结果文件可能正在被 Excel 打开，请先关闭该文件后再点击「生成合并结果」。"
                )
            else:
                gui_log("合并失败: " + str(err), self.status_var, is_error=True)
                messagebox.showerror("错误", str(err) + "\n" + tb)
            return
        self.merge_done = True
        self._merged_file_path = result["merged_file"]
        self._set_backup_info(result.get("backup_info"))
        gui_log("合并结果已生成：%s；备份=%s" % (self._merged_file_path, self._backup_dir_path or ""), self.status_var)
        _save_auto_open_merged(self.auto_open_var.get())
        self.btn_confirm.config(state=tk.NORMAL)
        if self.auto_open_var.get() and os.path.isfile(self._merged_file_path):
            open_excel_file(self._merged_file_path)

    def activate_and_refresh(self, path_local, path_base, path_remote, path_merged):
        """单实例复用：用新路径刷新列表并置前。"""
        self.path_local = path_local
        self.path_base = path_base
        self.path_remote = path_remote
        self.path_merged = path_merged
        self._merged_file_path = None
        self._set_backup_info(None)
        self.merge_done = False
        self.btn_confirm.config(state=tk.DISABLED)
        if self.target_path_var is not None:
            self.target_path_var.set(self._short_path(self.path_merged, 112))
        self._preview_cache.clear()
        self._schedule_preview_refresh(delay_ms=0)
        self.root.lift()
        self.root.focus_force()

    def _on_confirm_done(self):
        global _merge_instance
        if not self.merge_done:
            messagebox.showwarning("提示", "请先点击「生成合并结果」")
            return
        if self.completion_strategy is not None:
            try:
                result = self.completion_strategy.complete(self)
            except Exception as e:
                gui_log("确认失败: " + str(e), self.status_var, is_error=True)
                messagebox.showerror("确认失败", str(e))
                return
            if not getattr(result, "success", False):
                msg = "\n".join(getattr(result, "errors", None) or ["确认失败，Git 仍保持冲突状态。"])
                gui_log(msg, self.status_var, is_error=True)
                messagebox.showerror("确认失败", msg)
                return
            messagebox.showinfo("完成", getattr(result, "message", "合并结果已确认。"))
            self._close_after_success()
            return
        def _log_cb(msg, is_err=False):
            gui_log(msg, self.status_var, is_error=is_err)
        result = stage_merged_and_cleanup(
            self.path_merged, self.path_local, self.path_base, self.path_remote, log_callback=_log_cb,
        )
        if not result.success:
            msg = "\n".join(result.errors or ["确认失败：git add 未成功，已保留窗口和合并结果文件。"])
            messagebox.showerror("确认失败", msg)
            return
        messagebox.showinfo("完成", "冲突已解决：已 git add。Fork 将使用合并后的文件。")
        self._close_after_success()

    def _close_after_success(self):
        global _merge_instance
        release_merge_lock()
        if _merge_instance is self:
            _merge_instance = None
        self.root.quit()
        self.root.destroy()
        sys.exit(0)

    def _set_choice(self, choice):
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]
        tags = self.tree.item(item, "tags")
        if not tags:
            return
        try:
            idx = int(tags[0])
        except ValueError:
            return
        if "G" not in self._get_options() or idx < 0 or idx >= len(self.conflict_entries):
            return
        self.conflict_entries[idx]["choice"] = choice
        display = "将保留本地" if choice == "本地" else "将保留线上"
        self.conflict_entries[idx]["display"] = display
        vals = list(self.tree.item(item, "values"))
        vals[2] = display
        self.tree.item(item, values=vals)
        self._sync_merge_item_choice(idx, display)

    def _sync_merge_item_choice(self, conflict_idx, display):
        for i, item in enumerate(self._merge_items):
            tag = item[3] if len(item) > 3 else ()
            tags = tuple(tag) if isinstance(tag, (tuple, list)) else (tag,)
            if tags and tags[0] == str(conflict_idx):
                self._merge_items[i] = (item[0], item[1], display, item[3])
                break

    def _on_merge_tree_double_click(self, event):
        """双击列表行：打开详情面板，显示 BASE | LOCAL | REMOTE 三列对比。"""
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]
        vals = self.tree.item(item, "values")
        tags = self.tree.item(item, "tags") or ()
        if len(vals) < 3:
            return
        sheet_or_name, key_or_col, choice = vals[0], vals[1], vals[2]
        
        # 处理冲突项（带索引tag）
        if tags and tags[0] not in ("new", "del", "mod", "del_conflict", "conflict"):
            try:
                idx = int(tags[0])
                if 0 <= idx < len(self.conflict_entries):
                    entry = self.conflict_entries[idx]
                    c = entry["data"]
                    kind = entry.get("kind", "row")
                    conflict_type = c.get("type", "modify_conflict")
                    
                    # 提取三方数据
                    base_vals = [cell_str(x) for x in (c.get("base_row") or [])]
                    local_vals = [cell_str(x) for x in (c.get("local_row") or [])]
                    remote_vals = [cell_str(x) for x in (c.get("remote_row") or [])]
                    
                    # 根据冲突类型设置标题
                    type_names = {
                        "add_conflict": "新增冲突",
                        "delete_conflict_local": "删除冲突（本地删除）",
                        "delete_conflict_remote": "删除冲突（线上删除）",
                        "modify_conflict": "修改冲突",
                    }
                    type_name = type_names.get(conflict_type, "冲突")
                    title = "%s — %s / %s" % (type_name, c["sheet"], c["key"])
                    
                    self._show_merge_detail_panel_three(title, base_vals, local_vals, remote_vals, conflict_type)
                    return
            except (ValueError, IndexError):
                pass
        
        # 处理非冲突项（新增/删除）
        base_side = self._get_base_side()
        action = str(choice).strip("（）")
        if action == "将新增行":
            path_base = self.path_local if base_side == "local" else self.path_remote
            path_other = self.path_remote if base_side == "local" else self.path_local
            left_vals, right_vals = self._load_row_from_workbooks(sheet_or_name, key_or_col, path_base, path_other)
            self._show_merge_detail_panel("将新增行 — %s / %s" % (sheet_or_name, key_or_col), "基准(无)", "另一方", left_vals, right_vals)
        elif action == "将删除行":
            path_base = self.path_local if base_side == "local" else self.path_remote
            path_other = self.path_remote if base_side == "local" else self.path_local
            left_vals, right_vals = self._load_row_from_workbooks(sheet_or_name, key_or_col, path_base, path_other)
            self._show_merge_detail_panel("将删除行 — %s / %s" % (sheet_or_name, key_or_col), "基准", "另一方(无)", left_vals, right_vals)
        elif action == "将新增列":
            left_vals, right_vals = self._load_col_from_workbooks(
                sheet_or_name, key_or_col, base_side, has_in_other=True
            )
            self._show_merge_detail_panel("将新增列 — %s / %s" % (sheet_or_name, key_or_col), "基准(无)", "另一方", left_vals, right_vals)
        elif action == "将删除列":
            left_vals, right_vals = self._load_col_from_workbooks(
                sheet_or_name, key_or_col, base_side, has_in_other=False
            )
            self._show_merge_detail_panel("将删除列 — %s / %s" % (sheet_or_name, key_or_col), "基准", "另一方(无)", left_vals, right_vals)
        elif "新增 Sheet" in key_or_col or "删除 Sheet" in key_or_col or "将追加" in action or "将删除" in action:
            messagebox.showinfo("详情", "Sheet: %s\n%s" % (sheet_or_name, choice))
        else:
            gui_log("未找到详情处理方式: %s / %s / %s" % (sheet_or_name, key_or_col, choice), self.status_var)

    def _load_row_from_workbooks(self, sheet_name, key, path_base, path_other):
        """返回 (base_row_values, other_row_values)，缺失一方为空列表。"""
        left, right = [], []
        try:
            wb_b = openpyxl.load_workbook(path_base, data_only=True)
            wb_o = openpyxl.load_workbook(path_other, data_only=True)
            if sheet_name in wb_b.sheetnames and sheet_name in wb_o.sheetnames:
                ws_b = wb_b[sheet_name]
                ws_o = wb_o[sheet_name]
                max_col = max(ws_b.max_column or 1, ws_o.max_column or 1)
                rows_b, _ = load_sheet_rows_full(ws_b, max_col, use_cache=True)
                rows_o, _ = load_sheet_rows_full(ws_o, max_col, use_cache=True)
                dict_b = rows_to_dict(rows_b)
                dict_o = rows_to_dict(rows_o)
                key_norm = key_str_normalized(key)
                for d, out in [(dict_b, left), (dict_o, right)]:
                    for k, row in d.items():
                        if key_str_normalized(k) == key_norm or k == key:
                            out.extend([cell_str(c) for c in row])
                            break
            wb_b.close()
            wb_o.close()
        except Exception:
            pass
        return (left, right)

    def _load_col_from_workbooks(self, sheet_name, col_header, base_side, has_in_other):
        """返回 (基准列值, 另一方列值)，即 (left_vals, right_vals)。基准=本地时 left 来自 local。"""
        left, right = [], []
        try:
            wb_l = openpyxl.load_workbook(self.path_local, data_only=True)
            wb_r = openpyxl.load_workbook(self.path_remote, data_only=True)
            for wb, path_label in [(wb_l, "local"), (wb_r, "remote")]:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                max_col = ws.max_column or 1
                max_row = max(ws.max_row or 1, 1)
                headers = load_sheet_header(ws, max_col)
                col_idx = None
                for i, h in enumerate(headers):
                    if h and header_normalize_for_compare(h) == header_normalize_for_compare(col_header):
                        col_idx = i + 1
                        break
                if col_idx is None:
                    continue
                vals = get_column_values(ws, col_idx, max_row)
                out = [cell_str(v) for v in vals]
                if path_label == "local":
                    left = out
                else:
                    right = out
            wb_l.close()
            wb_r.close()
            # 基准为线上时，left 应为 remote、right 为 local
            if base_side == "remote":
                left, right = right, left
        except Exception:
            pass
        return (left, right)

    def _show_merge_detail_panel_three(self, title, base_vals, local_vals, remote_vals, conflict_type):
        """弹出 Toplevel：三栏显示 BASE | LOCAL | REMOTE，高亮差异。"""
        win = tk.Toplevel(self.root)
        win.title(title)
        win.minsize(900, 500)
        win.geometry("1200x600")
        pad = 8
        
        # 顶部说明
        header = ttk.Frame(win, padding=(pad, pad))
        header.pack(fill=tk.X)
        
        # 根据冲突类型显示不同的说明
        if conflict_type == "delete_conflict_local":
            hint = "本地删除了此行，线上保留/修改了此行"
        elif conflict_type == "delete_conflict_remote":
            hint = "线上删除了此行，本地保留/修改了此行"
        elif conflict_type == "add_conflict":
            hint = "BASE中不存在，双方都新增了此行但内容不同"
        else:
            hint = "BASE中存在，双方都修改了此行但内容不同"
        
        ttk.Label(header, text=hint, font=("Segoe UI", 9), foreground="#666").pack(anchor=tk.W)
        
        # 三栏布局
        cols_frame = ttk.Frame(win)
        cols_frame.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0, pad))
        
        # 创建三个列
        frames = []
        texts = []
        labels = [("BASE", base_vals), ("LOCAL", local_vals), ("REMOTE", remote_vals)]
        
        for i, (label, vals) in enumerate(labels):
            col = ttk.Frame(cols_frame)
            col.grid(row=0, column=i, sticky="nsew", padx=2)
            cols_frame.columnconfigure(i, weight=1)
            
            # 列标题
            lbl_frame = ttk.Frame(col)
            lbl_frame.pack(fill=tk.X)
            ttk.Label(lbl_frame, text=label, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=4)
            
            # 文本框
            txt = tk.Text(col, wrap=tk.WORD, font=("Consolas", 9), padx=6, pady=6)
            sb = ttk.Scrollbar(col, orient=tk.VERTICAL, command=txt.yview)
            
            # 填充内容
            if vals:
                lines = "\n".join("%d: %s" % (j + 1, v) for j, v in enumerate(vals))
            else:
                lines = "(不存在)"
            txt.insert(tk.END, lines)
            txt.config(state=tk.DISABLED)
            
            txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            txt.config(yscrollcommand=sb.set)
            
            frames.append(col)
            texts.append(txt)
        
        cols_frame.rowconfigure(0, weight=1)
        
        # 高亮差异
        n = max(len(base_vals), len(local_vals), len(remote_vals))
        for txt in texts:
            txt.tag_configure("diff", background="#FFF3E0", foreground="#CC6600")
            txt.tag_configure("missing", background="#FFEBEE", foreground="#999")
        
        for i in range(n):
            vb = str(base_vals[i]).strip() if i < len(base_vals) else ""
            vl = str(local_vals[i]).strip() if i < len(local_vals) else ""
            vr = str(remote_vals[i]).strip() if i < len(remote_vals) else ""
            
            # 标记差异
            if vb != vl or vb != vr or vl != vr:
                line_start = "%d.0" % (i + 1)
                line_end = "%d.0" % (i + 2)
                for j, (txt, val) in enumerate([(texts[0], vb), (texts[1], vl), (texts[2], vr)]):
                    try:
                        txt.config(state=tk.NORMAL)
                        if not val:
                            txt.tag_add("missing", line_start, line_end)
                        else:
                            txt.tag_add("diff", line_start, line_end)
                        txt.config(state=tk.DISABLED)
                    except tk.TclError:
                        pass

    def _show_merge_detail_panel(self, title, label_left, label_right, left_vals, right_vals):
        """弹出 Toplevel：左右两栏显示完整参数，每参数一行，双栏滚动同步。"""
        win = tk.Toplevel(self.root)
        win.title(title)
        win.minsize(640, 400)
        win.geometry("900x500")
        pad = 8
        header = ttk.Frame(win, padding=(pad, pad))
        header.pack(fill=tk.X)
        ttk.Label(header, text=label_left, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        ttk.Label(header, text="  |  ", font=("Segoe UI", 10)).pack(side=tk.LEFT)
        ttk.Label(header, text=label_right, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        paned = ttk.PanedWindow(win, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0, pad))
        n = max(len(left_vals), len(right_vals), 1)
        lines_left = "\n".join("%d: %s" % (i + 1, (left_vals[i] if i < len(left_vals) else "")) for i in range(n))
        lines_right = "\n".join("%d: %s" % (i + 1, (right_vals[i] if i < len(right_vals) else "")) for i in range(n))
        f_a = ttk.Frame(paned)
        f_b = ttk.Frame(paned)
        paned.add(f_a, weight=1)
        paned.add(f_b, weight=1)
        txt_a = tk.Text(f_a, wrap=tk.WORD, font=("Consolas", 9), padx=6, pady=6)
        txt_b = tk.Text(f_b, wrap=tk.WORD, font=("Consolas", 9), padx=6, pady=6)
        sb_a = ttk.Scrollbar(f_a, orient=tk.VERTICAL, command=txt_a.yview)
        sb_b = ttk.Scrollbar(f_b, orient=tk.VERTICAL, command=txt_b.yview)
        txt_a.insert(tk.END, lines_left)
        txt_b.insert(tk.END, lines_right)
        txt_a.tag_configure("diff", background="#FFF3E0", foreground="#CC6600")
        txt_b.tag_configure("diff", background="#FFF3E0", foreground="#CC6600")
        for i in range(n):
            vl = str(left_vals[i]).strip() if i < len(left_vals) else ""
            vr = str(right_vals[i]).strip() if i < len(right_vals) else ""
            if vl != vr:
                line_start = "%d.0" % (i + 1)
                line_end = "%d.0" % (i + 2)
                try:
                    txt_a.tag_add("diff", line_start, line_end)
                    txt_b.tag_add("diff", line_start, line_end)
                except tk.TclError:
                    pass

        def _sync_a(first, last):
            sb_a.set(first, last)
            sb_b.set(first, last)
            txt_b.yview_moveto(first)

        def _sync_b(first, last):
            sb_b.set(first, last)
            sb_a.set(first, last)
            txt_a.yview_moveto(first)

        txt_a.config(yscrollcommand=_sync_a)
        txt_b.config(yscrollcommand=_sync_b)
        txt_a.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_a.pack(side=tk.RIGHT, fill=tk.Y)
        txt_b.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_b.pack(side=tk.RIGHT, fill=tk.Y)

        def _strip_line_numbers(text):
            return "\n".join(re.sub(r"^\d+:\s*", "", line) for line in text.splitlines())

        def _copy_without_numbers(widget):
            try:
                sel = widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                win.clipboard_clear()
                win.clipboard_append(_strip_line_numbers(sel))
            except tk.TclError:
                pass
            return "break"

        txt_a.bind("<Control-c>", lambda e: _copy_without_numbers(txt_a))
        txt_b.bind("<Control-c>", lambda e: _copy_without_numbers(txt_b))

        def _copy_left():
            win.clipboard_clear()
            win.clipboard_append("\n".join(str(left_vals[i]) if i < len(left_vals) else "" for i in range(n)))
        def _copy_right():
            win.clipboard_clear()
            win.clipboard_append("\n".join(str(right_vals[i]) if i < len(right_vals) else "" for i in range(n)))
        def _copy_left_as_row():
            win.clipboard_clear()
            win.clipboard_append("\t".join(str(left_vals[i]) if i < len(left_vals) else "" for i in range(n)))
        def _copy_right_as_row():
            win.clipboard_clear()
            win.clipboard_append("\t".join(str(right_vals[i]) if i < len(right_vals) else "" for i in range(n)))

        btn_row = ttk.Frame(win, padding=(pad, 4))
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="复制左侧成列", command=_copy_left).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="复制右侧成列", command=_copy_right).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="复制左侧成行", command=_copy_left_as_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="复制右侧成行", command=_copy_right_as_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_row, text="关闭", command=win.destroy).pack(side=tk.LEFT, padx=8)

    def _on_cancel(self):
        global _merge_instance
        release_merge_lock()
        if _merge_instance is self:
            _merge_instance = None
        self.root.quit()
        self.root.destroy()
        sys.exit(1)

    def _quit_for_update(self):
        global _merge_instance
        release_merge_lock()
        if _merge_instance is self:
            _merge_instance = None
        self.root.quit()
        self.root.destroy()
        sys.exit(1)

    def run(self):
        self.root.mainloop()
