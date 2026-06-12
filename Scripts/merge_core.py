# -*- coding: utf-8 -*-
"""
多模式合并核心（A 新增行 / B 新增列 / C 新增 Sheet / D 冲突选择 / E 智能）。
给定 LOCAL、BASE、REMOTE、MERGED 路径，按 mode 与 base_side 执行合并并写入 MERGED，可选备份。
合并时尽量保留基准表的行高、列宽、单元格样式（颜色/字体等）与顺序。
"""

import os
import shutil
import sys
from copy import copy, deepcopy

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from backup_util import create_merge_backup
from config import ENABLE_MERGE_KEY_DIAGNOSTICS
from conflict import compute_auto_row_actions
from excel_io import (
    build_merged_cells_cache,
    cell_str,
    get_merged_cell_value,
    get_sheet_names,
    get_column_values,
    has_merged_cells,
    header_normalize_for_compare,
    key_str_normalized,
    load_sheet_header,
    load_sheet_rows,
    load_sheet_rows_full,
    merge_ordered_with_new_cols,
    merge_ordered_with_new_rows,
    ordered_keys,
    ordered_keys_normalized,
    row_equal,
    rows_to_dict,
    rows_to_dict_normalized,
)
from log_util import log


def _log_merged_sheet_keys(path_merged, base_side, options, sheet_name="Data_Language"):
    """合并完成后打日志：MERGED 中指定 sheet 的 key 数量及是否含 Store_DrawHero（便于排查基准/选项是否生效）。"""
    try:
        if not ENABLE_MERGE_KEY_DIAGNOSTICS:
            return
        if not path_merged or not os.path.isfile(path_merged):
            return
        wb = openpyxl.load_workbook(path_merged, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return
        ws = wb[sheet_name]
        max_col = ws.max_column or 1
        rows, _ = load_sheet_rows_full(ws, max_col, use_cache=True)
        keys = set(key_str_normalized(r[0]) if r else "" for r in rows)
        keys.discard("")
        has_draw = any((k or "").lower() == "store_drawhero" or ("store" in (k or "").lower() and "drawhero" in (k or "").lower()) for k in keys)
        log("[Options] MERGED %s keys=%d Store_DrawHero=%s（基准=%s）" % (
            sheet_name, len(keys), "有" if has_draw else "无", "本地" if base_side == "local" else "线上"))
        wb.close()
    except Exception as e:
        log("[Options] MERGED 校验日志异常: %s" % e, is_error=True)


# ---------------------------------------------------------------------------
# 复制单元格样式与行高/列宽（保留原表格式）
# ---------------------------------------------------------------------------

def _copy_cell_style(src_cell, dst_cell):
    """将 src 单元格的样式复制到 dst，不复制 value（由调用方赋值）。"""
    if src_cell.has_style:
        if src_cell.font:
            dst_cell.font = copy(src_cell.font)
        if src_cell.border:
            dst_cell.border = copy(src_cell.border)
        if src_cell.fill:
            dst_cell.fill = copy(src_cell.fill)
        if src_cell.number_format:
            dst_cell.number_format = copy(src_cell.number_format)
        if src_cell.alignment:
            dst_cell.alignment = copy(src_cell.alignment)
        if src_cell.protection:
            dst_cell.protection = copy(src_cell.protection)


def _copy_cell_value_and_style(src_cell, dst_cell, font=None):
    dst_cell.value = src_cell.value
    _copy_cell_style(src_cell, dst_cell)
    if font is not None:
        dst_cell.font = font


def _copy_row_with_style(ws_src, ws_dst, row_src, row_dst, max_col):
    """将 ws_src 的 row_src 行（值+样式）复制到 ws_dst 的 row_dst 行。合并格取区域左上角值。"""
    for c in range(1, max_col + 1):
        src_c = ws_src.cell(row=row_src, column=c)
        val = get_merged_cell_value(ws_src, row_src, c)
        dst_c = ws_dst.cell(row=row_dst, column=c, value=val)
        _copy_cell_style(src_c, dst_c)
    if row_src in ws_src.row_dimensions and ws_src.row_dimensions[row_src].height is not None:
        ws_dst.row_dimensions[row_dst].height = ws_src.row_dimensions[row_src].height


def _copy_col_with_style(ws_src, ws_dst, col_src, col_dst, max_row, merged_cache=None):
    """将 ws_src 的 col_src 列（值+样式）复制到 ws_dst 的 col_dst 列。合并格取区域左上角值。"""
    for r in range(1, max_row + 1):
        src_c = ws_src.cell(row=r, column=col_src)
        val = get_merged_cell_value(ws_src, r, col_src, merged_cache)
        dst_c = ws_dst.cell(row=r, column=col_dst, value=val)
        _copy_cell_style(src_c, dst_c)
    letter_src = get_column_letter(col_src)
    letter_dst = get_column_letter(col_dst)
    if letter_src in ws_src.column_dimensions and ws_src.column_dimensions[letter_src].width is not None:
        ws_dst.column_dimensions[letter_dst].width = ws_src.column_dimensions[letter_src].width


def _copy_merged_cells(ws_src, ws_dst):
    """将 ws_src 的合并单元格范围原样复制到 ws_dst（相同行列坐标）。"""
    try:
        for rng in list(ws_src.merged_cells.ranges):
            ws_dst.merge_cells(
                start_row=rng.min_row, start_column=rng.min_col,
                end_row=rng.max_row, end_column=rng.max_col,
            )
    except Exception:
        pass


def _unique_table_name(wb, preferred):
    name = preferred or "Table"
    try:
        if not wb._duplicate_name(name):
            return name
        base = name[:240] if len(name) > 240 else name
        for i in range(1, 10000):
            candidate = "%s_%d" % (base, i)
            if not wb._duplicate_name(candidate):
                return candidate
    except Exception:
        pass
    return name


def _copy_sheet_metadata(ws_src, ws_dst):
    """Copy common worksheet-level metadata that survives openpyxl round trips."""
    for attr in ("sheet_format", "sheet_properties", "page_margins", "page_setup", "print_options"):
        try:
            setattr(ws_dst, attr, copy(getattr(ws_src, attr)))
        except Exception:
            pass
    try:
        ws_dst.freeze_panes = ws_src.freeze_panes
    except Exception:
        pass
    try:
        if ws_src.auto_filter and ws_src.auto_filter.ref:
            ws_dst.auto_filter.ref = ws_src.auto_filter.ref
    except Exception:
        pass
    try:
        if ws_src.data_validations:
            ws_dst.data_validations = deepcopy(ws_src.data_validations)
    except Exception:
        pass
    try:
        ws_dst.conditional_formatting = deepcopy(ws_src.conditional_formatting)
    except Exception:
        pass
    for table in ws_src.tables.values():
        try:
            table_copy = deepcopy(table)
            table_copy.name = _unique_table_name(ws_dst.parent, table_copy.name)
            table_copy.displayName = table_copy.name
            ws_dst.add_table(table_copy)
        except Exception:
            pass
    try:
        for row_idx, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[row_idx] = copy(dim)
            ws_dst.row_dimensions[row_idx].worksheet = ws_dst
        for col_key, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col_key] = copy(dim)
            ws_dst.column_dimensions[col_key].worksheet = ws_dst
    except Exception:
        pass


def _copy_worksheet_to_new_sheet(wb_dst, ws_src, title=None):
    ws_new = wb_dst.create_sheet(title or ws_src.title)
    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
            _copy_cell_style(cell, dst)
    _copy_sheet_metadata(ws_src, ws_new)
    _copy_merged_cells(ws_src, ws_new)
    return ws_new


def _shift_merged_cells_after_insert_rows(ws, insert_at_row, amount=1):
    """在 insert_at_row 处插入了 amount 行后，更新 ws 中受影响的合并范围（下移）。"""
    try:
        to_reapply = []
        for rng in list(ws.merged_cells.ranges):
            if rng.max_row >= insert_at_row:
                new_min_r = rng.min_row + amount if rng.min_row >= insert_at_row else rng.min_row
                new_max_r = rng.max_row + amount
                to_reapply.append((str(rng), new_min_r, rng.min_col, new_max_r, rng.max_col))
        for old_ref, nr, nc, xr, xc in to_reapply:
            ws.unmerge_cells(old_ref)
            ws.merge_cells(start_row=nr, start_column=nc, end_row=xr, end_column=xc)
    except Exception:
        pass


def _shift_merged_cells_after_insert_cols(ws, insert_at_col, amount=1):
    """在 insert_at_col 处插入了 amount 列后，更新 ws 中受影响的合并范围（右移）。"""
    try:
        to_reapply = []
        for rng in list(ws.merged_cells.ranges):
            if rng.max_col >= insert_at_col:
                new_min_c = rng.min_col + amount if rng.min_col >= insert_at_col else rng.min_col
                new_max_c = rng.max_col + amount
                to_reapply.append((str(rng), rng.min_row, new_min_c, rng.max_row, new_max_c))
        for old_ref, nr, nc, xr, xc in to_reapply:
            ws.unmerge_cells(old_ref)
            ws.merge_cells(start_row=nr, start_column=nc, end_row=xr, end_column=xc)
    except Exception:
        pass


def _copy_row_merged_ranges_to(ws_src, row_src, ws_dst, row_dst):
    """把 ws_src 中与 row_src 相交的合并范围，在 ws_dst 的 row_dst 行上按列宽合并。"""
    try:
        for rng in list(ws_src.merged_cells.ranges):
            if rng.min_row <= row_src <= rng.max_row and rng.min_col < rng.max_col:
                ws_dst.merge_cells(
                    start_row=row_dst, start_column=rng.min_col,
                    end_row=row_dst, end_column=rng.max_col,
                )
    except Exception:
        pass


def _copy_col_merged_ranges_to(ws_src, col_src, ws_dst, col_dst):
    """把 ws_src 中与 col_src 相交的合并范围，在 ws_dst 的 col_dst 列上按行高合并。"""
    try:
        for rng in list(ws_src.merged_cells.ranges):
            if rng.min_col <= col_src <= rng.max_col and rng.min_row < rng.max_row:
                ws_dst.merge_cells(
                    start_row=rng.min_row, start_column=col_dst,
                    end_row=rng.max_row, end_column=col_dst,
                )
    except Exception:
        pass


def _row_key_to_index(ws, max_col):
    """构建 D 模式用的 key->原始行号，兼容空 key / 重复 key 的 __row_N 规则。"""
    rows, idx = load_sheet_rows_full(ws, max_col, use_cache=True) if ws else ([], [])
    key_to_row = {}
    seen = set()
    for i, row in enumerate(rows):
        if i >= len(idx):
            break
        raw = key_str_normalized(row[0]) if row else ""
        key = raw or "__row_%d" % idx[i]
        if key in seen:
            key = "__row_%d" % idx[i]
        seen.add(key)
        key_to_row[key] = idx[i]
    return key_to_row


def _normalized_header_to_index(headers):
    """构建规范化表头 -> 1-based 列号。重复表头保留第一次出现。"""
    out = {}
    for i, header in enumerate(headers):
        key = key_str_normalized(header)
        if key and key not in out:
            out[key] = i + 1
    return out


def _compare_header_to_index(headers):
    """构建 header_normalize_for_compare 规则下的表头索引。"""
    out = {}
    for i, header in enumerate(headers):
        key = header_normalize_for_compare(header)
        if key and key not in out:
            out[key] = i + 1
    return out


def _row_copy_plan(ws_out, ws_source, ws_base=None, max_col=None):
    max_col = max_col or ((ws_out.max_column or 1) if ws_out else 1)
    header_out = load_sheet_header(ws_out, max_col) if ws_out else []
    header_source = load_sheet_header(ws_source, max(ws_source.max_column or 1, max_col)) if ws_source else []
    header_base = load_sheet_header(ws_base, max((ws_base.max_column or 1) if ws_base else 1, max_col)) if ws_base else []
    source_map = _compare_header_to_index(header_source)
    base_map = _compare_header_to_index(header_base)
    has_named_headers = bool(source_map) and bool(_compare_header_to_index(header_out))
    plan = []
    for out_col in range(1, max_col + 1):
        src_col = out_col
        base_col = out_col
        if has_named_headers and out_col <= len(header_out):
            norm = header_normalize_for_compare(header_out[out_col - 1])
            src_col = source_map.get(norm, out_col)
            base_col = base_map.get(norm, out_col)
        plan.append((out_col, src_col, base_col))
    return plan


def _column_values_by_index(ws, col_indices, max_row):
    """一次性读取同一 Sheet 的多列，复用合并单元格缓存。"""
    if ws is None:
        return {}
    merged_cache = build_merged_cells_cache(ws) if has_merged_cells(ws) else None
    values = {}
    for col_idx in sorted(set(i for i in col_indices if i)):
        values[col_idx] = get_column_values(
            ws, col_idx, max_row, merged_cache=merged_cache,
        )
    return values


# ---------------------------------------------------------------------------
# 模式 A：新增行插入（基准 + 另一侧新增行，按前缀组末尾插入）
# ---------------------------------------------------------------------------

def _merge_mode_a_impl(path_base_side, path_other_side, path_merged, base_side):
    """以 base_side 文件为基准，将 other 中新增行按前缀插入，写入 path_merged。"""
    wb_base = openpyxl.load_workbook(path_base_side, data_only=True)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=True)
    font_new = Font(color="008000")

    seen = set()
    sheet_names = []
    for n in get_sheet_names(wb_base):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in get_sheet_names(wb_other):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in sheet_names:
        ws_base = wb_base[sheet_name] if sheet_name in wb_base.sheetnames else None
        ws_other = wb_other[sheet_name] if sheet_name in wb_other.sheetnames else None
        if not ws_base and not ws_other:
            continue
        max_col = max(
            (ws_base.max_column or 1) if ws_base else 1,
            (ws_other.max_column or 1) if ws_other else 1,
        )
        base_rows = {}
        base_ordered = []
        other_rows = {}
        other_ordered = []
        if ws_base:
            rows = load_sheet_rows(ws_base, max_col)
            base_rows = rows_to_dict(rows)
            base_ordered = ordered_keys(rows)
        if ws_other:
            rows = load_sheet_rows(ws_other, max_col)
            other_rows = rows_to_dict(rows)
            other_ordered = ordered_keys(rows)

        base_set = set(base_rows)
        new_keys = [k for k in other_ordered if k not in base_set]
        merged_ordered = merge_ordered_with_new_rows(base_ordered, new_keys)
        base_set_plus_new = set(merged_ordered)

        merged_rows = []
        for k in merged_ordered:
            if k in base_rows:
                merged_rows.append(list(base_rows[k]))
            else:
                merged_rows.append(list(other_rows.get(k, [])))
            while len(merged_rows[-1]) < max_col:
                merged_rows[-1].append("")

        if not merged_rows and not base_rows and not other_rows:
            continue
        ws_out = wb_out.create_sheet(sheet_name)
        for r, row_list in enumerate(merged_rows, start=1):
            key_str = cell_str(row_list[0]) if row_list else ""
            is_new = key_str in base_set_plus_new and key_str in new_keys
            for c, val in enumerate(row_list, start=1):
                cell = ws_out.cell(row=r, column=c, value=val)
                if is_new:
                    cell.font = font_new

    wb_base.close()
    wb_other.close()
    if not wb_out.sheetnames:
        wb_out.create_sheet("Data")
    os.makedirs(os.path.dirname(os.path.abspath(path_merged)) or ".", exist_ok=True)
    wb_out.save(path_merged)
    wb_out.close()
    log("[Mode A] 新增行插入完成 MERGED=%s" % path_merged)


# ---------------------------------------------------------------------------
# 模式 B：新增列插入（基准 + 另一侧新增列，按列名前缀组末尾插入）
# ---------------------------------------------------------------------------

def _merge_mode_b_impl(path_base_side, path_other_side, path_merged, base_side):
    """以 base_side 为基准，将 other 中新增列按表头前缀插入，写入 path_merged。"""
    wb_base = openpyxl.load_workbook(path_base_side, data_only=True)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=True)
    font_new = Font(color="008000")

    seen = set()
    sheet_names = []
    for n in get_sheet_names(wb_base):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in get_sheet_names(wb_other):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in sheet_names:
        ws_base = wb_base[sheet_name] if sheet_name in wb_base.sheetnames else None
        ws_other = wb_other[sheet_name] if sheet_name in wb_other.sheetnames else None
        if not ws_base and not ws_other:
            continue
        max_col_b = (ws_base.max_column or 1) if ws_base else 1
        max_col_o = (ws_other.max_column or 1) if ws_other else 1
        max_col = max(max_col_b, max_col_o)

        header_b = load_sheet_header(ws_base, max_col) if ws_base else []
        header_o = load_sheet_header(ws_other, max_col) if ws_other else []
        while len(header_b) < max_col:
            header_b.append("")
        while len(header_o) < max_col:
            header_o.append("")
        base_cols_ordered = [c for c in header_b if c]
        if not base_cols_ordered and header_b:
            base_cols_ordered = [cell_str(header_b[i]) or ("_col_%d" % (i + 1)) for i in range(len(header_b))]
        other_cols_ordered = [c for c in header_o if c]
        if not other_cols_ordered and header_o:
            other_cols_ordered = [cell_str(header_o[i]) or ("_col_%d" % (i + 1)) for i in range(len(header_o))]
        base_set = set(base_cols_ordered)
        new_cols = [c for c in other_cols_ordered if c not in base_set]
        merged_col_order = merge_ordered_with_new_cols(base_cols_ordered, new_cols)

        col_to_idx_base = {}
        for i, h in enumerate(header_b):
            if h:
                col_to_idx_base[h] = i + 1
        col_to_idx_other = {}
        for i, h in enumerate(header_o):
            if h:
                col_to_idx_other[h] = i + 1

        max_row = max(
            (ws_base.max_row or 1) if ws_base else 1,
            (ws_other.max_row or 1) if ws_other else 1,
        )
        needed_base_cols = []
        needed_other_cols = []
        for col_key in merged_col_order:
            idx_b = col_to_idx_base.get(col_key)
            idx_o = col_to_idx_other.get(col_key)
            if idx_o is not None and (idx_b is None or col_key in new_cols):
                needed_other_cols.append(idx_o)
            elif idx_b is not None:
                needed_base_cols.append(idx_b)
        base_col_values = _column_values_by_index(ws_base, needed_base_cols, max_row)
        other_col_values = _column_values_by_index(ws_other, needed_other_cols, max_row)
        ws_out = wb_out.create_sheet(sheet_name)
        for c, col_key in enumerate(merged_col_order, start=1):
            is_new_col = col_key in new_cols
            idx_b = col_to_idx_base.get(col_key)
            idx_o = col_to_idx_other.get(col_key)
            for r in range(1, max_row + 1):
                if idx_o is not None and (idx_b is None or is_new_col):
                    val = other_col_values.get(idx_o, [""] * max_row)[r - 1] if ws_other else ""
                else:
                    val = base_col_values.get(idx_b, [""] * max_row)[r - 1] if ws_base and idx_b is not None else ""
                cell = ws_out.cell(row=r, column=c, value=val)
                if is_new_col:
                    cell.font = font_new

    wb_base.close()
    wb_other.close()
    if not wb_out.sheetnames:
        wb_out.create_sheet("Data")
    os.makedirs(os.path.dirname(os.path.abspath(path_merged)) or ".", exist_ok=True)
    wb_out.save(path_merged)
    wb_out.close()
    log("[Mode B] 新增列插入完成 MERGED=%s" % path_merged)


# ---------------------------------------------------------------------------
# 模式 C：新增 Sheet 插入（基准 + 另一侧新增的 Sheet 整表追加）
# ---------------------------------------------------------------------------

def _merge_mode_c_impl(path_base_side, path_other_side, path_merged, base_side):
    """以 base_side 为基准，将 other 中多出的 Sheet 整表复制追加，保留格式。"""
    wb_base = openpyxl.load_workbook(path_base_side, data_only=False)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=False)

    base_sheets = set(get_sheet_names(wb_base))
    other_sheets = get_sheet_names(wb_other)
    new_sheets = [n for n in other_sheets if n not in base_sheets]

    os.makedirs(os.path.dirname(os.path.abspath(path_merged)) or ".", exist_ok=True)
    if not new_sheets:
        wb_base.close()
        wb_other.close()
        if os.path.abspath(path_base_side) != os.path.abspath(path_merged):
            shutil.copy2(path_base_side, path_merged)
        log("[Mode C] 无新增 Sheet，直接保留完整工作簿 MERGED=%s" % path_merged)
        return

    wb_out = wb_base
    for name in new_sheets:
        _copy_worksheet_to_new_sheet(wb_out, wb_other[name], name)

    wb_other.close()
    if not wb_out.sheetnames:
        wb_out.create_sheet("Data")
    wb_out.save(path_merged)
    wb_out.close()
    log("[Mode C] 新增 Sheet 插入完成 MERGED=%s new_sheets=%s" % (path_merged, new_sheets))


# ---------------------------------------------------------------------------
# 模式 D：冲突行/列选择（按 d_choices 写回 MERGED）
# ---------------------------------------------------------------------------

def _merge_mode_d_impl(path_local, path_remote, path_merged, path_base, d_choices, path_initial_merged=None):
    """
    d_choices: list of dict 每项 {"sheet", "key", "choice": "local"|"remote", "kind": "row"|"column"}。
    path_initial_merged: 若提供则以此为底写 path_merged，否则以 local 为底。E 模式传入 A+B+C 的结果。
    """
    if path_initial_merged and os.path.isfile(path_initial_merged):
        wb_out = openpyxl.load_workbook(path_initial_merged, data_only=False)
    else:
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        wb_local = openpyxl.load_workbook(path_local, data_only=False)
        for sheet_name in get_sheet_names(wb_local):
            _copy_worksheet_to_new_sheet(wb_out, wb_local[sheet_name], sheet_name)
        wb_local.close()

    wb_local = openpyxl.load_workbook(path_local, data_only=False)
    wb_base = openpyxl.load_workbook(path_base, data_only=False)
    wb_remote = openpyxl.load_workbook(path_remote, data_only=False)

    sheet_names = list(wb_out.sheetnames)

    for item in d_choices or []:
        sheet_name = item.get("sheet")
        key = item.get("key")
        choice = item.get("choice", "local")
        kind = item.get("kind", "row")
        if sheet_name not in wb_out.sheetnames:
            continue
        ws_out = wb_out[sheet_name]
        ws_l = wb_local[sheet_name] if sheet_name in wb_local.sheetnames else None
        ws_b = wb_base[sheet_name] if sheet_name in wb_base.sheetnames else None
        ws_r = wb_remote[sheet_name] if sheet_name in wb_remote.sheetnames else None
        max_col = ws_out.max_column or 1
        max_row = ws_out.max_row or 1
        if kind == "row":
            key_norm = key_str_normalized(key) or key
            key_to_row_out = _row_key_to_index(ws_out, max_col)
            row_idx_out = key_to_row_out.get(key_norm)
            source = ws_r if choice == "remote" else ws_l
            if not source:
                if row_idx_out is not None:
                    ws_out.delete_rows(row_idx_out, 1)
                continue
            key_to_row_source = _row_key_to_index(source, max_col)
            row_idx_source = key_to_row_source.get(key_norm)
            if row_idx_source is None:
                if row_idx_out is not None:
                    ws_out.delete_rows(row_idx_out, 1)
                continue
            if row_idx_out is None:
                row_idx_out = min(row_idx_source, (ws_out.max_row or 0) + 1)
                ws_out.insert_rows(row_idx_out, 1)
                _shift_merged_cells_after_insert_rows(ws_out, row_idx_out, 1)
            _font_mod = Font(color="CC6600")
            auto_type = item.get("type")
            copy_plan = _row_copy_plan(ws_out, source, ws_b, max_col)
            if auto_type in ("take_local", "take_remote"):
                row_idx_base = None
                if ws_b is not None:
                    key_to_row_base = _row_key_to_index(ws_b, max_col)
                    row_idx_base = key_to_row_base.get(key_norm)
                for out_col, src_col, base_col in copy_plan:
                    src_c = source.cell(row=row_idx_source, column=src_col)
                    base_val = ws_b.cell(row=row_idx_base, column=base_col).value if ws_b is not None and row_idx_base else None
                    if cell_str(src_c.value) == cell_str(base_val):
                        continue
                    dst_c = ws_out.cell(row=row_idx_out, column=out_col)
                    _copy_cell_value_and_style(src_c, dst_c, _font_mod)
                continue
            for out_col, src_col, _base_col in copy_plan:
                src_c = source.cell(row=row_idx_source, column=src_col)
                dst_c = ws_out.cell(row=row_idx_out, column=out_col)
                _copy_cell_value_and_style(src_c, dst_c, _font_mod)
        else:
            header_out = load_sheet_header(ws_out, max_col)
            header_l = load_sheet_header(ws_l, max_col) if ws_l else []
            header_r = load_sheet_header(ws_r, max_col) if ws_r else []
            norm_key = header_normalize_for_compare(key)
            col_idx_out = _compare_header_to_index(header_out).get(norm_key)
            col_idx_l = _compare_header_to_index(header_l).get(norm_key)
            col_idx_r = _compare_header_to_index(header_r).get(norm_key)
            col_idx = col_idx_l if choice == "local" else col_idx_r
            source = ws_l if choice == "local" else ws_r
            if col_idx is None or not source:
                continue
            if col_idx_out is None:
                col_idx_out = (ws_out.max_column or 0) + 1
            _font_mod = Font(color="CC6600")
            for r in range(1, max_row + 1):
                src_c = source.cell(row=r, column=col_idx)
                dst_c = ws_out.cell(row=r, column=col_idx_out, value=src_c.value)
                _copy_cell_style(src_c, dst_c)
                dst_c.font = _font_mod

    wb_local.close()
    wb_base.close()
    wb_remote.close()
    if not wb_out.sheetnames:
        wb_out.create_sheet("Data")
    os.makedirs(os.path.dirname(os.path.abspath(path_merged)) or ".", exist_ok=True)
    wb_out.save(path_merged)
    wb_out.close()
    log("[Mode D] 冲突选择写回完成 MERGED=%s" % path_merged)


# ---------------------------------------------------------------------------
# 模式 E：智能（A → B → C → D，D 需 d_choices）
# ---------------------------------------------------------------------------

def _merge_mode_e_impl(path_local, path_base, path_remote, path_merged, base_side, d_choices):
    """顺序执行 A、B、C、D；前一步输出作为下一步输入，D 在 C 的结果上应用 d_choices 写入 path_merged。"""
    import tempfile
    path_base_side = path_local if base_side == "local" else path_remote
    path_other_side = path_remote if base_side == "local" else path_local

    fd, path_a = tempfile.mkstemp(suffix=".xlsx", prefix="merge_a_")
    os.close(fd)
    try:
        _merge_mode_a_impl(path_base_side, path_other_side, path_a, base_side)
        fd2, path_b = tempfile.mkstemp(suffix=".xlsx", prefix="merge_b_")
        os.close(fd2)
        try:
            _merge_mode_b_impl(path_a, path_other_side, path_b, base_side)
            fd3, path_c = tempfile.mkstemp(suffix=".xlsx", prefix="merge_c_")
            os.close(fd3)
            try:
                _merge_mode_c_impl(path_b, path_other_side, path_c, base_side)
                _merge_mode_d_impl(
                    path_local, path_remote, path_merged, path_base,
                    d_choices, path_initial_merged=path_c,
                )
            finally:
                try:
                    os.unlink(path_c)
                except Exception:
                    pass
        finally:
            try:
                os.unlink(path_b)
            except Exception:
                pass
    finally:
        try:
            os.unlink(path_a)
        except Exception:
            pass
    log("[Mode E] 智能合并完成 MERGED=%s" % path_merged)


# ---------------------------------------------------------------------------
# 选项驱动：C 删除行 / D 删除列 / F 删除 Sheet（path_in 为当前工作簿，other 为对照）
# ---------------------------------------------------------------------------

def _merge_delete_rows_impl(path_in, path_other_side, path_out):
    """从 path_in 中删除「在 base 侧有而 other 侧没有」的行（按首列 key）。"""
    wb_in = openpyxl.load_workbook(path_in, data_only=False)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=True)
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col = max(ws_in.max_column or 1, ws_o.max_column or 1)
        rows_o, _ = load_sheet_rows_full(ws_o, max_col, use_cache=True)
        keys_other = set(key_str_normalized(r[0]) if r else "" for r in rows_o)
        keys_other.discard("")
        rows_in, idx_in = load_sheet_rows_full(ws_in, max_col, use_cache=True)
        to_delete = []
        for i, r in enumerate(rows_in):
            k = key_str_normalized(r[0]) if r else ""
            if k and k not in keys_other:
                to_delete.append(idx_in[i] if i < len(idx_in) else i + 1)
        for row_idx in sorted(set(to_delete), reverse=True):
            ws_in.delete_rows(row_idx, 1)
    wb_other.close()
    os.makedirs(os.path.dirname(os.path.abspath(path_out)) or ".", exist_ok=True)
    wb_in.save(path_out)
    wb_in.close()
    log("[Option C] 删除行完成 %s" % path_out)


def _merge_delete_cols_impl(path_in, path_other_side, path_out):
    """从 path_in 中删除「在 base 有而 other 没有」的列（按表头）。"""
    wb_in = openpyxl.load_workbook(path_in, data_only=False)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=True)
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col = max(ws_in.max_column or 1, ws_o.max_column or 1)
        header_in = load_sheet_header(ws_in, max_col)
        header_o = load_sheet_header(ws_o, max_col)
        keys_other_norm = set(header_normalize_for_compare(h) for h in header_o if h)
        to_delete = []
        for c in range(len(header_in) - 1, -1, -1):
            h = header_in[c] if c < len(header_in) else ""
            if h and header_normalize_for_compare(h) not in keys_other_norm:
                col_1based = c + 1
                # 禁止删除第 1 列（key 列），否则"仅本地有"的行会因 key 丢失而表现为整行丢失
                if col_1based != 1:
                    to_delete.append(col_1based)
        for col_idx in sorted(to_delete, reverse=True):
            ws_in.delete_cols(col_idx, 1)
    wb_other.close()
    os.makedirs(os.path.dirname(os.path.abspath(path_out)) or ".", exist_ok=True)
    wb_in.save(path_out)
    wb_in.close()
    log("[Option D] 删除列完成 %s" % path_out)


def _merge_delete_sheets_impl(path_in, path_other_side, path_out):
    """从 path_in 中删除「在 base 有而 other 没有」的 Sheet。"""
    wb_in = openpyxl.load_workbook(path_in, data_only=False)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=True)
    other_sheets = set(get_sheet_names(wb_other))
    wb_other.close()
    for name in list(wb_in.sheetnames):
        if name not in other_sheets:
            del wb_in[name]
    if not wb_in.sheetnames:
        wb_in.create_sheet("Data")
    os.makedirs(os.path.dirname(os.path.abspath(path_out)) or ".", exist_ok=True)
    wb_in.save(path_out)
    wb_in.close()
    log("[Option F] 删除 Sheet 完成 %s" % path_out)


# ---------------------------------------------------------------------------
# 按选项集合执行管道（A 不变=不增行 B 不变=不增列 C 删除行 D 删除列 E 新增Sheet F 删除Sheet G 冲突）
# ---------------------------------------------------------------------------

def _merge_mode_a_preserve_format(path_in, path_other_side, path_out, base_side):
    """在基准副本 path_in 上插入另一侧新增行，保留行高与单元格样式，写入 path_out。"""
    wb_in = openpyxl.load_workbook(path_in, data_only=False)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=False)
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col = max(ws_in.max_column or 1, ws_o.max_column or 1)
        rows_in, idx_in = load_sheet_rows_full(ws_in, max_col, use_cache=True)
        rows_o, idx_o = load_sheet_rows_full(ws_o, max_col, use_cache=True)
        base_keys = set(key_str_normalized(r[0]) if r else "" for r in rows_in)
        base_keys.discard("")
        key_to_row_in = {}
        for i, r in enumerate(rows_in):
            k = key_str_normalized(r[0]) if r else ""
            if k and i < len(idx_in):
                key_to_row_in[k] = idx_in[i]
        base_ordered = ordered_keys_normalized(rows_in)
        other_ordered = ordered_keys_normalized(rows_o)
        new_keys = [k for k in other_ordered if k not in base_keys]
        merged_ordered = merge_ordered_with_new_rows(base_ordered, new_keys)
        # 每个 key 可能对应多行（首列非唯一），按行顺序保存所有行号
        key_to_rows_other = {}
        for i, r in enumerate(rows_o):
            k = key_str_normalized(r[0]) if r else ""
            if k and i < len(idx_o):
                key_to_rows_other.setdefault(k, []).append(idx_o[i])
        last_base_row = None
        inserts = []  # (insert_after, new_key, order_idx) 其中 order_idx 用于 insert_after=0 时保持 merged 顺序
        for order_idx, k in enumerate(merged_ordered):
            if k in key_to_row_in:
                last_base_row = key_to_row_in[k]
            elif k in key_to_rows_other:
                insert_after = (last_base_row if last_base_row is not None else 0)
                inserts.append((insert_after, k, order_idx))
        # 从高 insert_after 往低处理；同一 insert_after 内按 order_idx 倒序，这样先插的不会把后插的顶下去，最终顺序与 merged 一致
        def _sort_key(item):
            ia, key, oi = item[0], item[1], item[2]
            if ia >= 1:
                return (-ia, -oi, 0)
            return (0, -oi, 0)  # insert_after=0 排最后，同组内 -order_idx 大的先处理
        for insert_after, new_key, _ in sorted(inserts, key=_sort_key):
            other_rows = key_to_rows_other[new_key]
            insert_at_row = (insert_after + 1) if insert_after >= 1 else 2  # 0 时插在表头后（第 2 行）
            # 批量插入多行，减少 openpyxl 内部操作
            num_rows = len(other_rows)
            if num_rows > 0:
                ws_in.insert_rows(insert_at_row, num_rows)
                _shift_merged_cells_after_insert_rows(ws_in, insert_at_row, num_rows)
                # 倒序填充，保持与原始逻辑一致的顺序（后插入的行在上面）
                for i, other_row_idx in enumerate(reversed(other_rows)):
                    row_dst = insert_at_row + (num_rows - 1 - i)
                    _copy_row_with_style(ws_o, ws_in, other_row_idx, row_dst, max_col)
                    _copy_row_merged_ranges_to(ws_o, other_row_idx, ws_in, row_dst)
    wb_other.close()
    os.makedirs(os.path.dirname(os.path.abspath(path_out)) or ".", exist_ok=True)
    wb_in.save(path_out)
    wb_in.close()
    log("[Option A] 新增行插入（保留格式）完成 %s" % path_out)


def _merge_mode_b_preserve_format(path_in, path_other_side, path_out, base_side):
    """在基准副本 path_in 上插入另一侧新增列，保留列宽与单元格样式，写入 path_out。"""
    wb_in = openpyxl.load_workbook(path_in, data_only=False)
    wb_other = openpyxl.load_workbook(path_other_side, data_only=False)
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col_in = ws_in.max_column or 1
        max_col_o = ws_o.max_column or 1
        max_col = max(max_col_in, max_col_o)
        header_in = load_sheet_header(ws_in, max_col)
        header_o = load_sheet_header(ws_o, max_col)
        header_in_norm_to_idx = _compare_header_to_index(header_in)
        base_set_norm = set(header_in_norm_to_idx)
        other_ordered = [h for h in header_o if h]
        new_cols = [h for h in other_ordered if header_normalize_for_compare(h) not in base_set_norm]
        merged_col_order = merge_ordered_with_new_cols([h for h in header_in if h], new_cols)
        col_to_idx_o = {}
        for i, h in enumerate(header_o):
            if h:
                col_to_idx_o[h] = i + 1
        last_base_col = None
        inserts = []
        for h in merged_col_order:
            idx_in = header_in_norm_to_idx.get(header_normalize_for_compare(h))
            if idx_in is not None:
                last_base_col = idx_in
            elif h in col_to_idx_o:
                # 新增列排在首列前时 last_base_col 为 None，用 0 表示插在第 1 列前
                insert_after = (last_base_col if last_base_col is not None else 0)
                inserts.append((insert_after, h))
        max_row = max(ws_in.max_row or 1, ws_o.max_row or 1)
        merged_cache_o = build_merged_cells_cache(ws_o) if has_merged_cells(ws_o) and inserts else None
        # 从后往前插入，避免列索引变化
        for insert_after, col_key in sorted(inserts, key=lambda x: -x[0]):
            col_o = col_to_idx_o[col_key]
            # 禁止在列 1 前插入：插在列 1 会挤掉 key 列；插在列 2 会形成"双 key 列"导致错位。改为追加到表尾。
            if insert_after >= 1:
                insert_at_col = insert_after + 1
            else:
                insert_at_col = (ws_in.max_column or 1) + 1
            ws_in.insert_cols(insert_at_col, 1)
            _shift_merged_cells_after_insert_cols(ws_in, insert_at_col, 1)
            _copy_col_with_style(ws_o, ws_in, col_o, insert_at_col, max_row, merged_cache_o)
            _copy_col_merged_ranges_to(ws_o, col_o, ws_in, insert_at_col)
    wb_other.close()
    os.makedirs(os.path.dirname(os.path.abspath(path_out)) or ".", exist_ok=True)
    wb_in.save(path_out)
    wb_in.close()
    log("[Option B] 新增列插入（保留格式）完成 %s" % path_out)


def _merge_mode_a_preserve_format_wb(wb_in, wb_other, base_side):
    """在已打开的基准工作簿中插入另一侧新增行，返回新增行数。"""
    inserted_count = 0
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col = max(ws_in.max_column or 1, ws_o.max_column or 1)
        rows_in, idx_in = load_sheet_rows_full(ws_in, max_col, use_cache=True)
        rows_o, idx_o = load_sheet_rows_full(ws_o, max_col, use_cache=True)
        base_keys = set(key_str_normalized(r[0]) if r else "" for r in rows_in)
        base_keys.discard("")
        key_to_row_in = {}
        for i, r in enumerate(rows_in):
            k = key_str_normalized(r[0]) if r else ""
            if k and i < len(idx_in):
                key_to_row_in[k] = idx_in[i]
        base_ordered = ordered_keys_normalized(rows_in)
        other_ordered = ordered_keys_normalized(rows_o)
        new_keys = [k for k in other_ordered if k not in base_keys]
        merged_ordered = merge_ordered_with_new_rows(base_ordered, new_keys)
        key_to_rows_other = {}
        for i, r in enumerate(rows_o):
            k = key_str_normalized(r[0]) if r else ""
            if k and i < len(idx_o):
                key_to_rows_other.setdefault(k, []).append(idx_o[i])
        last_base_row = None
        inserts = []
        for order_idx, k in enumerate(merged_ordered):
            if k in key_to_row_in:
                last_base_row = key_to_row_in[k]
            elif k in key_to_rows_other:
                insert_after = (last_base_row if last_base_row is not None else 0)
                inserts.append((insert_after, k, order_idx))

        def _sort_key(item):
            ia, _key, oi = item[0], item[1], item[2]
            if ia >= 1:
                return (-ia, -oi, 0)
            return (0, -oi, 0)

        for insert_after, new_key, _ in sorted(inserts, key=_sort_key):
            other_rows = key_to_rows_other[new_key]
            insert_at_row = (insert_after + 1) if insert_after >= 1 else 2
            num_rows = len(other_rows)
            if num_rows <= 0:
                continue
            ws_in.insert_rows(insert_at_row, num_rows)
            _shift_merged_cells_after_insert_rows(ws_in, insert_at_row, num_rows)
            for i, other_row_idx in enumerate(reversed(other_rows)):
                row_dst = insert_at_row + (num_rows - 1 - i)
                _copy_row_with_style(ws_o, ws_in, other_row_idx, row_dst, max_col)
                _copy_row_merged_ranges_to(ws_o, other_row_idx, ws_in, row_dst)
            inserted_count += num_rows
    return inserted_count


def _merge_mode_b_preserve_format_wb(wb_in, wb_other, base_side):
    """在已打开的基准工作簿中插入另一侧新增列，返回新增列数。"""
    inserted_count = 0
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col_in = ws_in.max_column or 1
        max_col_o = ws_o.max_column or 1
        max_col = max(max_col_in, max_col_o)
        header_in = load_sheet_header(ws_in, max_col)
        header_o = load_sheet_header(ws_o, max_col)
        header_in_norm_to_idx = _compare_header_to_index(header_in)
        base_set_norm = set(header_in_norm_to_idx)
        other_ordered = [h for h in header_o if h]
        new_cols = [h for h in other_ordered if header_normalize_for_compare(h) not in base_set_norm]
        merged_col_order = merge_ordered_with_new_cols([h for h in header_in if h], new_cols)
        col_to_idx_o = {}
        for i, h in enumerate(header_o):
            if h:
                col_to_idx_o[h] = i + 1
        last_base_col = None
        inserts = []
        for h in merged_col_order:
            idx_in = header_in_norm_to_idx.get(header_normalize_for_compare(h))
            if idx_in is not None:
                last_base_col = idx_in
            elif h in col_to_idx_o:
                insert_after = (last_base_col if last_base_col is not None else 0)
                inserts.append((insert_after, h))
        max_row = max(ws_in.max_row or 1, ws_o.max_row or 1)
        merged_cache_o = build_merged_cells_cache(ws_o) if has_merged_cells(ws_o) and inserts else None
        for insert_after, col_key in sorted(inserts, key=lambda x: -x[0]):
            col_o = col_to_idx_o[col_key]
            if insert_after >= 1:
                insert_at_col = insert_after + 1
            else:
                insert_at_col = (ws_in.max_column or 1) + 1
            ws_in.insert_cols(insert_at_col, 1)
            _shift_merged_cells_after_insert_cols(ws_in, insert_at_col, 1)
            _copy_col_with_style(ws_o, ws_in, col_o, insert_at_col, max_row, merged_cache_o)
            _copy_col_merged_ranges_to(ws_o, col_o, ws_in, insert_at_col)
            inserted_count += 1
    return inserted_count


def _merge_delete_rows_wb(wb_in, wb_other):
    """从已打开的工作簿中删除另一侧缺失的行，返回删除行数。"""
    deleted_count = 0
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col = max(ws_in.max_column or 1, ws_o.max_column or 1)
        rows_o, _ = load_sheet_rows_full(ws_o, max_col, use_cache=True)
        keys_other = set(key_str_normalized(r[0]) if r else "" for r in rows_o)
        keys_other.discard("")
        rows_in, idx_in = load_sheet_rows_full(ws_in, max_col, use_cache=True)
        to_delete = []
        for i, r in enumerate(rows_in):
            k = key_str_normalized(r[0]) if r else ""
            if k and k not in keys_other:
                to_delete.append(idx_in[i] if i < len(idx_in) else i + 1)
        for row_idx in sorted(set(to_delete), reverse=True):
            ws_in.delete_rows(row_idx, 1)
            deleted_count += 1
    return deleted_count


def _merge_delete_cols_wb(wb_in, wb_other):
    """从已打开的工作簿中删除另一侧缺失的列，返回删除列数。"""
    deleted_count = 0
    for sheet_name in list(wb_in.sheetnames):
        if sheet_name not in wb_other.sheetnames:
            continue
        ws_in = wb_in[sheet_name]
        ws_o = wb_other[sheet_name]
        max_col = max(ws_in.max_column or 1, ws_o.max_column or 1)
        header_in = load_sheet_header(ws_in, max_col)
        header_o = load_sheet_header(ws_o, max_col)
        keys_other_norm = set(header_normalize_for_compare(h) for h in header_o if h)
        to_delete = []
        for c in range(len(header_in) - 1, -1, -1):
            h = header_in[c] if c < len(header_in) else ""
            if h and header_normalize_for_compare(h) not in keys_other_norm:
                col_1based = c + 1
                if col_1based != 1:
                    to_delete.append(col_1based)
        for col_idx in sorted(to_delete, reverse=True):
            ws_in.delete_cols(col_idx, 1)
            deleted_count += 1
    return deleted_count


def _merge_mode_c_wb(wb_in, wb_other):
    """在已打开的工作簿中追加另一侧新增 Sheet，返回 Sheet 名列表。"""
    base_sheets = set(get_sheet_names(wb_in))
    new_sheets = [n for n in get_sheet_names(wb_other) if n not in base_sheets]
    for name in new_sheets:
        _copy_worksheet_to_new_sheet(wb_in, wb_other[name], name)
    return new_sheets


def _merge_delete_sheets_wb(wb_in, wb_other):
    """从已打开的工作簿中删除另一侧缺失的 Sheet，返回删除 Sheet 名列表。"""
    other_sheets = set(get_sheet_names(wb_other))
    deleted = []
    for name in list(wb_in.sheetnames):
        if name not in other_sheets:
            del wb_in[name]
            deleted.append(name)
    if not wb_in.sheetnames:
        wb_in.create_sheet("Data")
    return deleted


def _row_maps_for_actions(ws, max_col):
    rows, idx = load_sheet_rows_full(ws, max_col, use_cache=True) if ws else ([], [])
    row_by_key = {}
    row_index_by_key = {}
    seen = set()
    for i, row in enumerate(rows):
        if i >= len(idx):
            break
        raw = cell_str(row[0]) if row else ""
        key = key_str_normalized(raw) if raw else "__row_%d" % idx[i]
        if key in seen:
            key = "__row_%d" % idx[i]
        seen.add(key)
        row_by_key[key] = row
        row_index_by_key[key] = idx[i]
    return row_by_key, row_index_by_key


def _compute_auto_row_actions_from_workbooks(wb_l, wb_b, wb_r):
    """基于已打开的三方工作簿计算非冲突自动行动作。"""
    seen = set()
    sheet_names = []
    for wb in (wb_b, wb_l, wb_r):
        for name in get_sheet_names(wb):
            if name not in seen:
                seen.add(name)
                sheet_names.append(name)

    actions = []
    for sheet_name in sheet_names:
        ws_l = wb_l[sheet_name] if sheet_name in wb_l.sheetnames else None
        ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None
        ws_r = wb_r[sheet_name] if sheet_name in wb_r.sheetnames else None
        max_col = max(
            (ws_l.max_column or 1) if ws_l else 1,
            (ws_b.max_column or 1) if ws_b else 1,
            (ws_r.max_column or 1) if ws_r else 1,
        )
        dict_l, key_to_row_l = _row_maps_for_actions(ws_l, max_col) if ws_l else ({}, {})
        dict_b, key_to_row_b = _row_maps_for_actions(ws_b, max_col) if ws_b else ({}, {})
        dict_r, key_to_row_r = _row_maps_for_actions(ws_r, max_col) if ws_r else ({}, {})

        for key in set(dict_b) | set(dict_l) | set(dict_r):
            if 1 in (
                key_to_row_l.get(key),
                key_to_row_b.get(key),
                key_to_row_r.get(key),
            ):
                continue
            row_b = dict_b.get(key)
            if row_b is None:
                continue
            row_l = dict_l.get(key)
            row_r = dict_r.get(key)
            if row_l is None and row_r is None:
                continue
            if row_l is None:
                if row_equal(row_r, row_b):
                    actions.append({
                        "sheet": sheet_name,
                        "key": key,
                        "choice": "local",
                        "kind": "row",
                        "type": "delete_local",
                    })
                continue
            if row_r is None:
                if row_equal(row_l, row_b):
                    actions.append({
                        "sheet": sheet_name,
                        "key": key,
                        "choice": "remote",
                        "kind": "row",
                        "type": "delete_remote",
                    })
                continue
            if row_equal(row_l, row_r):
                continue
            local_changed = not row_equal(row_l, row_b)
            remote_changed = not row_equal(row_r, row_b)
            if local_changed and not remote_changed:
                actions.append({
                    "sheet": sheet_name,
                    "key": key,
                    "choice": "local",
                    "kind": "row",
                    "type": "take_local",
                })
            elif remote_changed and not local_changed:
                actions.append({
                    "sheet": sheet_name,
                    "key": key,
                    "choice": "remote",
                    "kind": "row",
                    "type": "take_remote",
                })
    return actions


def _shift_row_map_after_delete(key_to_row, deleted_row):
    for key, row_idx in list(key_to_row.items()):
        if row_idx == deleted_row:
            del key_to_row[key]
        elif row_idx > deleted_row:
            key_to_row[key] = row_idx - 1


def _shift_row_map_after_insert(key_to_row, insert_at_row, amount=1):
    for key, row_idx in list(key_to_row.items()):
        if row_idx >= insert_at_row:
            key_to_row[key] = row_idx + amount


def _apply_merge_choices_to_workbook(wb_out, wb_local, wb_base, wb_remote, d_choices):
    """将冲突选择和自动动作批量写入已打开的输出工作簿。"""
    if not d_choices:
        return 0
    font_mod = Font(color="CC6600")
    sheet_cache = {}
    applied = 0

    def cache_for_sheet(sheet_name):
        if sheet_name in sheet_cache:
            return sheet_cache[sheet_name]
        if sheet_name not in wb_out.sheetnames:
            sheet_cache[sheet_name] = None
            return None
        ws_out = wb_out[sheet_name]
        ws_l = wb_local[sheet_name] if sheet_name in wb_local.sheetnames else None
        ws_b = wb_base[sheet_name] if sheet_name in wb_base.sheetnames else None
        ws_r = wb_remote[sheet_name] if sheet_name in wb_remote.sheetnames else None
        max_col = ws_out.max_column or 1
        cache = {
            "ws_out": ws_out,
            "ws_l": ws_l,
            "ws_b": ws_b,
            "ws_r": ws_r,
            "row_out": _row_key_to_index(ws_out, max_col),
            "row_l": _row_key_to_index(ws_l, max_col) if ws_l else {},
            "row_b": _row_key_to_index(ws_b, max_col) if ws_b else {},
            "row_r": _row_key_to_index(ws_r, max_col) if ws_r else {},
            "header_out": None,
            "header_l": None,
            "header_r": None,
            "header_max_col": None,
        }
        sheet_cache[sheet_name] = cache
        return cache

    def header_maps(cache):
        ws_out = cache["ws_out"]
        max_col = ws_out.max_column or 1
        if cache["header_max_col"] != max_col:
            ws_l = cache["ws_l"]
            ws_r = cache["ws_r"]
            cache["header_out"] = _compare_header_to_index(load_sheet_header(ws_out, max_col))
            cache["header_l"] = _compare_header_to_index(load_sheet_header(ws_l, max_col) if ws_l else [])
            cache["header_r"] = _compare_header_to_index(load_sheet_header(ws_r, max_col) if ws_r else [])
            cache["header_max_col"] = max_col
        return cache["header_out"], cache["header_l"], cache["header_r"]

    for item in d_choices or []:
        sheet_name = item.get("sheet")
        cache = cache_for_sheet(sheet_name)
        if not cache:
            continue
        key = item.get("key")
        choice = item.get("choice", "local")
        kind = item.get("kind", "row")
        ws_out = cache["ws_out"]
        ws_l = cache["ws_l"]
        ws_b = cache["ws_b"]
        ws_r = cache["ws_r"]
        max_col = ws_out.max_column or 1
        if kind == "row":
            key_norm = key_str_normalized(key) or key
            row_idx_out = cache["row_out"].get(key_norm)
            source = ws_r if choice == "remote" else ws_l
            source_rows = cache["row_r"] if choice == "remote" else cache["row_l"]
            row_idx_source = source_rows.get(key_norm) if source else None
            if not source or row_idx_source is None:
                if row_idx_out is not None:
                    ws_out.delete_rows(row_idx_out, 1)
                    _shift_row_map_after_delete(cache["row_out"], row_idx_out)
                    applied += 1
                continue
            if row_idx_out is None:
                row_idx_out = min(row_idx_source, (ws_out.max_row or 0) + 1)
                ws_out.insert_rows(row_idx_out, 1)
                _shift_merged_cells_after_insert_rows(ws_out, row_idx_out, 1)
                _shift_row_map_after_insert(cache["row_out"], row_idx_out, 1)
                cache["row_out"][key_norm] = row_idx_out
            auto_type = item.get("type")
            copy_plan = _row_copy_plan(ws_out, source, ws_b, max_col)
            if auto_type in ("take_local", "take_remote"):
                row_idx_base = cache["row_b"].get(key_norm) if ws_b is not None else None
                for out_col, src_col, base_col in copy_plan:
                    src_c = source.cell(row=row_idx_source, column=src_col)
                    base_val = ws_b.cell(row=row_idx_base, column=base_col).value if ws_b is not None and row_idx_base else None
                    if cell_str(src_c.value) == cell_str(base_val):
                        continue
                    dst_c = ws_out.cell(row=row_idx_out, column=out_col)
                    _copy_cell_value_and_style(src_c, dst_c, font_mod)
            else:
                for out_col, src_col, _base_col in copy_plan:
                    src_c = source.cell(row=row_idx_source, column=src_col)
                    dst_c = ws_out.cell(row=row_idx_out, column=out_col)
                    _copy_cell_value_and_style(src_c, dst_c, font_mod)
            applied += 1
        else:
            header_out, header_l, header_r = header_maps(cache)
            norm_key = header_normalize_for_compare(key)
            col_idx_out = header_out.get(norm_key)
            col_idx_l = header_l.get(norm_key)
            col_idx_r = header_r.get(norm_key)
            col_idx = col_idx_l if choice == "local" else col_idx_r
            source = ws_l if choice == "local" else ws_r
            if col_idx is None or not source:
                continue
            if col_idx_out is None:
                col_idx_out = (ws_out.max_column or 0) + 1
                header_out[norm_key] = col_idx_out
            max_row = ws_out.max_row or 1
            for r in range(1, max_row + 1):
                src_c = source.cell(row=r, column=col_idx)
                dst_c = ws_out.cell(row=r, column=col_idx_out, value=src_c.value)
                _copy_cell_style(src_c, dst_c)
                dst_c.font = font_mod
            applied += 1
    return applied


def _merge_choices(d_choices, auto_actions):
    merged = []
    seen = set()
    for item in list(d_choices or []) + list(auto_actions or []):
        key = (
            item.get("sheet"),
            item.get("key"),
            item.get("kind", "row"),
        )
        if key in seen:
            continue
        seen.add(key)
        merged.append(item)
    return merged


def _do_merge_by_options(path_local, path_base, path_remote, path_merged, options, base_side, d_choices):
    options = set(options or [])
    path_base_side = path_local if base_side == "local" else path_remote
    log("[Options] 合并管道 基准=%s（%s 为底）勾选项=%s（仅当 C 勾选时执行删除行）" % (
        "本地" if base_side == "local" else "线上",
        "LOCAL" if base_side == "local" else "REMOTE",
        sorted(options) or ["无"],
    ))
    import tempfile
    fd, path_cur = tempfile.mkstemp(suffix=".xlsx", prefix="merge_opt_")
    os.close(fd)
    wb_out = None
    wb_local = None
    wb_base = None
    wb_remote = None
    try:
        shutil.copy2(path_base_side, path_cur)
        wb_out = openpyxl.load_workbook(path_cur, data_only=False)
        wb_local = openpyxl.load_workbook(path_local, data_only=False)
        wb_base = openpyxl.load_workbook(path_base, data_only=False)
        wb_remote = openpyxl.load_workbook(path_remote, data_only=False)
        wb_other = wb_remote if base_side == "local" else wb_local

        if "A" not in options:
            inserted_rows = _merge_mode_a_preserve_format_wb(wb_out, wb_other, base_side)
            log("[Option A] 新增行插入（保留格式）完成 rows=%d" % inserted_rows)
        if "C" in options:
            deleted_rows = _merge_delete_rows_wb(wb_out, wb_other)
            log("[Option C] 删除行完成 rows=%d" % deleted_rows)
        else:
            log("[Options] 跳过删除行（C 未勾选，基准侧独有行将保留）")
        if "B" not in options:
            inserted_cols = _merge_mode_b_preserve_format_wb(wb_out, wb_other, base_side)
            log("[Option B] 新增列插入（保留格式）完成 cols=%d" % inserted_cols)
        if "D" in options:
            deleted_cols = _merge_delete_cols_wb(wb_out, wb_other)
            log("[Option D] 删除列完成 cols=%d" % deleted_cols)
        if "E" in options:
            new_sheets = _merge_mode_c_wb(wb_out, wb_other)
            if new_sheets:
                log("[Mode C] 新增 Sheet 插入完成 new_sheets=%s" % new_sheets)
            else:
                log("[Mode C] 无新增 Sheet，保留当前工作簿")
        if "F" in options:
            deleted_sheets = _merge_delete_sheets_wb(wb_out, wb_other)
            log("[Option F] 删除 Sheet 完成 sheets=%s" % deleted_sheets)
        auto_actions = _compute_auto_row_actions_from_workbooks(wb_local, wb_base, wb_remote)
        all_choices = _merge_choices(d_choices if "G" in options else [], auto_actions)
        if all_choices:
            applied = _apply_merge_choices_to_workbook(wb_out, wb_local, wb_base, wb_remote, all_choices)
            log("[Mode D] 冲突/自动选择写回完成 actions=%d applied=%d" % (len(all_choices), applied))
        if not wb_out.sheetnames:
            wb_out.create_sheet("Data")
        os.makedirs(os.path.dirname(os.path.abspath(path_merged)) or ".", exist_ok=True)
        wb_out.save(path_merged)
        log("[Options] 管道完成 MERGED=%s options=%s auto_actions=%d" % (path_merged, options, len(auto_actions)))
        _log_merged_sheet_keys(path_merged, base_side, options)
    finally:
        for wb in (wb_out, wb_local, wb_base, wb_remote):
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
        try:
            os.unlink(path_cur)
        except Exception:
            pass


# ---------------------------------------------------------------------------
# 统一入口与备份
# ---------------------------------------------------------------------------

def do_merge(path_local, path_base, path_remote, path_merged, mode="E", base_side="local", d_choices=None, options=None, backup_root=None, backup_context_path=None):
    """
    执行合并并写入 MERGED，并备份到 指定根目录/项目/时间。
    options: 若提供则为多选集合 {"A","B","C","D","E","F","G"}，此时忽略 mode。
    A=行不变 B=列不变 C=删除行 D=删除列 E=新增Sheet F=删除Sheet G=冲突。
    mode: 未提供 options 时生效，"A"|"B"|"C"|"D"|"E"。
    base_side: "local"|"remote"
    d_choices: G 或冲突时需要；list of {"sheet", "key", "choice", "kind"}。
    backup_root: 可选备份根目录；未提供时读取本地配置，仍未设置则使用 MERGED 同目录下 MergeExcelBackup。
    backup_context_path: 可选逻辑目标路径；driver 模式可用真实仓库路径决定备份目录和文件名。
    返回 0 成功，2 异常；成功后 do_merge.last_backup_info 保存本次备份信息。
    """
    do_merge.last_backup_info = None
    path_base_side = path_local if base_side == "local" else path_remote
    path_other_side = path_remote if base_side == "local" else path_local
    try:
        if options is not None:
            _do_merge_by_options(path_local, path_base, path_remote, path_merged, options, base_side, d_choices or [])
        elif mode == "A":
            _merge_mode_a_impl(path_base_side, path_other_side, path_merged, base_side)
        elif mode == "B":
            _merge_mode_b_impl(path_base_side, path_other_side, path_merged, base_side)
        elif mode == "C":
            _merge_mode_c_impl(path_base_side, path_other_side, path_merged, base_side)
        elif mode == "D":
            auto_actions = compute_auto_row_actions(path_local, path_base, path_remote)
            _merge_mode_d_impl(path_local, path_remote, path_merged, path_base, _merge_choices(d_choices or [], auto_actions))
        elif mode == "E":
            _do_merge_by_options(
                path_local, path_base, path_remote, path_merged,
                options={"E", "G"}, base_side=base_side, d_choices=d_choices or [],
            )
        else:
            log("未知 mode=%s，回退到 A" % mode)
            _merge_mode_a_impl(path_base_side, path_other_side, path_merged, base_side)
    except Exception as e:
        log("合并异常: %s" % e, is_error=True)
        import traceback
        log(traceback.format_exc(), is_error=True)
        print("ERROR: " + str(e), file=sys.stderr)
        return 2

    try:
        backup_info = create_merge_backup(
            path_local, path_remote, path_merged,
            backup_root=backup_root, context_path=backup_context_path,
        )
        do_merge.last_backup_info = backup_info
        backup_dir = backup_info["dir"]
    except Exception as e:
        log("备份异常: %s" % e, is_error=True)
        print("ERROR: 备份失败。" + str(e), file=sys.stderr)
        return 2
    log("合并完成 MERGED=%s 备份=%s" % (path_merged, backup_dir))
    print("OK: 合并完成。MERGED=%s 备份=%s" % (path_merged, backup_dir), file=sys.stdout)
    return 0


do_merge.last_backup_info = None
