# -*- coding: utf-8 -*-
"""
冲突检测：三份 Excel（LOCAL、BASE、REMOTE）逐 Sheet 比较，得到“需选择”的冲突项与“仅本地有/仅线上有”项。
只做一件事：给定三个路径，返回 (conflicts, sheet_data, sheet_names)，供合并 GUI 展示与写回；
使用 data_only=False 加载，避免公式未计算导致误判相同。不包含合并写回逻辑。
"""

import openpyxl

from excel_io import (
    build_merged_cells_cache,
    cell_str,
    col_equal,
    get_column_values,
    get_sheet_names,
    has_merged_cells,
    key_str_normalized,
    load_sheet_header,
    load_sheet_rows_full,
    row_equal,
)
from log_util import log_path


def _normalized_header_to_index(headers):
    out = {}
    for i, h in enumerate(headers):
        key = key_str_normalized(h or "")
        if key and key not in out:
            out[key] = i + 1
    return out


def _dict_and_order(rows, row_indices):
    """
    将 (rows, row_indices) 转为 dict、key->行号、顺序列表。
    首列空或 key 重复时使用 __row_N 作为 key，保证每行唯一且可写回时定位行号。
    """
    d = {}
    key_to_row = {}
    ord_list = []
    seen = set()
    for i, r in enumerate(rows):
        if i >= len(row_indices):
            break
        raw = cell_str(r[0]) if r else ""
        if not raw:
            k = "__row_%d" % row_indices[i]
        else:
            k = key_str_normalized(raw)
        if k in seen:
            k = "__row_%d" % row_indices[i]
        seen.add(k)
        d[k] = r
        key_to_row[k] = row_indices[i]
        ord_list.append(k)
    return d, key_to_row, ord_list


def _log_merge_diagnostic(sheet_name, n_l, n_b, n_r, n_keys, n_conflict):
    """写入单 Sheet 合并诊断到日志，便于排查“0 冲突”等问题。"""
    try:
        from datetime import datetime
        with open(log_path(), "a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write("%s [MERGE] Sheet=%s 行数 local=%d base=%d remote=%d all_keys=%d 冲突=%d\n" % (
                ts, sheet_name, n_l, n_b, n_r, n_keys, n_conflict))
    except Exception:
        pass


def compute_conflicts(path_local, path_base, path_remote, include_sheet_data=True):
    """
    计算三向合并的冲突列表与每 Sheet 的数据结构。
    使用 data_only=False，避免公式单元格缓存为空导致误判。
    返回:
      - conflicts: list of dict，每项含 sheet, key, local_row, remote_row, base_row，
        以及可选 _only_local / _only_remote（仅一方有的行）；
      - sheet_data: dict[sheet_name] = { base_rows, local_rows, remote_rows, base_ordered, local_ordered, remote_ordered, key_to_row_l, key_to_row_r, key_to_row_b, max_col }；
        include_sheet_data=False 时返回空 dict，适合 GUI 只展示冲突列表的大文件路径；
      - sheet_names: list of str。
    """
    wb_l = openpyxl.load_workbook(path_local, data_only=False)
    wb_b = openpyxl.load_workbook(path_base, data_only=False)
    wb_r = openpyxl.load_workbook(path_remote, data_only=False)

    seen = set()
    sheet_names = []
    for n in get_sheet_names(wb_b):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in get_sheet_names(wb_l):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in get_sheet_names(wb_r):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)

    conflicts = []
    sheet_data = {}

    for sheet_name in sheet_names:
        ws_l = wb_l[sheet_name] if sheet_name in wb_l.sheetnames else None
        ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None
        ws_r = wb_r[sheet_name] if sheet_name in wb_r.sheetnames else None

        max_col_sheet = max(
            (ws_l.max_column or 1) if ws_l else 1,
            (ws_b.max_column or 1) if ws_b else 1,
            (ws_r.max_column or 1) if ws_r else 1,
        )
        rows_l, idx_l = load_sheet_rows_full(ws_l, max_col_sheet, use_cache=True) if ws_l else ([], [])
        rows_b, idx_b = load_sheet_rows_full(ws_b, max_col_sheet, use_cache=True) if ws_b else ([], [])
        rows_r, idx_r = load_sheet_rows_full(ws_r, max_col_sheet, use_cache=True) if ws_r else ([], [])

        dict_l, key_to_row_l, ord_l = _dict_and_order(rows_l, idx_l)
        dict_b, key_to_row_b, ord_b = _dict_and_order(rows_b, idx_b)
        dict_r, key_to_row_r, ord_r = _dict_and_order(rows_r, idx_r)

        base_set = set(dict_b)
        local_set = set(dict_l)
        remote_set = set(dict_r)
        all_keys = base_set | local_set | remote_set

        n_conflict_sheet = 0
        for key in all_keys:
            row_l = dict_l.get(key)
            row_r = dict_r.get(key)
            row_b = dict_b.get(key)
            
            # 场景1：BASE无 - 新增相关
            if row_b is None:
                if row_l is not None and row_r is not None:
                    # 双方都新增
                    if row_equal(row_l, row_r):
                        # 新增相同内容，无冲突
                        continue
                    else:
                        # 新增不同内容，冲突
                        n_conflict_sheet += 1
                        conflicts.append({
                            "sheet": sheet_name,
                            "key": key,
                            "type": "add_conflict",
                            "local_row": row_l,
                            "remote_row": row_r,
                            "base_row": None,
                        })
                elif row_l is not None:
                    # 仅本地新增
                    conflicts.append({
                        "sheet": sheet_name,
                        "key": key,
                        "type": "add_local",
                        "local_row": row_l,
                        "remote_row": None,
                        "base_row": None,
                        "_only_local": True,
                    })
                elif row_r is not None:
                    # 仅线上新增
                    conflicts.append({
                        "sheet": sheet_name,
                        "key": key,
                        "type": "add_remote",
                        "local_row": None,
                        "remote_row": row_r,
                        "base_row": None,
                        "_only_remote": True,
                    })
            
            # 场景2：BASE有，至少一方删除
            elif row_l is None or row_r is None:
                if row_l is None and row_r is None:
                    # 双方都删除，无冲突
                    continue
                elif row_l is None:
                    # 本地删除，线上保留/修改 → 删除冲突
                    n_conflict_sheet += 1
                    conflicts.append({
                        "sheet": sheet_name,
                        "key": key,
                        "type": "delete_conflict_local",
                        "local_row": None,
                        "remote_row": row_r,
                        "base_row": row_b,
                        "_delete_conflict": True,
                    })
                else:
                    # 线上删除，本地保留/修改 → 删除冲突
                    n_conflict_sheet += 1
                    conflicts.append({
                        "sheet": sheet_name,
                        "key": key,
                        "type": "delete_conflict_remote",
                        "local_row": row_l,
                        "remote_row": None,
                        "base_row": row_b,
                        "_delete_conflict": True,
                    })
            
            # 场景3：BASE有，双方都有
            else:
                if row_equal(row_l, row_r):
                    # 双方相同，无冲突
                    continue
                # 双方不同，检查是否为冲突
                if not row_equal(row_b, row_l) and not row_equal(row_b, row_r):
                    # 双方都修改且不同 → 修改冲突
                    n_conflict_sheet += 1
                    conflicts.append({
                        "sheet": sheet_name,
                        "key": key,
                        "type": "modify_conflict",
                        "local_row": row_l,
                        "remote_row": row_r,
                        "base_row": row_b,
                    })

        if include_sheet_data:
            sheet_data[sheet_name] = {
                "base_rows": dict_b,
                "local_rows": dict_l,
                "remote_rows": dict_r,
                "base_ordered": ord_b,
                "local_ordered": ord_l,
                "remote_ordered": ord_r,
                "key_to_row_l": key_to_row_l,
                "key_to_row_r": key_to_row_r,
                "key_to_row_b": key_to_row_b,
                "max_col": max(
                    max(len(r) for r in rows_l) if rows_l else 1,
                    max(len(r) for r in rows_b) if rows_b else 1,
                    max(len(r) for r in rows_r) if rows_r else 1,
                ),
            }
        _log_merge_diagnostic(sheet_name, len(rows_l), len(rows_b), len(rows_r), len(all_keys), n_conflict_sheet)

    wb_l.close()
    wb_b.close()
    wb_r.close()
    return conflicts, sheet_data, sheet_names


def compute_auto_row_actions(path_local, path_base, path_remote):
    """
    计算 BASE 中心的非冲突行级自动动作。
    返回 list，每项含 sheet/key/choice/type/kind，可直接转为 D 模式选择：
      - take_local / take_remote: 单侧修改，自动采用修改侧
      - delete_local / delete_remote: 单侧删除且另一侧未改，自动删除
    """
    wb_l = openpyxl.load_workbook(path_local, data_only=False)
    wb_b = openpyxl.load_workbook(path_base, data_only=False)
    wb_r = openpyxl.load_workbook(path_remote, data_only=False)
    try:
        seen = set()
        sheet_names = []
        for wb in (wb_b, wb_l, wb_r):
            for name in get_sheet_names(wb):
                if name not in seen:
                    seen.add(name)
                    sheet_names.append(name)

        actions = []
        for sheet_name in sheet_names:
            ws_l = wb_l[sheet_name] if sheet_name in wb_l.sheetnames else None
            ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None
            ws_r = wb_r[sheet_name] if sheet_name in wb_r.sheetnames else None
            max_col = max(
                (ws_l.max_column or 1) if ws_l else 1,
                (ws_b.max_column or 1) if ws_b else 1,
                (ws_r.max_column or 1) if ws_r else 1,
            )
            rows_l, idx_l = load_sheet_rows_full(ws_l, max_col, use_cache=True) if ws_l else ([], [])
            rows_b, idx_b = load_sheet_rows_full(ws_b, max_col, use_cache=True) if ws_b else ([], [])
            rows_r, idx_r = load_sheet_rows_full(ws_r, max_col, use_cache=True) if ws_r else ([], [])
            dict_l, key_to_row_l, _ = _dict_and_order(rows_l, idx_l)
            dict_b, key_to_row_b, _ = _dict_and_order(rows_b, idx_b)
            dict_r, key_to_row_r, _ = _dict_and_order(rows_r, idx_r)

            for key in set(dict_b) | set(dict_l) | set(dict_r):
                if 1 in (
                    key_to_row_l.get(key),
                    key_to_row_b.get(key),
                    key_to_row_r.get(key),
                ):
                    continue
                row_b = dict_b.get(key)
                if row_b is None:
                    continue
                row_l = dict_l.get(key)
                row_r = dict_r.get(key)
                if row_l is None and row_r is None:
                    continue
                if row_l is None:
                    if row_equal(row_r, row_b):
                        actions.append({
                            "sheet": sheet_name,
                            "key": key,
                            "choice": "local",
                            "kind": "row",
                            "type": "delete_local",
                        })
                    continue
                if row_r is None:
                    if row_equal(row_l, row_b):
                        actions.append({
                            "sheet": sheet_name,
                            "key": key,
                            "choice": "remote",
                            "kind": "row",
                            "type": "delete_remote",
                        })
                    continue
                if row_equal(row_l, row_r):
                    continue
                local_changed = not row_equal(row_l, row_b)
                remote_changed = not row_equal(row_r, row_b)
                if local_changed and not remote_changed:
                    actions.append({
                        "sheet": sheet_name,
                        "key": key,
                        "choice": "local",
                        "kind": "row",
                        "type": "take_local",
                    })
                elif remote_changed and not local_changed:
                    actions.append({
                        "sheet": sheet_name,
                        "key": key,
                        "choice": "remote",
                        "kind": "row",
                        "type": "take_remote",
                    })
        return actions
    finally:
        wb_l.close()
        wb_b.close()
        wb_r.close()


def compute_conflicts_d(path_local, path_remote, path_base=None):
    """
    D 模式：二向冲突检测（冲突行 + 冲突列）。
    两边都有且内容不同即为冲突。path_base 可选，暂不用于过滤。
    返回 (conflict_rows, conflict_cols, sheet_names)。
    - conflict_rows: list of {"sheet", "key", "local_row", "remote_row", "kind": "row"}
    - conflict_cols: list of {"sheet", "key", "local_col", "remote_col", "kind": "column"}，key 为表头（列名）
    """
    wb_l = openpyxl.load_workbook(path_local, data_only=False)
    wb_r = openpyxl.load_workbook(path_remote, data_only=False)

    seen = set()
    sheet_names = []
    for n in get_sheet_names(wb_l):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in get_sheet_names(wb_r):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)

    conflict_rows = []
    conflict_cols = []

    for sheet_name in sheet_names:
        ws_l = wb_l[sheet_name] if sheet_name in wb_l.sheetnames else None
        ws_r = wb_r[sheet_name] if sheet_name in wb_r.sheetnames else None
        if not ws_l or not ws_r:
            continue
        max_col = max(ws_l.max_column or 1, ws_r.max_column or 1)
        max_row = max(ws_l.max_row or 1, ws_r.max_row or 1)

        # 使用缓存加速合并单元格查找
        rows_l, idx_l = load_sheet_rows_full(ws_l, max_col, use_cache=True)
        rows_r, idx_r = load_sheet_rows_full(ws_r, max_col, use_cache=True)
        dict_l, _, ord_l = _dict_and_order(rows_l, idx_l)
        dict_r, _, ord_r = _dict_and_order(rows_r, idx_r)
        common_keys = set(dict_l) & set(dict_r)
        for key in common_keys:
            if not row_equal(dict_l[key], dict_r[key]):
                conflict_rows.append({
                    "sheet": sheet_name,
                    "key": key,
                    "local_row": dict_l[key],
                    "remote_row": dict_r[key],
                    "kind": "row",
                })

        header_l = load_sheet_header(ws_l, max_col)
        header_r = load_sheet_header(ws_r, max_col)
        norm_to_col_l = _normalized_header_to_index(header_l[:max_col])
        norm_to_col_r = _normalized_header_to_index(header_r[:max_col])
        common_headers = set(norm_to_col_l) & set(norm_to_col_r)
        merged_cache_l = build_merged_cells_cache(ws_l) if has_merged_cells(ws_l) else None
        merged_cache_r = build_merged_cells_cache(ws_r) if has_merged_cells(ws_r) else None
        for h_norm in common_headers:
            col_l = get_column_values(
                ws_l, norm_to_col_l[h_norm], max_row, merged_cache=merged_cache_l,
            )
            col_r = get_column_values(
                ws_r, norm_to_col_r[h_norm], max_row, merged_cache=merged_cache_r,
            )
            if not col_equal(col_l, col_r):
                conflict_cols.append({
                    "sheet": sheet_name,
                    "key": header_l[norm_to_col_l[h_norm] - 1] or header_r[norm_to_col_r[h_norm] - 1],
                    "local_col": col_l,
                    "remote_col": col_r,
                    "kind": "column",
                })

    wb_l.close()
    wb_r.close()
    return conflict_rows, conflict_cols, sheet_names
