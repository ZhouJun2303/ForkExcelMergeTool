# -*- coding: utf-8 -*-
"""
二向对比核心（A vs B → 对比表 Excel）。
只做一件事：给定两个 Excel 路径，计算差异、生成对比 Excel（含颜色标记），可选打开文件。
"""

import os
import argparse
import subprocess
import sys

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill

from config import COMPARE_SUFFIX
from excel_io import (
    cell_str,
    get_sheet_names,
    header_normalize_for_compare,
    load_sheet_header,
    load_sheet_rows,
    rows_to_dict,
)
from log_util import log


def get_compare_data(path_a, path_b, include_same=False):
    """
    计算 A、B 两个 Excel 的差异。
    返回 (path_out, sheet_names, diff_rows)：
      - path_out: 将生成的对比文件路径（与 path_a 同目录，文件名为 {base}_compare.xlsx）
      - sheet_names: 参与对比的 Sheet 名列表
      - diff_rows: [(sheet_name, key, status, str_a, str_b), ...]，
        status 为 "新增行"|"删除行"|"新增列"|"删除列"|"修改"|"相同"。
        include_same=False 时跳过相同行，适合 GUI 大文件预览。
    """
    out_dir = os.path.dirname(os.path.abspath(path_a))
    base_name = os.path.splitext(os.path.basename(path_a))[0]
    path_out = os.path.join(out_dir, base_name + COMPARE_SUFFIX + ".xlsx")

    wb_a = openpyxl.load_workbook(path_a, data_only=False, read_only=True)
    wb_b = openpyxl.load_workbook(path_b, data_only=False, read_only=True)

    seen = set()
    sheet_names = []
    for n in get_sheet_names(wb_a):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in get_sheet_names(wb_b):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    if not sheet_names:
        for wb in (wb_a, wb_b):
            for n in (wb.sheetnames or []):
                sheet_names.append(n)
                break
            if sheet_names:
                break
        if not sheet_names:
            sheet_names = ["Sheet1"]

    diff_rows = []
    for sheet_name in sheet_names:
        ws_a = wb_a[sheet_name] if sheet_name in wb_a.sheetnames else None
        ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None

        # 列差异（表头）：新增列 = B 有 A 无，删除列 = A 有 B 无
        max_col = max(
            (ws_a.max_column or 1) if ws_a else 1,
            (ws_b.max_column or 1) if ws_b else 1,
        )
        header_a = load_sheet_header(ws_a, max_col) if ws_a else []
        header_b = load_sheet_header(ws_b, max_col) if ws_b else []
        norm_a = set(header_normalize_for_compare(h) for h in header_a if h)
        norm_b = set(header_normalize_for_compare(h) for h in header_b if h)
        new_cols = [h for h in header_b if h and header_normalize_for_compare(h) not in norm_a]
        del_cols = [h for h in header_a if h and header_normalize_for_compare(h) not in norm_b]
        if new_cols:
            diff_rows.append((sheet_name, "[新增列]", "新增列", "", " | ".join(new_cols)))
        if del_cols:
            diff_rows.append((sheet_name, "[删除列]", "删除列", " | ".join(del_cols), ""))

        # 行差异
        rows_a = load_sheet_rows(ws_a) if ws_a else []
        rows_b = load_sheet_rows(ws_b) if ws_b else []
        dict_a = rows_to_dict(rows_a)
        dict_b = rows_to_dict(rows_b)
        all_keys = sorted(set(dict_a) | set(dict_b))
        max_col = max(
            max(len(r) for r in rows_a) if rows_a else 0,
            max(len(r) for r in rows_b) if rows_b else 0,
            1,
        )
        for key in all_keys:
            row_a = dict_a.get(key)
            row_b = dict_b.get(key)
            vals_a = [cell_str(c) for c in (row_a or [])]
            vals_b = [cell_str(c) for c in (row_b or [])]
            while len(vals_a) < max_col:
                vals_a.append("")
            while len(vals_b) < max_col:
                vals_b.append("")
            if row_a is None:
                status = "新增行"
            elif row_b is None:
                status = "删除行"
            elif vals_a != vals_b:
                status = "修改"
            else:
                status = "相同"
            if status == "相同" and not include_same:
                continue
            str_a = " | ".join(vals_a)
            str_b = " | ".join(vals_b)
            diff_rows.append((sheet_name, key, status, str_a, str_b))

    wb_a.close()
    wb_b.close()
    return path_out, sheet_names, diff_rows


def write_compare_excel(path_out, sheet_names, diff_rows, open_file=False):
    """根据 diff 数据写入对比 Excel（带颜色），可选用系统默认程序打开。"""
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    wb_out = openpyxl.Workbook(write_only=True)
    sheets = {}
    for sheet_name in sheet_names:
        ws_out = wb_out.create_sheet(sheet_name[:31])
        ws_out.append(["[Key]", "[A-LEFT]", "[B-RIGHT]", "[Status]"])
        sheets[sheet_name] = ws_out

    for sheet_name, key, status, str_a, str_b in diff_rows:
        ws_out = sheets[sheet_name]
        fill = green_fill if status in ("新增行", "新增列") else (red_fill if status in ("删除行", "删除列") else (yellow_fill if status == "修改" else None))
        cells = []
        for value in (key, str_a, str_b, status):
            cell = WriteOnlyCell(ws_out, value=value)
            if fill:
                cell.fill = fill
            cells.append(cell)
        ws_out.append(cells)

    wb_out.save(path_out)
    wb_out.close()

    if open_file:
        try:
            if sys.platform == "win32":
                os.startfile(path_out)
            elif sys.platform == "darwin":
                subprocess.run(["open", path_out], check=False)
            else:
                subprocess.run(["xdg-open", path_out], check=False)
        except Exception:
            pass
    return 0


def do_compare(path_a, path_b, open_file=True, include_same=False):
    """
    二向对比：生成对比 Excel 并可选打开。
    返回 0 成功，2 异常（如缺 openpyxl 或 get_compare_data 失败）。
    """
    result = get_compare_data(path_a, path_b, include_same=include_same)
    if result[0] is None:
        return 2
    path_out, sheet_names, diff_rows = result
    log("对比模式 输出: %s" % path_out)
    write_compare_excel(path_out, sheet_names, diff_rows, open_file=open_file)

    print("OK: 对比已生成 %s" % path_out, file=sys.stdout)
    return 0


def main(argv=None):
    parser = argparse.ArgumentParser(description="Compare two Excel files.")
    parser.add_argument("path_a")
    parser.add_argument("path_b")
    parser.add_argument("--include-same", action="store_true", help="输出相同行（默认只输出差异）")
    args = parser.parse_args(argv)
    return do_compare(args.path_a, args.path_b, include_same=args.include_same)


if __name__ == "__main__":
    sys.exit(main())
