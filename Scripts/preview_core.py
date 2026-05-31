# -*- coding: utf-8 -*-
"""
合并预览数据构造。
只做一件事：给 GUI 提供可展示的合并预览列表，不触碰任何 Tk 控件。
"""

import os
import time

import openpyxl

from conflict import compute_conflicts
from conflict import compute_auto_row_actions
from excel_io import (
    get_sheet_names,
    header_normalize_for_compare,
    key_str_normalized,
    load_sheet_header,
    load_sheet_rows_full,
    ordered_keys_normalized,
)
from log_util import log


def build_merge_preview(path_local, path_base_file, path_remote, options, base_side):
    """
    构建合并预览数据。

    返回 dict:
      - items: [(sheet, key_or_desc, action, tag), ...]
      - conflict_entries: GUI 选择冲突时需要的原始数据
      - summary: {new, delete, conflict, info}
      - base_side/options: 本次预览参数
      - elapsed_ms/sheet_count: 诊断信息
    """
    started = time.time()
    options = set(options)
    path_base_side = path_local if base_side == "local" else path_remote
    path_other_side = path_remote if base_side == "local" else path_local

    items_to_insert = []
    conflict_entries = []
    summary = {"new": 0, "delete": 0, "conflict": 0, "info": 0}

    row_col_data = _load_preview_row_col_data(options, path_base_side, path_other_side)
    sheet_data_base = row_col_data["base_data"]
    sheet_data_other = row_col_data["other_data"]
    base_sheets = row_col_data["base_sheets"]
    other_sheets = row_col_data["other_sheets"]
    other_sheet_set = row_col_data["other_sheet_set"]
    sheet_names_common = row_col_data["common_sheets"]
    wb_base_sheetnames = row_col_data["base_sheetnames"]

    if "A" not in options:
        for sheet_name in sheet_names_common:
            data_b = sheet_data_base[sheet_name]
            data_o = sheet_data_other[sheet_name]
            base_keys = data_b["keys"] - {""}
            for k in data_o["keys_ordered"]:
                if k and k not in base_keys:
                    items_to_insert.append((sheet_name, k, "将新增行", "new"))
                    summary["new"] += 1

    if "C" in options:
        for sheet_name in sheet_names_common:
            data_b = sheet_data_base[sheet_name]
            data_o = sheet_data_other[sheet_name]
            base_keys = data_b["keys"] - {""}
            other_keys = data_o["keys"] - {""}
            for k in data_b["keys_ordered"]:
                if k and k in base_keys and k not in other_keys:
                    items_to_insert.append((sheet_name, k, "将删除行", "del"))
                    summary["delete"] += 1

    if "B" not in options:
        for sheet_name in sheet_names_common:
            data_b = sheet_data_base[sheet_name]
            data_o = sheet_data_other[sheet_name]
            header_b_norm = set(header_normalize_for_compare(h) for h in data_b["header"] if h)
            for h in data_o["header"]:
                if h and header_normalize_for_compare(h) not in header_b_norm:
                    items_to_insert.append((sheet_name, h, "将新增列", "new"))
                    summary["new"] += 1

    if "D" in options:
        for sheet_name in sheet_names_common:
            data_b = sheet_data_base[sheet_name]
            data_o = sheet_data_other[sheet_name]
            header_o_norm = set(header_normalize_for_compare(h) for h in data_o["header"] if h)
            seen_headers = set()
            for h in data_b["header"]:
                norm = header_normalize_for_compare(h)
                if h and norm not in seen_headers and norm not in header_o_norm:
                    seen_headers.add(norm)
                    items_to_insert.append((sheet_name, h, "将删除列", "del"))
                    summary["delete"] += 1

    if "E" in options:
        for name in other_sheets:
            if name not in base_sheets:
                items_to_insert.append((name, "新增 Sheet", "将追加", "new"))
                summary["new"] += 1

    if "F" in options:
        for name in wb_base_sheetnames:
            if name not in other_sheet_set:
                items_to_insert.append((name, "删除 Sheet", "将删除", "del"))
                summary["delete"] += 1

    if "G" in options:
        _append_conflict_preview_items(
            path_local, path_base_file, path_remote, items_to_insert, conflict_entries, summary
        )
        _append_auto_action_preview_items(
            path_local, path_base_file, path_remote, items_to_insert, summary
        )

    elapsed_ms = int((time.time() - started) * 1000)
    result = {
        "items": items_to_insert,
        "conflict_entries": conflict_entries,
        "summary": summary,
        "base_side": base_side,
        "options": tuple(sorted(options)),
        "elapsed_ms": elapsed_ms,
        "sheet_count": len(set(base_sheets) | other_sheet_set | set(wb_base_sheetnames)),
    }
    log("合并预览完成: base=%s options=%s sheets=%d items=%d elapsed=%dms" % (
        base_side,
        ",".join(sorted(options)) or "none",
        result["sheet_count"],
        len(items_to_insert),
        elapsed_ms,
    ))
    return result


def _load_preview_row_col_data(options, path_base_side, path_other_side):
    sheet_data_base = {}
    sheet_data_other = {}
    need_row_col_data = ("A" not in options) or ("B" not in options) or ("C" in options) or ("D" in options)
    need_sheet_names = need_row_col_data or ("E" in options) or ("F" in options)
    if not need_sheet_names:
        return {
            "base_data": sheet_data_base,
            "other_data": sheet_data_other,
            "base_sheets": set(),
            "other_sheets": [],
            "other_sheet_set": set(),
            "common_sheets": [],
            "base_sheetnames": [],
        }

    wb_base = None
    wb_other = None
    base_sheets = set()
    other_sheets = []
    other_sheet_set = set()
    sheet_names_common = []
    wb_base_sheetnames = []
    try:
        wb_base = openpyxl.load_workbook(path_base_side, data_only=True, read_only=True)
        wb_other = openpyxl.load_workbook(path_other_side, data_only=True, read_only=True)
        base_sheets = set(get_sheet_names(wb_base))
        other_sheets = get_sheet_names(wb_other)
        other_sheet_set = set(other_sheets)
        sheet_names_common = [n for n in get_sheet_names(wb_base) if n in wb_other.sheetnames]
        wb_base_sheetnames = list(wb_base.sheetnames)

        if need_row_col_data:
            for sheet_name in sheet_names_common:
                ws_b = wb_base[sheet_name]
                ws_o = wb_other[sheet_name]
                max_col = max(ws_b.max_column or 1, ws_o.max_column or 1)
                rows_b, _ = load_sheet_rows_full(ws_b, max_col, use_cache=True)
                rows_o, _ = load_sheet_rows_full(ws_o, max_col, use_cache=True)
                sheet_data_base[sheet_name] = {
                    "keys": set(key_str_normalized(r[0]) if r else "" for r in rows_b),
                    "keys_ordered": ordered_keys_normalized(rows_b),
                    "header": load_sheet_header(ws_b, max_col),
                }
                sheet_data_other[sheet_name] = {
                    "keys": set(key_str_normalized(r[0]) if r else "" for r in rows_o),
                    "keys_ordered": ordered_keys_normalized(rows_o),
                    "header": load_sheet_header(ws_o, max_col),
                }
    finally:
        if wb_base is not None:
            try:
                wb_base.close()
            except Exception:
                pass
        if wb_other is not None:
            try:
                wb_other.close()
            except Exception:
                pass

    return {
        "base_data": sheet_data_base,
        "other_data": sheet_data_other,
        "base_sheets": base_sheets,
        "other_sheets": other_sheets,
        "other_sheet_set": other_sheet_set,
        "common_sheets": sheet_names_common,
        "base_sheetnames": wb_base_sheetnames,
    }


def _append_conflict_preview_items(path_local, path_base_file, path_remote, items_to_insert, conflict_entries, summary):
    conflicts, _, _ = compute_conflicts(
        path_local, path_base_file, path_remote, include_sheet_data=False,
    )
    type_display = {
        "add_local": ("仅本地新增", "new", False),
        "add_remote": ("仅线上新增", "new", False),
        "add_conflict": ("新增冲突", "conflict", True),
        "delete_conflict_local": ("删除冲突：本地删", "del_conflict", True),
        "delete_conflict_remote": ("删除冲突：线上删", "del_conflict", True),
        "modify_conflict": ("修改冲突", "conflict", True),
    }
    for c in conflicts:
        conflict_type = c.get("type", "modify_conflict")
        suffix, tag, need_choice = type_display.get(conflict_type, ("冲突", "conflict", True))
        if need_choice:
            idx = len(conflict_entries)
            default_choice = "线上" if conflict_type == "delete_conflict_local" else "本地"
            choice_text = "将保留%s" % default_choice
            if conflict_type == "delete_conflict_local":
                choice_text = "将保留线上（本地已删）"
            elif conflict_type == "delete_conflict_remote":
                choice_text = "将保留本地（线上已删）"
            conflict_entries.append({
                "choice": default_choice,
                "data": c,
                "kind": "row",
                "display": choice_text,
            })
            items_to_insert.append((c["sheet"], "%s (%s)" % (c["key"], suffix), choice_text, (str(idx), tag)))
            summary["conflict"] += 1
        else:
            choice_text = "信息：本地新增" if conflict_type == "add_local" else "信息：线上新增"
            items_to_insert.append((c["sheet"], "%s (%s)" % (c["key"], suffix), choice_text, tag))
            summary["info"] += 1


def _append_auto_action_preview_items(path_local, path_base_file, path_remote, items_to_insert, summary):
    labels = {
        "take_local": ("自动采用本地修改", "mod"),
        "take_remote": ("自动采用线上修改", "mod"),
        "delete_local": ("自动删除（本地已删，线上未改）", "del"),
        "delete_remote": ("自动删除（线上已删，本地未改）", "del"),
    }
    seen = set((item[0], str(item[1]).split(" (", 1)[0], item[2]) for item in items_to_insert)
    for action in compute_auto_row_actions(path_local, path_base_file, path_remote):
        label, tag = labels.get(action.get("type"), ("自动合并", "info"))
        key = (action.get("sheet"), action.get("key"), label)
        if key in seen:
            continue
        seen.add(key)
        items_to_insert.append((action.get("sheet"), action.get("key"), label, tag))
        if tag == "del":
            summary["delete"] += 1
        else:
            summary["info"] += 1
