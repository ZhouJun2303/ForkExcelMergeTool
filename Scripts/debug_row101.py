# -*- coding: utf-8 -*-
"""诊断 Data_HeroGEarSkin 表第 101 行附近：对比 local / remote / merged。"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from excel_io import (
    get_sheet_names,
    key_str_normalized,
    load_sheet_rows_full,
    ordered_keys_normalized,
    merge_ordered_with_new_rows,
)

def main():
    base_dir = r"d:\MyGit\ExcelMergeDiffTools_B\MergeExcelBackup"
    name = "Y英雄齿轮养成表"
    path_local = os.path.join(base_dir, name + "_local.xlsx")
    path_remote = os.path.join(base_dir, name + "_remote.xlsx")
    path_merged = os.path.join(base_dir, name + "_merged.xlsx")

    for p in (path_local, path_remote, path_merged):
        if not os.path.isfile(p):
            print("文件不存在:", p)
            return

    sheet_name = "Data_HeroGearSkin"

    def dump_sheet(path, label, row_range=None):
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name not in wb.sheetnames:
            print("%s: 无表 %s" % (label, sheet_name))
            return None, None, None
        ws = wb[sheet_name]
        max_col = min(ws.max_column or 1, 6)
        rows, indices = load_sheet_rows_full(ws, max_col)
        print("\n--- %s (%s) ---" % (label, path))
        print("总行数(含表头): %d, row_indices 范围: %s .. %s" % (len(rows), indices[0] if indices else "-", indices[-1] if indices else "-"))
        # 打印 row_index 95-110 附近
        rlo, rhi = (row_range or (95, 112))[0], (row_range or (95, 112))[1]
        for i in range(len(rows)):
            ri = indices[i] if i < len(indices) else i + 1
            if rlo <= ri <= rhi:
                key = key_str_normalized(rows[i][0]) if rows[i] else ""
                preview = [rows[i][c] if c < len(rows[i]) else "" for c in range(min(4, max_col))]
                print("  行号=%3d  key=%r  preview=%s" % (ri, key, preview))
        base_keys = set(key_str_normalized(r[0]) if r else "" for r in rows)
        base_keys.discard("")
        ordered = ordered_keys_normalized(rows)
        return rows, indices, (base_keys, ordered)

    dump_sheet(path_local, "LOCAL")
    dump_sheet(path_remote, "REMOTE")
    dump_sheet(path_merged, "MERGED")

    # 模拟合并顺序（基准=remote）
    wb_r = openpyxl.load_workbook(path_remote, data_only=True)
    wb_l = openpyxl.load_workbook(path_local, data_only=True)
    if sheet_name not in wb_r.sheetnames or sheet_name not in wb_l.sheetnames:
        print("\n表不存在，跳过合并顺序模拟")
        return
    ws_r = wb_r[sheet_name]
    ws_l = wb_l[sheet_name]
    max_col = max(ws_r.max_column or 1, ws_l.max_column or 1)
    rows_in, idx_in = load_sheet_rows_full(ws_r, max_col)
    rows_o, idx_o = load_sheet_rows_full(ws_l, max_col)
    base_keys = set(key_str_normalized(r[0]) if r else "" for r in rows_in)
    base_keys.discard("")
    base_ordered = ordered_keys_normalized(rows_in)
    other_ordered = ordered_keys_normalized(rows_o)
    new_keys = [k for k in other_ordered if k not in base_keys]
    merged_ordered = merge_ordered_with_new_rows(base_ordered, new_keys)

    key_to_row_in = {}
    for i, r in enumerate(rows_in):
        k = key_str_normalized(r[0]) if r else ""
        if k and i < len(idx_in):
            key_to_row_in[k] = idx_in[i]
    key_to_row_other = {}
    for i, r in enumerate(rows_o):
        k = key_str_normalized(r[0]) if r else ""
        if k and i < len(idx_o):
            key_to_row_other[k] = idx_o[i]

    print("\n--- 合并顺序模拟（基准=REMOTE，插入 LOCAL 新增行）---")
    print("REMOTE 行数(含表头): %d, LOCAL 行数: %d" % (len(rows_in), len(rows_o)))
    print("REMOTE 独有 key 数: %d, LOCAL 独有(将新增) key 数: %d" % (
        len(base_keys - set(other_ordered)), len(new_keys)))
    print("merged_ordered 长度: %d" % len(merged_ordered))
    # 找到 101 附近的 key
    for pos, k in enumerate(merged_ordered):
        if pos >= 97 and pos <= 105:
            r_in = key_to_row_in.get(k)
            r_o = key_to_row_other.get(k)
            src = "REMOTE" if r_in else ("LOCAL" if r_o else "?")
            print("  merged[%d] key=%r  -> REMOTE行=%s LOCAL行=%s (来源:%s)" % (pos + 1, k, r_in, r_o, src))

    # 插入列表
    last_base_row = None
    inserts = []
    for k in merged_ordered:
        if k in key_to_row_in:
            last_base_row = key_to_row_in[k]
        elif k in key_to_row_other:
            insert_after = (last_base_row if last_base_row is not None else 0)
            inserts.append((insert_after, k))
    print("\n插入列表 (insert_after, key) 在 98-105 附近:")
    for j, (ia, k) in enumerate(inserts):
        if 97 <= j <= 105 or (ia and 97 <= ia <= 105):
            print("  insert_after=%s key=%r" % (ia, k))
    print("总插入数: %d" % len(inserts))

if __name__ == "__main__":
    main()
