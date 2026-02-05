# -*- coding: utf-8 -*-
"""
Excel 合并与对比工具（Fork Merge/Diff Tool 兼容）
- 合并模式（4 参数）：三向合并 LOCAL + BASE + REMOTE -> MERGED，并备份三方
- 对比模式（2 参数）：二向 diff LOCAL vs REMOTE，生成对比 Excel 并打开

用法:
  合并: python MergeExcelFork.py <local> <base> <remote> <merged>
  对比: python MergeExcelFork.py <local> <remote>

退出码: 0 成功, 1 参数/文件错误, 2 合并/对比异常

日志: 与 exe/脚本同目录下的 MergeExcelFork.log
输出: 对比模式生成 {原文件名}_compare.xlsx，与第一个文件同目录
"""

import sys
import os
import shutil
import subprocess
import platform
import traceback
from datetime import datetime

# Sheet 名过滤：跳过 # 开头、Sheet 默认名
_SKIP_PREFIX = "#"
_LOG_FILE = "MergeExcelFork.log"


def _log_dir():
    """日志所在目录：exe 或脚本所在目录"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _log_path():
    return os.path.join(_log_dir(), _LOG_FILE)


def _log(msg, is_error=False):
    try:
        with open(_log_path(), "a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            prefix = "[ERROR] " if is_error else ""
            f.write("%s %s%s\n" % (ts, prefix, msg))
    except Exception:
        pass
_SKIP_SHEET_PREFIX = "Sheet"
_BACKUP_SUBDIR = "MergeExcelBackup"
_COMPARE_SUFFIX = "_compare"


def _cell_str(c):
    if c is None:
        return ""
    return str(c).strip()


def _should_skip_sheet(name):
    """跳过 # 开头或 Sheet 开头的表名"""
    s = (name or "").strip()
    return s.startswith(_SKIP_PREFIX) or s.startswith(_SKIP_SHEET_PREFIX)


def _get_sheet_names(wb):
    """返回需要处理的 Sheet 名称列表"""
    return [n for n in wb.sheetnames if not _should_skip_sheet(n)]


def _load_sheet_rows(ws, max_col=None):
    """加载 Sheet 所有行，每行为 list；第一列空则跳过该行"""
    if ws is None:
        return []
    rows = []
    if max_col is None:
        max_col = ws.max_column or 1
    for row in ws.iter_rows(min_row=1, max_col=max_col, values_only=True):
        row_list = list(row) if row else []
        while len(row_list) < max_col:
            row_list.append("")
        key = _cell_str(row_list[0]) if row_list else ""
        if not key:
            continue
        rows.append(row_list)
    return rows


def _rows_to_dict(rows):
    """按第一列 key 转为 dict[key] = row_list"""
    d = {}
    for r in rows:
        k = _cell_str(r[0]) if r else ""
        if k:
            d[k] = r
    return d


def _ordered_keys(rows):
    """从 rows 提取 key 顺序（按行出现顺序）"""
    keys = []
    seen = set()
    for r in rows:
        k = _cell_str(r[0]) if r else ""
        if k and k not in seen:
            seen.add(k)
            keys.append(k)
    return keys


# ---- 合并模式 ----
# Git: LOCAL=ours=当前分支, REMOTE=theirs=被合并入的分支. 冲突时取 LOCAL.
def _merge_sheet(base_rows, local_rows, remote_rows, base_ordered, local_ordered, remote_ordered, max_col):
    """
    三向合并。顺序：BASE 原有 key 保持 BASE 顺序，新增 key 放到最后（LOCAL 新增先，REMOTE 新增后）
    返回 (merged_rows, row_types) 其中 row_types 为 {key: "新增"|"修改"|"冲突"}
    """
    base_set = set(base_rows)
    local_set = set(local_rows)
    remote_set = set(remote_rows)
    all_keys = base_set | local_set | remote_set
    merged = []
    row_types = {}

    def process_key(key):
        base_row = base_rows.get(key)
        local_row = local_rows.get(key)
        remote_row = remote_rows.get(key)
        if local_row is not None and remote_row is not None:
            if _row_equal(local_row, remote_row):
                merged.append(list(local_row))
                if base_row is None:
                    row_types[key] = "新增"
                elif not _row_equal(base_row, local_row):
                    row_types[key] = "修改"
            elif base_row is None:
                merged.append(list(local_row))
                row_types[key] = "冲突"
            elif _row_equal(base_row, local_row):
                merged.append(list(remote_row))
                row_types[key] = "修改"
            elif _row_equal(base_row, remote_row):
                merged.append(list(local_row))
                row_types[key] = "修改"
            else:
                merged.append(list(local_row))
                row_types[key] = "冲突"
        elif local_row is not None:
            merged.append(list(local_row))
            row_types[key] = "新增" if base_row is None else ("修改" if not _row_equal(base_row, local_row) else None)
        elif remote_row is not None:
            merged.append(list(remote_row))
            row_types[key] = "新增" if base_row is None else ("修改" if not _row_equal(base_row, remote_row) else None)

    for key in base_ordered:
        if key in all_keys:
            process_key(key)
    for key in local_ordered:
        if key not in base_set and key in all_keys:
            process_key(key)
    for key in remote_ordered:
        if key not in base_set and key not in local_set and key in all_keys:
            process_key(key)

    # 补齐列数
    for r in merged:
        while len(r) < max_col:
            r.append("")

    return merged, row_types


def _row_equal(a, b):
    if a is b:
        return True
    if a is None or b is None:
        return a == b
    return [_cell_str(c) for c in a] == [_cell_str(c) for c in b]


def _do_merge(path_local, path_base, path_remote, path_merged):
    """执行三向合并，写入 MERGED，并备份三方。新增/修改/冲突用颜色标记"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("ERROR: 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        return 2

    red_font = Font(color="FF0000", bold=True)
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    merged_dir = os.path.dirname(os.path.abspath(path_merged))
    base_name = os.path.splitext(os.path.basename(path_merged))[0]
    backup_dir = os.path.join(merged_dir, _BACKUP_SUBDIR)

    wb_local = openpyxl.load_workbook(path_local, data_only=True)
    wb_base = openpyxl.load_workbook(path_base, data_only=True)
    wb_remote = openpyxl.load_workbook(path_remote, data_only=True)

    # 保持 Sheet 顺序：以 BASE 顺序为主，再追加仅出现在 LOCAL/REMOTE 的 Sheet
    seen = set()
    sheet_names = []
    for n in _get_sheet_names(wb_base):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in _get_sheet_names(wb_local):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in _get_sheet_names(wb_remote):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in sheet_names:
        ws_local = wb_local[sheet_name] if sheet_name in wb_local.sheetnames else None
        ws_base = wb_base[sheet_name] if sheet_name in wb_base.sheetnames else None
        ws_remote = wb_remote[sheet_name] if sheet_name in wb_remote.sheetnames else None

        max_col = 1
        base_rows = {}
        local_rows = {}
        remote_rows = {}
        base_ordered = []
        local_ordered = []
        remote_ordered = []

        if ws_local:
            rows = _load_sheet_rows(ws_local)
            local_rows = _rows_to_dict(rows)
            local_ordered = _ordered_keys(rows)
            max_col = max(max_col, max(len(r) for r in rows) if rows else 1)
        if ws_base:
            rows = _load_sheet_rows(ws_base, max_col)
            base_rows = _rows_to_dict(rows)
            base_ordered = _ordered_keys(rows)
        if ws_remote:
            rows = _load_sheet_rows(ws_remote, max_col)
            remote_rows = _rows_to_dict(rows)
            remote_ordered = _ordered_keys(rows)
            max_col = max(max_col, max(len(r) for r in rows) if rows else 1)

        merged_rows, row_types = _merge_sheet(
            base_rows, local_rows, remote_rows,
            base_ordered, local_ordered, remote_ordered,
            max_col
        )
        if not merged_rows and not local_rows and not remote_rows:
            continue

        ws_out = wb_out.create_sheet(sheet_name)
        for r, row_list in enumerate(merged_rows, start=1):
            key_str = _cell_str(row_list[0]) if row_list else ""
            rtype = row_types.get(key_str)
            fill = green_fill if rtype == "新增" else (yellow_fill if rtype == "修改" else (red_fill if rtype == "冲突" else None))
            for c, val in enumerate(row_list, start=1):
                cell = ws_out.cell(row=r, column=c, value=val)
                if fill:
                    cell.fill = fill
                if rtype == "冲突":
                    cell.font = red_font

    if not wb_out.sheetnames:
        wb_out.create_sheet("Data")
    if merged_dir:
        os.makedirs(merged_dir, exist_ok=True)
    wb_out.save(path_merged)
    wb_local.close()
    wb_base.close()
    wb_remote.close()
    wb_out.close()

    os.makedirs(backup_dir, exist_ok=True)
    shutil.copy2(path_local, os.path.join(backup_dir, base_name + "_local.xlsx"))   # LOCAL=ours=当前分支
    shutil.copy2(path_remote, os.path.join(backup_dir, base_name + "_remote.xlsx"))  # REMOTE=theirs=被合并分支
    shutil.copy2(path_merged, os.path.join(backup_dir, base_name + "_merged.xlsx"))

    _log("合并完成 MERGED=%s 备份=%s" % (path_merged, backup_dir))
    print("OK: 合并完成。MERGED=%s 备份=%s" % (path_merged, backup_dir), file=sys.stdout)
    return 0


# ---- 对比模式 ----
def get_compare_data(path_a, path_b):
    """返回 (path_out, sheet_names, [(sheet, key, status, str_a, str_b), ...]) 供 GUI 显示"""
    try:
        import openpyxl
    except ImportError:
        return None, [], []

    out_dir = os.path.dirname(os.path.abspath(path_a))
    base_name = os.path.splitext(os.path.basename(path_a))[0]
    path_out = os.path.join(out_dir, base_name + _COMPARE_SUFFIX + ".xlsx")

    wb_a = openpyxl.load_workbook(path_a, data_only=True)
    wb_b = openpyxl.load_workbook(path_b, data_only=True)

    seen = set()
    sheet_names = []
    for n in _get_sheet_names(wb_a):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in _get_sheet_names(wb_b):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    if not sheet_names:
        for wb in (wb_a, wb_b):
            for n in (wb.sheetnames or []):
                sheet_names.append(n)
                break
            if sheet_names:
                break
        if not sheet_names:
            sheet_names = ["Sheet1"]

    diff_rows = []
    for sheet_name in sheet_names:
        ws_a = wb_a[sheet_name] if sheet_name in wb_a.sheetnames else None
        ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None
        rows_a = _load_sheet_rows(ws_a) if ws_a else []
        rows_b = _load_sheet_rows(ws_b) if ws_b else []
        dict_a = _rows_to_dict(rows_a)
        dict_b = _rows_to_dict(rows_b)
        all_keys = sorted(set(dict_a) | set(dict_b))
        max_col = max(
            max(len(r) for r in rows_a) if rows_a else 0,
            max(len(r) for r in rows_b) if rows_b else 0,
            1
        )
        for key in all_keys:
            row_a = dict_a.get(key)
            row_b = dict_b.get(key)
            vals_a = [_cell_str(c) for c in (row_a or [])]
            vals_b = [_cell_str(c) for c in (row_b or [])]
            while len(vals_a) < max_col:
                vals_a.append("")
            while len(vals_b) < max_col:
                vals_b.append("")
            str_a = " | ".join(vals_a)
            str_b = " | ".join(vals_b)
            if row_a is None:
                status = "B新增"
            elif row_b is None:
                status = "A独有"
            elif str_a != str_b:
                status = "修改"
            else:
                status = "相同"
            diff_rows.append((sheet_name, key, status, str_a, str_b))

    wb_a.close()
    wb_b.close()
    return path_out, sheet_names, diff_rows


def _write_compare_excel(path_out, sheet_names, diff_rows, open_file=False):
    """根据 diff 数据写入 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
    except ImportError:
        return 2
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    row_idx = {}
    for sheet_name in sheet_names:
        ws_out = wb_out.create_sheet(sheet_name[:31])
        ws_out.cell(row=1, column=1, value="[Key]")
        ws_out.cell(row=1, column=2, value="[A-LEFT]")
        ws_out.cell(row=1, column=3, value="[B-RIGHT]")
        ws_out.cell(row=1, column=4, value="[Status]")
        row_idx[sheet_name] = 2
    for sheet_name, key, status, str_a, str_b in diff_rows:
        ws_out = wb_out[sheet_name[:31]]
        r = row_idx[sheet_name]
        ws_out.cell(row=r, column=1, value=key)
        ws_out.cell(row=r, column=2, value=str_a)
        ws_out.cell(row=r, column=3, value=str_b)
        ws_out.cell(row=r, column=4, value=status)
        fill = green_fill if status == "B新增" else (red_fill if status == "A独有" else (yellow_fill if status == "修改" else None))
        if fill:
            for col in range(1, 5):
                ws_out.cell(row=r, column=col).fill = fill
        row_idx[sheet_name] = r + 1
    wb_out.save(path_out)
    wb_out.close()
    if open_file:
        try:
            if platform.system() == "Windows":
                os.startfile(path_out)
            elif platform.system() == "Darwin":
                subprocess.run(["open", path_out], check=False)
            else:
                subprocess.run(["xdg-open", path_out], check=False)
        except Exception:
            pass
    return 0


def _do_compare(path_a, path_b, open_file=True):
    """二向对比，生成对比 Excel，可选打开"""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        return 2

    result = get_compare_data(path_a, path_b)
    if result[0] is None:
        return 2
    path_out, sheet_names, diff_rows = result
    _log("对比模式 输出: %s" % path_out)
    _write_compare_excel(path_out, sheet_names, diff_rows, open_file=open_file)

    if open_file:
        try:
            if platform.system() == "Windows":
                os.startfile(path_out)
            elif platform.system() == "Darwin":
                subprocess.run(["open", path_out], check=False)
            else:
                subprocess.run(["xdg-open", path_out], check=False)
        except Exception:
            pass

    print("OK: 对比已生成 %s" % path_out, file=sys.stdout)
    return 0


def stage_merged_and_cleanup(path_merged, path_local, path_base, path_remote, log_callback=None):
    """
    解决冲突后：1) git add 合并文件使 Fork 识别为已解决  2) 清理临时文件
    log_callback(msg, is_error=False) 可选，用于 GUI 日志
    """
    def _log_cb(msg, is_err=False):
        _log(msg, is_error=is_err)
        if log_callback:
            try:
                log_callback(msg, is_err)
            except Exception:
                pass

    work_dir = os.path.dirname(os.path.abspath(path_merged))
    if not work_dir:
        work_dir = "."
    abs_merged = os.path.abspath(path_merged)
    try:
        rr = subprocess.run(
            ["git", "rev-parse", "--show-toplevel"],
            cwd=work_dir,
            capture_output=True,
            text=True,
            timeout=5,
        )
        repo_root = rr.stdout.strip() if rr.returncode == 0 and rr.stdout else work_dir
    except Exception:
        repo_root = work_dir
    rel_path = os.path.relpath(abs_merged, repo_root).replace("\\", "/")
    if rel_path.startswith(".."):
        rel_path = os.path.basename(path_merged)

    try:
        r = subprocess.run(
            ["git", "add", rel_path],
            cwd=repo_root,
            capture_output=True,
            text=True,
            timeout=10,
        )
        if r.returncode == 0:
            _log_cb("已执行 git add，冲突已标记为已解决")
        else:
            _log_cb("git add 失败: %s" % (r.stderr or r.stdout or "未知"), is_err=True)
    except Exception as e:
        _log_cb("git add 异常: %s" % e, is_err=True)

    def _is_temp(p):
        if not p or not os.path.isfile(p):
            return False
        pn = p.lower()
        return "temp" in pn or "tmp" in pn or "fork" in pn or "appdata" in pn

    for label, p in [("LOCAL", path_local), ("BASE", path_base), ("REMOTE", path_remote)]:
        if _is_temp(p):
            try:
                os.remove(p)
                _log_cb("已清理临时文件 %s: %s" % (label, p))
            except Exception as e:
                _log_cb("清理 %s 失败: %s" % (label, e), is_err=True)

    # 删除 MergeExcelBackup 中当前文件的备份（含从 git 索引移除）
    merged_dir = os.path.dirname(os.path.abspath(path_merged))
    base_name = os.path.splitext(os.path.basename(path_merged))[0]
    backup_dir = os.path.join(merged_dir, _BACKUP_SUBDIR)
    for suf in ["_local.xlsx", "_remote.xlsx", "_merged.xlsx"]:
        bp = os.path.join(backup_dir, base_name + suf)
        if not os.path.isfile(bp):
            continue
        bp_rel = os.path.relpath(bp, repo_root).replace("\\", "/")
        try:
            r = subprocess.run(["git", "rm", "-f", bp_rel], cwd=repo_root, capture_output=True, text=True, timeout=5)
            if r.returncode == 0:
                _log_cb("已删除备份(含从 git 移除): %s" % bp)
            else:
                os.remove(bp)
                _log_cb("已删除备份: %s" % bp)
        except Exception as e:
            try:
                os.remove(bp)
                _log_cb("已删除备份: %s" % bp)
            except Exception as e2:
                _log_cb("删除备份失败 %s: %s" % (bp, e2), is_err=True)
    if os.path.isdir(backup_dir) and not os.listdir(backup_dir):
        try:
            os.rmdir(backup_dir)
            _log_cb("已删除空备份目录: %s" % backup_dir)
        except Exception:
            pass


def get_git_merge_info(path_merged):
    """
    从 MERGED 文件所在仓库获取 LOCAL/REMOTE 的提交信息，便于判断哪边是自己的修改。
    返回 (local_info, remote_info)，每个为 dict: hash, short_hash, author, email, date, message
    获取失败时返回 (None, None) 或部分 None。
    """
    local_info = None
    remote_info = None
    try:
        work_dir = os.path.dirname(os.path.abspath(path_merged))
        if not work_dir:
            work_dir = "."
        rel_path = os.path.relpath(path_merged, work_dir).replace("\\", "/")
        if rel_path.startswith(".."):
            return None, None

        fmt = "%H%n%h%n%an%n%ae%n%ci%n%s"

        def run_git(ref):
            r = subprocess.run(
                ["git", "log", "-1", "--format=" + fmt, ref, "--", rel_path],
                cwd=work_dir,
                capture_output=True,
                text=True,
                timeout=5,
            )
            if r.returncode != 0 or not r.stdout.strip():
                return None
            parts = r.stdout.strip().split("\n")
            if len(parts) >= 6:
                return {
                    "hash": parts[0],
                    "short_hash": parts[1],
                    "author": parts[2],
                    "email": parts[3],
                    "date": parts[4],
                    "message": parts[5],
                }
            return None

        local_info = run_git("HEAD")

        for ref in ["MERGE_HEAD", "REBASE_HEAD", "CHERRY_PICK_HEAD"]:
            remote_info = run_git(ref)
            if remote_info:
                break
        if remote_info is None:
            remote_info = {}
    except Exception:
        pass
    return local_info, remote_info


def _normalize_args():
    """
    Fork 可能将多个路径合并为单个参数 "path1,path2" 或 "path1,path2,path3,path4"
    需按逗号拆分。返回 (mode, args) 其中 mode='merge'|'compare'|None
    """
    raw = sys.argv[1:]
    _log("启动 原始 args=%s" % raw)
    args = []
    for a in raw:
        for p in a.split(","):
            p = p.strip().strip('"').strip("'")
            if p:
                args.append(p)
    argc = len(args)
    if argc == 4:
        return "merge", args
    if argc == 2:
        return "compare", args
    return None, args


def main():
    try:
        mode, args = _normalize_args()
        argc = len(args)
        _log("解析后 argc=%d args=%s" % (argc, args))
        if mode == "merge":
            path_local, path_base, path_remote, path_merged = args[0], args[1], args[2], args[3]
            for p, name in [(path_local, "LOCAL"), (path_base, "BASE"), (path_remote, "REMOTE")]:
                if not os.path.isfile(p):
                    msg = "%s 不存在: %s" % (name, p)
                    _log(msg, is_error=True)
                    print("ERROR: " + msg, file=sys.stderr)
                    sys.exit(1)
            try:
                from ExcelMergeGUI import MergeWindow
                win = MergeWindow(path_local, path_base, path_remote, path_merged)
                win.run()
                sys.exit(0)
            except Exception as gui_err:
                _log("GUI 合并失败，回退命令行: %s" % gui_err)
                code = _do_merge(path_local, path_base, path_remote, path_merged)
                sys.exit(code)
        elif mode == "compare":
            path_a, path_b = args[0], args[1]
            if not os.path.isfile(path_a):
                msg = "文件 A 不存在: %s" % path_a
                _log(msg, is_error=True)
                print("ERROR: " + msg, file=sys.stderr)
                sys.exit(1)
            if not os.path.isfile(path_b):
                msg = "文件 B 不存在: %s" % path_b
                _log(msg, is_error=True)
                print("ERROR: " + msg, file=sys.stderr)
                sys.exit(1)
            try:
                from ExcelMergeGUI import DiffWindow
                win = DiffWindow(path_a, path_b)
                win.run()
                sys.exit(0)
            except Exception as gui_err:
                _log("GUI 对比失败，回退命令行: %s" % gui_err)
                code = _do_compare(path_a, path_b)
                sys.exit(code if isinstance(code, int) else 0)
        else:
            msg = "Usage: Merge (4 args) | Compare (2 args). Fork 可能传 path1,path2 单参数，已支持拆分"
            _log(msg)
            print(msg, file=sys.stderr)
            sys.exit(1)
    except Exception as e:
        tb = traceback.format_exc()
        _log("异常: %s\n%s" % (e, tb), is_error=True)
        print("ERROR: %s\n%s" % (e, tb), file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
