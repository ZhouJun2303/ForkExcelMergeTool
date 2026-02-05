# -*- coding: utf-8 -*-
"""
生成三向合并测试数据：BASE、LOCAL、REMOTE

覆盖场景（Sheet Data）：
  k1: 无变化
  k2: 仅 LOCAL 修改
  k3: 仅 REMOTE 修改
  k4: 两边同改（相同值）
  k5: 两边异改 -> 冲突（取 LOCAL 标红）
  k6: LOCAL 删除（取 REMOTE）
  k7: REMOTE 删除（取 LOCAL）
  k8: 仅 LOCAL 新增
  k9: 仅 REMOTE 新增
  k10: 两边同增同值
  k11: 两边异增 -> 冲突（取 LOCAL 标红）

Sheet Data2: 多 Sheet 测试，a/b 有修改、c/d 分别新增
"""

import os
import sys


def _dir_of_script():
    return os.path.dirname(os.path.abspath(__file__))


def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: 请先安装 openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    out_dir = _dir_of_script()

    # Sheet1: 主测试表，覆盖所有场景
    # k1: 无变化
    # k2: 仅 LOCAL 修改
    # k3: 仅 REMOTE 修改
    # k4: 两边同改（相同值）
    # k5: 两边异改 -> 冲突
    # k6: LOCAL 删除（base+remote 有，local 无）
    # k7: REMOTE 删除（base+local 有，remote 无）
    # k8: 仅 LOCAL 新增
    # k9: 仅 REMOTE 新增
    # k10: 两边同增（相同值）
    # k11: 两边异增 -> 冲突

    base_rows = [
        ["Key", "Col1", "Col2"],
        ["k1", "v1", "v1"],   # 无变化
        ["k2", "v2", "v2"],   # 仅 local 改
        ["k3", "v3", "v3"],   # 仅 remote 改
        ["k4", "v4", "v4"],   # 两边同改
        ["k5", "v5", "v5"],   # 两边异改-冲突
        ["k6", "v6", "v6"],   # local 删
        ["k7", "v7", "v7"],   # remote 删
    ]

    local_rows = [
        ["Key", "Col1", "Col2"],
        ["k1", "v1", "v1"],
        ["k2", "v2_L", "v2_L"],
        ["k3", "v3", "v3"],
        ["k4", "v4_both", "v4_both"],
        ["k5", "v5_L", "v5_L"],
        # k6 删除
        ["k7", "v7", "v7"],
        ["k8", "v8", "v8"],
        ["k9", "v9_L", "v9_L"],   # 两边同增-相同则取一即可，这里 local 先有
        ["k10", "v10_L", "v10_L"],
        ["k11", "v11_L", "v11_L"],
    ]

    remote_rows = [
        ["Key", "Col1", "Col2"],
        ["k1", "v1", "v1"],
        ["k2", "v2", "v2"],
        ["k3", "v3_R", "v3_R"],
        ["k4", "v4_both", "v4_both"],
        ["k5", "v5_R", "v5_R"],
        ["k6", "v6", "v6"],
        ["k8", "v8", "v8"],       # 仅 remote 新增（base 无，local 无）
        ["k9", "v9_L", "v9_L"],   # 两边同增相同
        ["k10", "v10_R", "v10_R"],
        ["k11", "v11_R", "v11_R"],
    ]

    # 重新梳理：
    # k8: 仅 LOCAL 新增 -> remote 无 k8
    # k9: 仅 REMOTE 新增 -> local 无 k9？不对，k10/k11 是两边都有
    # 设：k8 仅 local, k9 仅 remote, k10 两边同增同值, k11 两边异增
    local_rows = [
        ["Key", "Col1", "Col2"],
        ["k1", "v1", "v1"],
        ["k2", "v2_L", "v2_L"],
        ["k3", "v3", "v3"],
        ["k4", "v4_both", "v4_both"],
        ["k5", "v5_L", "v5_L"],
        ["k7", "v7", "v7"],
        ["k8", "v8", "v8"],       # 仅 local 新增
        ["k10", "v10_both", "v10_both"],  # 两边同增同值
        ["k11", "v11_L", "v11_L"],        # 两边异增-冲突
    ]
    remote_rows = [
        ["Key", "Col1", "Col2"],
        ["k1", "v1", "v1"],
        ["k2", "v2", "v2"],
        ["k3", "v3_R", "v3_R"],
        ["k4", "v4_both", "v4_both"],
        ["k5", "v5_R", "v5_R"],
        ["k6", "v6", "v6"],
        ["k9", "v9", "v9"],       # 仅 remote 新增
        ["k10", "v10_both", "v10_both"],
        ["k11", "v11_R", "v11_R"],
    ]

    def write_xlsx(path, sheets_dict):
        """sheets_dict: {sheet_name: rows}"""
        wb = openpyxl.Workbook()
        first = True
        for name, rows in sheets_dict.items():
            if first:
                ws = wb.active
                ws.title = name
                first = False
            else:
                ws = wb.create_sheet(name)
            for r, row in enumerate(rows, start=1):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=val)
        wb.save(path)
        wb.close()

    # Sheet2: 简表，测试多 Sheet
    base_sheet2 = [["Id", "Val"], ["a", "1"], ["b", "2"]]
    local_sheet2 = [["Id", "Val"], ["a", "1_changed"], ["b", "2"], ["c", "3"]]
    remote_sheet2 = [["Id", "Val"], ["a", "1"], ["b", "2_remote"], ["d", "4"]]

    write_xlsx(os.path.join(out_dir, "base.xlsx"), {"Data": base_rows, "Data2": base_sheet2})
    write_xlsx(os.path.join(out_dir, "local.xlsx"), {"Data": local_rows, "Data2": local_sheet2})
    write_xlsx(os.path.join(out_dir, "remote.xlsx"), {"Data": remote_rows, "Data2": remote_sheet2})

    print("OK: 已生成 base.xlsx, local.xlsx, remote.xlsx（含 Data/Data2 两 Sheet）", file=sys.stdout)
    print("场景: k1无改 k2仅L改 k3仅R改 k4两边同改 k5冲突 k6 L删 k7 R删 k8仅L增 k9仅R增 k10两边同增 k11两边异增", file=sys.stdout)
    sys.exit(0)


if __name__ == "__main__":
    main()
