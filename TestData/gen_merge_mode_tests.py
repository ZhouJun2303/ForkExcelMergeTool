# -*- coding: utf-8 -*-
"""
生成多模式合并测试用 10 张 Excel：每模式 2 个（local + remote）。
- mode_a: 一侧多若干行，首列 key 有前缀（A-1, A-2, B-1）便于验证前缀插入
- mode_b: 一侧多若干列，表头有前缀（单价、总价等）
- mode_c: 一侧多若干 Sheet
- mode_d: 同一 key 行或列两边内容不同（冲突）
- mode_e: 综合 A+B+C+D
"""

import os
import sys


def _dir_of_script():
    return os.path.dirname(os.path.abspath(__file__))


def write_xlsx(path, sheets_dict):
    """sheets_dict: {sheet_name: list of rows}"""
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets_dict.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r, row in enumerate(rows, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
    wb.save(path)
    wb.close()


def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    out_dir = _dir_of_script()
    os.makedirs(out_dir, exist_ok=True)

    # ---- Mode A: 新增行。基准少行，另一侧多 A-1,A-2,B-1 等
    mode_a_local = {
        "Data": [
            ["Key", "Col1", "Col2"],
            ["A-1", "a1", "x"],
            ["B-1", "b1", "y"],
        ],
    }
    mode_a_remote = {
        "Data": [
            ["Key", "Col1", "Col2"],
            ["A-1", "a1", "x"],
            ["A-2", "a2", "x"],   # 新增，同前缀 A
            ["B-1", "b1", "y"],
            ["B-2", "b2", "y"],   # 新增，同前缀 B
        ],
    }
    write_xlsx(os.path.join(out_dir, "mode_a_local.xlsx"), mode_a_local)
    write_xlsx(os.path.join(out_dir, "mode_a_remote.xlsx"), mode_a_remote)

    # ---- Mode B: 新增列。基准少列，另一侧多「单价」「总价」等
    mode_b_local = {
        "Data": [
            ["Key", "名称", "数量"],
            ["1", "苹果", "10"],
            ["2", "梨", "20"],
        ],
    }
    mode_b_remote = {
        "Data": [
            ["Key", "名称", "数量", "单价", "总价"],
            ["1", "苹果", "10", "2.5", "25"],
            ["2", "梨", "20", "3", "60"],
        ],
    }
    write_xlsx(os.path.join(out_dir, "mode_b_local.xlsx"), mode_b_local)
    write_xlsx(os.path.join(out_dir, "mode_b_remote.xlsx"), mode_b_remote)

    # ---- Mode C: 新增 Sheet。基准只有 Sheet1，另一侧多 Sheet2、Sheet3
    mode_c_local = {
        "Sheet1": [["Key", "V"], ["1", "a"]],
    }
    mode_c_remote = {
        "Sheet1": [["Key", "V"], ["1", "a"]],
        "Sheet2": [["Id", "X"], ["x", "1"]],
        "Sheet3": [["Name", "Y"], ["n", "2"]],
    }
    write_xlsx(os.path.join(out_dir, "mode_c_local.xlsx"), mode_c_local)
    write_xlsx(os.path.join(out_dir, "mode_c_remote.xlsx"), mode_c_remote)

    # ---- Mode D: 冲突行/列。两边都有同一 key 但内容不同
    mode_d_local = {
        "Data": [
            ["Key", "Col1", "Col2"],
            ["k1", "local1", "local2"],
            ["k2", "local_a", "local_b"],
        ],
        "Head": ["姓名", "年龄", "城市"],
        "Data2": [
            ["张三", "25", "北京"],
            ["李四", "30", "上海"],
        ],
    }
    mode_d_remote = {
        "Data": [
            ["Key", "Col1", "Col2"],
            ["k1", "remote1", "remote2"],
            ["k2", "remote_a", "remote_b"],
        ],
        "Head": ["姓名", "年龄", "城市"],
        "Data2": [
            ["张三", "26", "北京"],
            ["李四", "30", "广州"],
        ],
    }
    write_xlsx(os.path.join(out_dir, "mode_d_local.xlsx"), mode_d_local)
    write_xlsx(os.path.join(out_dir, "mode_d_remote.xlsx"), mode_d_remote)

    # ---- Mode E: 综合（既有新增行/列/Sheet，也有冲突）
    mode_e_local = {
        "Main": [
            ["Key", "V1", "V2"],
            ["A-1", "a", "b"],
            ["id1", "local_val", "x"],
        ],
    }
    mode_e_remote = {
        "Main": [
            ["Key", "V1", "V2"],
            ["A-1", "a", "b"],
            ["A-2", "c", "d"],
            ["id1", "remote_val", "x"],
        ],
        "Extra": [["X", "Y"], ["1", "2"]],
    }
    write_xlsx(os.path.join(out_dir, "mode_e_local.xlsx"), mode_e_local)
    write_xlsx(os.path.join(out_dir, "mode_e_remote.xlsx"), mode_e_remote)

    print("OK: 已生成 10 个测试 Excel：mode_a_local/remote ... mode_e_local/remote", file=sys.stdout)
    sys.exit(0)


if __name__ == "__main__":
    main()
