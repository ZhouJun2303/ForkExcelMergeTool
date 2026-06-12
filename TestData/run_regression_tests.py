# -*- coding: utf-8 -*-
"""
针对核心 bugfix 的轻量回归测试。
不依赖 pytest，输出写入 TestData/_output/regression。
"""

import os
import shutil
import sys
import time

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def _setup_paths():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(script_dir)
    if root not in sys.path:
        sys.path.insert(0, root)
    scripts_dir = os.path.join(root, "Scripts")
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    os.chdir(root)
    out_dir = os.path.join(script_dir, "_output", "regression")
    if os.path.isdir(out_dir):
        shutil.rmtree(out_dir)
    os.makedirs(out_dir, exist_ok=True)
    return root, out_dir


def _write_xlsx(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(path)
    wb.close()


def _rows(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Data"]
    values = [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]
    wb.close()
    return values


def _assert(cond, msg):
    if not cond:
        raise AssertionError(msg)


def _add_common_sheet_metadata(ws):
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:B2"
    ws.column_dimensions["B"].width = 24
    validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    validation.add("B2:B10")
    ws.add_data_validation(validation)
    ws.conditional_formatting.add(
        "B2:B10",
        CellIsRule(
            operator="notEqual",
            formula=['""'],
            fill=PatternFill(fill_type="solid", fgColor="FFF2CC"),
        ),
    )
    table = Table(displayName="MetaTable", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def test_compare_opens_once(out_dir):
    import compare_core

    left = os.path.join(out_dir, "compare_left.xlsx")
    right = os.path.join(out_dir, "compare_right.xlsx")
    _write_xlsx(left, [["Key", "V"], ["a", "1"]])
    _write_xlsx(right, [["Key", "V"], ["a", "2"]])

    calls = []
    original = compare_core.write_compare_excel

    def wrapped(path_out, sheet_names, diff_rows, open_file=False):
        calls.append(open_file)
        return original(path_out, sheet_names, diff_rows, open_file=False)

    compare_core.write_compare_excel = wrapped
    try:
        code = compare_core.do_compare(left, right, open_file=True)
    finally:
        compare_core.write_compare_excel = original

    _assert(code == 0, "compare should succeed")
    _assert(calls == [True], "do_compare should delegate opening exactly once")


def test_compare_formula_changes(out_dir):
    import compare_core

    left = os.path.join(out_dir, "formula_left.xlsx")
    right = os.path.join(out_dir, "formula_right.xlsx")
    _write_xlsx(left, [["Key", "V"], ["a", "=1+1"]])
    _write_xlsx(right, [["Key", "V"], ["a", "=1+2"]])

    _, _, diff_rows = compare_core.get_compare_data(left, right)
    statuses = [row[2] for row in diff_rows if row[1] == "a"]
    _assert(statuses == ["修改"], "formula text changes should be reported")


def test_key_normalization_edges(out_dir):
    from excel_io import key_str_normalized

    _assert(key_str_normalized("1.0") == "1", "1.0 should match 1")
    _assert(key_str_normalized("001") == "001", "leading zero business IDs should be preserved")
    _assert(
        key_str_normalized("9007199254740993") == "9007199254740993",
        "long integer keys should not lose precision",
    )
    _assert(key_str_normalized("1e309") == "1e309", "scientific overflow-like keys should stay textual")


def test_backup_same_second(out_dir):
    from backup_util import create_merge_backup

    local = os.path.join(out_dir, "backup_local.xlsx")
    remote = os.path.join(out_dir, "backup_remote.xlsx")
    merged = os.path.join(out_dir, "backup_merged.xlsx")
    backup_root = os.path.join(out_dir, "backup_root")
    for path, value in [(local, "l"), (remote, "r"), (merged, "m")]:
        _write_xlsx(path, [["Key", "V"], ["a", value]])

    infos = [
        create_merge_backup(local, remote, merged, backup_root=backup_root, timestamp="20260101_010101")
        for _ in range(3)
    ]
    dirs = [info["dir"] for info in infos]
    _assert(len(set(dirs)) == 3, "same-second backups should create unique dirs")
    _assert(all(os.path.isdir(d) for d in dirs), "backup dirs should exist")


def test_merge_diff_rejects_macro_extensions(out_dir):
    from excel_format import merge_diff_supported

    _assert(merge_diff_supported("demo.xlsx"), "xlsx should remain supported")
    _assert(merge_diff_supported("template.xltx"), "xltx should remain supported")
    _assert(not merge_diff_supported("macro.xlsm"), "xlsm should not be parsed because VBA would be lost")
    _assert(not merge_diff_supported("macro_template.xltm"), "xltm should not be parsed because VBA would be lost")


def test_backup_path_shortens_when_root_is_long(out_dir):
    from backup_util import MAX_BACKUP_PATH_LEN, _backup_path_for_dir

    backup_dir = os.path.join(out_dir, "x" * 60, "y" * 45)
    context = os.path.join(out_dir, "very_long_excel_name_" + "n" * 100 + ".xlsx")
    path = _backup_path_for_dir(
        backup_dir,
        context,
        "merged",
        {"author": "a" * 60, "short_hash": "1234567890abcdef", "message": "m" * 120},
        {"author": "b" * 60, "short_hash": "abcdef1234567890", "message": "r" * 120},
    )
    _assert(len(os.path.abspath(path)) <= MAX_BACKUP_PATH_LEN, "backup file path should stay within configured budget")
    _assert("__" in os.path.basename(path), "short backup name should include a stable digest")


def test_delete_conflict_choices(out_dir):
    from merge_core import do_merge

    base = os.path.join(out_dir, "delete_base.xlsx")
    local_deleted = os.path.join(out_dir, "delete_local_deleted.xlsx")
    remote_kept = os.path.join(out_dir, "delete_remote_kept.xlsx")
    local_kept = os.path.join(out_dir, "delete_local_kept.xlsx")
    remote_deleted = os.path.join(out_dir, "delete_remote_deleted.xlsx")

    _write_xlsx(base, [["Key", "V"], ["keep", "base"], ["gone", "base"]])
    _write_xlsx(local_deleted, [["Key", "V"], ["keep", "local"]])
    _write_xlsx(remote_kept, [["Key", "V"], ["keep", "remote"], ["gone", "remote"]])
    _write_xlsx(local_kept, [["Key", "V"], ["keep", "local"], ["gone", "local"]])
    _write_xlsx(remote_deleted, [["Key", "V"], ["keep", "remote"]])

    out_keep_remote = os.path.join(out_dir, "delete_keep_remote.xlsx")
    code = do_merge(
        local_deleted, base, remote_kept, out_keep_remote,
        mode="D",
        d_choices=[{"sheet": "Data", "key": "gone", "choice": "remote", "kind": "row"}],
        backup_root=os.path.join(out_dir, "backup_delete_keep_remote"),
    )
    _assert(code == 0, "remote keep delete-conflict merge should succeed")
    rows = _rows(out_keep_remote)
    _assert(["gone", "remote"] in rows, "remote-kept row should be copied by key")

    out_delete_local = os.path.join(out_dir, "delete_choose_local_deleted.xlsx")
    code = do_merge(
        local_deleted, base, remote_kept, out_delete_local,
        mode="D",
        d_choices=[{"sheet": "Data", "key": "gone", "choice": "local", "kind": "row"}],
        backup_root=os.path.join(out_dir, "backup_delete_choose_local"),
    )
    _assert(code == 0, "local delete choice should succeed")
    keys = [row[0] for row in _rows(out_delete_local)]
    _assert("gone" not in keys, "choosing deleted local side should remove output row")

    out_delete_remote = os.path.join(out_dir, "delete_choose_remote_deleted.xlsx")
    code = do_merge(
        local_kept, base, remote_deleted, out_delete_remote,
        mode="D",
        d_choices=[{"sheet": "Data", "key": "gone", "choice": "remote", "kind": "row"}],
        backup_root=os.path.join(out_dir, "backup_delete_choose_remote"),
    )
    _assert(code == 0, "remote delete choice should succeed")
    keys = [row[0] for row in _rows(out_delete_remote)]
    _assert("gone" not in keys, "choosing deleted remote side should remove output row")

    out_keep_local = os.path.join(out_dir, "delete_keep_local.xlsx")
    code = do_merge(
        local_kept, base, remote_deleted, out_keep_local,
        mode="D",
        d_choices=[{"sheet": "Data", "key": "gone", "choice": "local", "kind": "row"}],
        backup_root=os.path.join(out_dir, "backup_delete_keep_local"),
    )
    _assert(code == 0, "local keep delete-conflict merge should succeed")
    rows = _rows(out_keep_local)
    _assert(["gone", "local"] in rows, "local-kept row should be copied by key")


def test_auto_single_side_modifications(out_dir):
    from merge_core import do_merge

    base = os.path.join(out_dir, "auto_base.xlsx")
    local = os.path.join(out_dir, "auto_local.xlsx")
    remote = os.path.join(out_dir, "auto_remote.xlsx")
    merged = os.path.join(out_dir, "auto_merged.xlsx")
    _write_xlsx(base, [["Key", "V"], ["local_only_change", "base"], ["remote_only_change", "base"], ["remote_deleted", "base"]])
    _write_xlsx(local, [["Key", "V"], ["local_only_change", "local"], ["remote_only_change", "base"], ["remote_deleted", "base"]])
    _write_xlsx(remote, [["Key", "V"], ["local_only_change", "base"], ["remote_only_change", "remote"]])

    code = do_merge(
        local, base, remote, merged,
        options={"A", "B", "E", "G"},
        backup_root=os.path.join(out_dir, "backup_auto"),
    )
    _assert(code == 0, "auto single-side merge should succeed")
    rows = _rows(merged)
    _assert(["local_only_change", "local"] in rows, "local-only modification should be preserved")
    _assert(["remote_only_change", "remote"] in rows, "remote-only modification should be preserved")
    keys = [row[0] for row in rows]
    _assert("remote_deleted" not in keys, "single-side delete with other side unchanged should auto delete")


def test_auto_single_side_cell_merge_preserves_new_columns(out_dir):
    from merge_core import do_merge

    base = os.path.join(out_dir, "auto_cell_base.xlsx")
    local = os.path.join(out_dir, "auto_cell_local.xlsx")
    remote = os.path.join(out_dir, "auto_cell_remote.xlsx")
    merged = os.path.join(out_dir, "auto_cell_merged.xlsx")
    _write_xlsx(base, [["Key", "Name"], ["k1", "base"]])
    _write_xlsx(local, [["Key", "Name"], ["k1", "local"]])
    _write_xlsx(remote, [["Key", "Name", "Price"], ["k1", "base", "9"]])

    code = do_merge(
        local, base, remote, merged,
        options={"E", "G"},
        backup_root=os.path.join(out_dir, "backup_auto_cell"),
    )
    _assert(code == 0, "auto cell-level merge should succeed")
    rows = _rows(merged)
    _assert(rows[0] == ["Key", "Name", "Price"], "new remote column should remain")
    _assert(rows[1] == ["k1", "local", "9"], "local changed cell and remote new column should both survive")


def test_mode_c_without_new_sheets_preserves_metadata(out_dir):
    from merge_core import do_merge

    local = os.path.join(out_dir, "metadata_local.xlsx")
    remote = os.path.join(out_dir, "metadata_remote.xlsx")
    merged = os.path.join(out_dir, "metadata_merged.xlsx")
    _write_xlsx(local, [["Key", "V"], ["a", "1"]])
    wb = openpyxl.load_workbook(local)
    ws = wb["Data"]
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:B2"
    wb.save(local)
    wb.close()
    shutil.copy2(local, remote)

    code = do_merge(
        local, local, remote, merged,
        mode="C",
        backup_root=os.path.join(out_dir, "backup_metadata"),
    )
    _assert(code == 0, "mode C metadata preservation merge should succeed")
    wb_m = openpyxl.load_workbook(merged)
    ws_m = wb_m["Data"]
    _assert(ws_m.freeze_panes == "B2", "freeze panes should be preserved when no sheet rebuild is needed")
    _assert(ws_m.auto_filter.ref == "A1:B2", "auto filter should be preserved when no sheet rebuild is needed")
    wb_m.close()


def test_new_sheet_preserves_metadata(out_dir):
    from merge_core import do_merge

    local = os.path.join(out_dir, "new_sheet_meta_local.xlsx")
    remote = os.path.join(out_dir, "new_sheet_meta_remote.xlsx")
    merged = os.path.join(out_dir, "new_sheet_meta_merged.xlsx")
    _write_xlsx(local, [["Key", "V"], ["a", "1"]])
    shutil.copy2(local, remote)
    wb = openpyxl.load_workbook(remote)
    ws = wb.create_sheet("Extra")
    ws.append(["Key", "V"])
    ws.append(["x", "2"])
    _add_common_sheet_metadata(ws)
    wb.save(remote)
    wb.close()

    code = do_merge(
        local, local, remote, merged,
        options={"E"},
        backup_root=os.path.join(out_dir, "backup_new_sheet_meta"),
    )
    _assert(code == 0, "new sheet metadata merge should succeed")
    wb_m = openpyxl.load_workbook(merged)
    ws_m = wb_m["Extra"]
    _assert(ws_m.freeze_panes == "B2", "new sheet freeze panes should be copied")
    _assert(ws_m.auto_filter.ref == "A1:B2", "new sheet auto filter should be copied")
    _assert(int(ws_m.column_dimensions["B"].width) == 24, "new sheet column width should be copied")
    _assert(len(ws_m.data_validations.dataValidation) == 1, "new sheet data validation should be copied")
    _assert(str(ws_m.data_validations.dataValidation[0].sqref) == "B2:B10", "new sheet data validation range should be copied")
    _assert(len(ws_m.conditional_formatting) == 1, "new sheet conditional formatting should be copied")
    _assert(list(ws_m.tables.keys()) == ["MetaTable"], "new sheet tables should be copied")
    wb_m.close()


def test_column_conflict_writes_by_header(out_dir):
    from merge_core import do_merge

    base = os.path.join(out_dir, "col_conflict_base.xlsx")
    local = os.path.join(out_dir, "col_conflict_local.xlsx")
    remote = os.path.join(out_dir, "col_conflict_remote.xlsx")
    merged = os.path.join(out_dir, "col_conflict_merged.xlsx")
    _write_xlsx(base, [["Key", "A", "B"], ["k1", "base_a", "base_b"]])
    _write_xlsx(local, [["Key", "A", "B"], ["k1", "local_a", "base_b"]])
    _write_xlsx(remote, [["Key", "B", "A"], ["k1", "base_b", "remote_a"]])

    code = do_merge(
        local, base, remote, merged,
        mode="D",
        d_choices=[{"sheet": "Data", "key": "A", "choice": "remote", "kind": "column"}],
        backup_root=os.path.join(out_dir, "backup_col_conflict"),
    )
    _assert(code == 0, "column conflict merge should succeed")
    rows = _rows(merged)
    _assert(rows[0] == ["Key", "A", "B"], "target column order should remain stable")
    _assert(rows[1] == ["k1", "remote_a", "base_b"], "remote column A should write into target A, not source index")


def test_row_choice_writes_by_header(out_dir):
    from merge_core import do_merge

    base = os.path.join(out_dir, "row_header_base.xlsx")
    local = os.path.join(out_dir, "row_header_local.xlsx")
    remote = os.path.join(out_dir, "row_header_remote.xlsx")
    merged = os.path.join(out_dir, "row_header_merged.xlsx")
    _write_xlsx(base, [["Key", "A", "B"], ["k1", "base_a", "base_b"]])
    _write_xlsx(local, [["Key", "A", "B"], ["k1", "local_a", "local_b"]])
    _write_xlsx(remote, [["Key", "B", "A"], ["k1", "remote_b", "remote_a"]])

    code = do_merge(
        local, base, remote, merged,
        mode="D",
        d_choices=[{"sheet": "Data", "key": "k1", "choice": "remote", "kind": "row"}],
        backup_root=os.path.join(out_dir, "backup_row_header"),
    )
    _assert(code == 0, "row choice with reordered headers should succeed")
    rows = _rows(merged)
    _assert(rows[0] == ["Key", "A", "B"], "output header order should remain local/base order")
    _assert(rows[1] == ["k1", "remote_a", "remote_b"], "row choice should map source cells by header")


def test_column_conflicts_are_in_merge_preview(out_dir):
    from preview_core import build_merge_preview

    base = os.path.join(out_dir, "preview_col_base.xlsx")
    local = os.path.join(out_dir, "preview_col_local.xlsx")
    remote = os.path.join(out_dir, "preview_col_remote.xlsx")
    _write_xlsx(base, [["Key", "A", "B"], ["k1", "base_a", "same"], ["k2", "base_a2", "same"]])
    _write_xlsx(local, [["Key", "A", "B"], ["k1", "local_a", "same"], ["k2", "local_a2", "same"]])
    _write_xlsx(remote, [["Key", "B", "A"], ["k1", "same", "remote_a"], ["k2", "same", "remote_a2"]])

    result = build_merge_preview(local, base, remote, {"A", "B", "E", "G"}, "local")
    col_entries = [e for e in result["conflict_entries"] if e.get("kind") == "column"]
    _assert(any((e.get("data") or {}).get("key") == "A" for e in col_entries), "column conflict A should be exposed in G preview")
    _assert(any("列冲突" in str(item[1]) for item in result["items"]), "preview list should show column conflict")


def test_git_cleanup_policy_does_not_delete_repo_paths(out_dir):
    from git_util import CleanupPolicy

    repo_path = os.path.join(out_dir, "ForkTempProject")
    os.makedirs(repo_path, exist_ok=True)
    real_file = os.path.join(repo_path, "real.xlsx")
    _write_xlsx(real_file, [["Key", "V"], ["a", "1"]])
    policy = CleanupPolicy.default()
    _assert(not policy.allows(real_file), "repo path containing fork/temp words must not be treated as temp")


def test_compare_temp_detection_uses_cleanup_policy(out_dir):
    from git_util import CleanupPolicy

    repo_path = os.path.join(out_dir, "ForkNamedRealProject")
    os.makedirs(repo_path, exist_ok=True)
    compare_file = os.path.join(repo_path, "book_compare.xlsx")
    _write_xlsx(compare_file, [["Key", "V"], ["a", "1"]])
    _assert(not CleanupPolicy.default().allows(compare_file), "compare output in a real Fork-named project should not be considered temp")


def test_git_root_discovery_uses_marker_before_rev_parse(out_dir):
    import git_util

    repo_path = os.path.join(out_dir, "marker_repo")
    nested_path = os.path.join(repo_path, "a", "b")
    os.makedirs(os.path.join(repo_path, ".git"), exist_ok=True)
    os.makedirs(nested_path, exist_ok=True)

    original_run = git_util.subprocess.run

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("git rev-parse should not run when a .git marker exists")

    git_util.subprocess.run = fail_if_called
    try:
        root, err = git_util.discover_git_worktree_root(nested_path)
    finally:
        git_util.subprocess.run = original_run

    _assert(err is None, "marker-based git discovery should not return an error")
    _assert(os.path.normcase(root) == os.path.normcase(repo_path), "nearest .git marker should define repo root")


def test_stage_confirmation_survives_rev_parse_timeout(out_dir):
    import subprocess
    import git_util

    repo_path = os.path.join(out_dir, "stage_marker_repo")
    nested_path = os.path.join(repo_path, "dir")
    os.makedirs(os.path.join(repo_path, ".git"), exist_ok=True)
    os.makedirs(nested_path, exist_ok=True)

    merged = os.path.join(nested_path, "merged.xlsx")
    local = os.path.join(out_dir, "stage_local.xlsx")
    base = os.path.join(out_dir, "stage_base.xlsx")
    remote = os.path.join(out_dir, "stage_remote.xlsx")
    for path, value in [(merged, "m"), (local, "l"), (base, "b"), (remote, "r")]:
        _write_xlsx(path, [["Key", "V"], ["a", value]])

    calls = []
    original_run = git_util.subprocess.run

    def fake_run(cmd, *args, **kwargs):
        calls.append(cmd)
        if cmd[:3] == ["git", "rev-parse", "--show-toplevel"]:
            raise subprocess.TimeoutExpired(cmd, kwargs.get("timeout"))
        if cmd[:2] == ["git", "add"]:
            return subprocess.CompletedProcess(cmd, 0, "", "")
        if cmd[:2] == ["git", "rm"]:
            return subprocess.CompletedProcess(cmd, 0, "", "")
        return original_run(cmd, *args, **kwargs)

    git_util.subprocess.run = fake_run
    try:
        result = git_util.stage_merged_and_cleanup(merged, local, base, remote)
    finally:
        git_util.subprocess.run = original_run

    _assert(result.success, "confirmation should succeed without git rev-parse when .git marker exists")
    _assert(result.staged, "merged workbook should be staged")
    _assert(not any(cmd[:3] == ["git", "rev-parse", "--show-toplevel"] for cmd in calls), "rev-parse should be skipped")
    _assert(any(cmd[:2] == ["git", "add"] for cmd in calls), "git add should still run")


def test_git_driver_completion_strategy_writes_target(out_dir):
    from git_merge_driver import GitDriverCompletionStrategy

    target = os.path.join(out_dir, "driver_target.xlsx")
    merged = os.path.join(out_dir, "driver_merged.xlsx")
    _write_xlsx(target, [["Key", "V"], ["a", "old"]])
    _write_xlsx(merged, [["Key", "V"], ["a", "new"]])

    strategy = GitDriverCompletionStrategy(target, merged)
    result = strategy.complete(None)
    _assert(result.success, "driver completion should succeed")
    _assert(strategy.completed, "driver strategy should record completed state")
    rows = _rows(target)
    _assert(["a", "new"] in rows, "driver completion should replace %A with merged result")


def test_multi_new_columns(out_dir):
    from merge_core import do_merge

    local = os.path.join(out_dir, "cols_local.xlsx")
    remote = os.path.join(out_dir, "cols_remote.xlsx")
    merged = os.path.join(out_dir, "cols_merged.xlsx")
    _write_xlsx(local, [["Key", "Name"], ["1", "Apple"], ["2", "Pear"]])
    _write_xlsx(
        remote,
        [
            ["Key", "Name", "Price", "Qty", "Total"],
            ["1", "Apple", "2", "3", "6"],
            ["2", "Pear", "4", "5", "20"],
        ],
    )
    code = do_merge(
        local, local, remote, merged,
        mode="B",
        backup_root=os.path.join(out_dir, "backup_cols"),
    )
    _assert(code == 0, "mode B merge should succeed")
    rows = _rows(merged)
    _assert(rows[0] == ["Key", "Name", "Price", "Qty", "Total"], "new columns should keep expected order")
    _assert(rows[1] == ["1", "Apple", "2", "3", "6"], "first data row should include all new col values")
    _assert(rows[2] == ["2", "Pear", "4", "5", "20"], "second data row should include all new col values")


def test_performance_smoke(out_dir):
    from merge_core import do_merge

    local = os.path.join(out_dir, "perf_local.xlsx")
    remote = os.path.join(out_dir, "perf_remote.xlsx")
    merged = os.path.join(out_dir, "perf_merged.xlsx")
    headers_local = ["Key"] + ["B%02d" % i for i in range(1, 16)]
    headers_remote = headers_local + ["N%02d" % i for i in range(1, 16)]
    rows_local = [headers_local]
    rows_remote = [headers_remote]
    for i in range(1, 301):
        rows_local.append(["k%03d" % i] + ["v%d_%02d" % (i, c) for c in range(1, 16)])
        rows_remote.append(
            ["k%03d" % i]
            + ["v%d_%02d" % (i, c) for c in range(1, 16)]
            + ["n%d_%02d" % (i, c) for c in range(1, 16)]
        )
    _write_xlsx(local, rows_local)
    _write_xlsx(remote, rows_remote)

    started = time.time()
    code = do_merge(
        local, local, remote, merged,
        mode="B",
        backup_root=os.path.join(out_dir, "backup_perf"),
    )
    elapsed = time.time() - started
    _assert(code == 0, "performance smoke merge should succeed")
    _assert(elapsed < 20, "performance smoke should finish quickly enough, elapsed=%.2fs" % elapsed)
    print("Performance smoke elapsed: %.2fs" % elapsed)


def main():
    _, out_dir = _setup_paths()
    tests = [
        test_compare_opens_once,
        test_compare_formula_changes,
        test_key_normalization_edges,
        test_backup_same_second,
        test_merge_diff_rejects_macro_extensions,
        test_backup_path_shortens_when_root_is_long,
        test_delete_conflict_choices,
        test_auto_single_side_modifications,
        test_auto_single_side_cell_merge_preserves_new_columns,
        test_mode_c_without_new_sheets_preserves_metadata,
        test_new_sheet_preserves_metadata,
        test_column_conflict_writes_by_header,
        test_row_choice_writes_by_header,
        test_column_conflicts_are_in_merge_preview,
        test_git_cleanup_policy_does_not_delete_repo_paths,
        test_compare_temp_detection_uses_cleanup_policy,
        test_git_root_discovery_uses_marker_before_rev_parse,
        test_stage_confirmation_survives_rev_parse_timeout,
        test_git_driver_completion_strategy_writes_target,
        test_multi_new_columns,
        test_performance_smoke,
    ]
    for test in tests:
        test(out_dir)
        print("OK %s" % test.__name__)
    print("All regression tests passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
