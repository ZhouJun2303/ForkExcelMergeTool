# -*- coding: utf-8 -*-
"""
Excel 合并/对比 图形化界面（tkinter）
- Merge: 冲突可选手动选本地/线上，确认后合并
- Diff: 显示差异，关闭时自动清理缓存
"""

import os
import sys
import shutil
import subprocess
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

# 导入主模块的共享函数
import MergeExcelFork as _core

_LOG_FILE = "MergeExcelFork.log"


def _gui_log_path():
    if getattr(sys, "frozen", False):
        return os.path.join(os.path.dirname(sys.executable), _LOG_FILE)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), _LOG_FILE)


def _gui_log(msg, log_widget=None, is_error=False):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = "[ERROR] " if is_error else ""
    line = "%s %s%s" % (ts, prefix, msg)
    try:
        with open(_gui_log_path(), "a", encoding="utf-8") as f:
            f.write(datetime.now().strftime("%Y-%m-%d ") + line + "\n")
    except Exception:
        pass
    if log_widget:
        try:
            if isinstance(log_widget, tk.StringVar):
                log_widget.set(line.strip())
            else:
                log_widget.insert(tk.END, line + "\n")
                log_widget.see(tk.END)
            if hasattr(log_widget, "update_idletasks"):
                log_widget.update_idletasks()
        except Exception:
            pass


def _key_str_normalized(c):
    """合并时统一 key：数字 1 与 1.0 视为同一行，避免漏检冲突"""
    if c is None:
        return ""
    s = str(c).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return str(f)
    except ValueError:
        return s


def _rows_to_dict_normalized(rows):
    """按首列建 dict，key 用规范化字符串"""
    d = {}
    for r in rows:
        k = _key_str_normalized(r[0]) if r else ""
        if k:
            d[k] = r
    return d


def _ordered_keys_normalized(rows):
    """按行顺序返回规范化 key 列表"""
    keys, seen = [], set()
    for r in rows:
        k = _key_str_normalized(r[0]) if r else ""
        if k and k not in seen:
            seen.add(k)
            keys.append(k)
    return keys


def _compute_conflicts(path_local, path_base, path_remote):
    """计算冲突列表，返回 (冲突项列表, 各 sheet 数据)。
    用 data_only=False 加载，避免公式单元格缓存为空导致本地/线上被误判为相同。
    """
    import openpyxl
    # 冲突检测用 data_only=False，这样公式/未打开过的文件也能正确比较出差异
    wb_l = openpyxl.load_workbook(path_local, data_only=False)
    wb_b = openpyxl.load_workbook(path_base, data_only=False)
    wb_r = openpyxl.load_workbook(path_remote, data_only=False)

    seen = set()
    sheet_names = []
    for n in _core._get_sheet_names(wb_b):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in _core._get_sheet_names(wb_l):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)
    for n in _core._get_sheet_names(wb_r):
        if n not in seen:
            seen.add(n)
            sheet_names.append(n)

    conflicts = []
    sheet_data = {}

    for sheet_name in sheet_names:
        ws_l = wb_l[sheet_name] if sheet_name in wb_l.sheetnames else None
        ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None
        ws_r = wb_r[sheet_name] if sheet_name in wb_r.sheetnames else None

        rows_l = _core._load_sheet_rows(ws_l) if ws_l else []
        rows_b = _core._load_sheet_rows(ws_b) if ws_b else []
        rows_r = _core._load_sheet_rows(ws_r) if ws_r else []

        # 用规范化 key，避免 1 与 1.0 被当成不同行导致漏冲突
        dict_l = _rows_to_dict_normalized(rows_l)
        dict_b = _rows_to_dict_normalized(rows_b)
        dict_r = _rows_to_dict_normalized(rows_r)
        ord_b = _ordered_keys_normalized(rows_b)
        ord_l = _ordered_keys_normalized(rows_l)
        ord_r = _ordered_keys_normalized(rows_r)

        base_set = set(dict_b)
        local_set = set(dict_l)
        remote_set = set(dict_r)
        all_keys = base_set | local_set | remote_set

        n_conflict_sheet = 0
        for key in all_keys:
            row_l = dict_l.get(key)
            row_r = dict_r.get(key)
            row_b = dict_b.get(key)
            if row_l is not None and row_r is not None:
                if _core._row_equal(row_l, row_r):
                    continue
                if row_b is None or (not _core._row_equal(row_b, row_l) and not _core._row_equal(row_b, row_r)):
                    n_conflict_sheet += 1
                    conflicts.append({
                        "sheet": sheet_name,
                        "key": key,
                        "local_row": row_l,
                        "remote_row": row_r,
                        "base_row": row_b,
                    })

        # 仅本地有、仅线上有的 key（用于界面展示，让用户看到“两边不同”的全貌）
        only_local_set = local_set - remote_set
        only_remote_set = remote_set - local_set
        for k in only_local_set:
            conflicts.append({
                "sheet": sheet_name,
                "key": k,
                "local_row": dict_l.get(k),
                "remote_row": None,
                "base_row": dict_b.get(k),
                "_only_local": True,
            })
        for k in only_remote_set:
            conflicts.append({
                "sheet": sheet_name,
                "key": k,
                "local_row": None,
                "remote_row": dict_r.get(k),
                "base_row": dict_b.get(k),
                "_only_remote": True,
            })

        sheet_data[sheet_name] = {
            "base_rows": dict_b, "local_rows": dict_l, "remote_rows": dict_r,
            "base_ordered": ord_b, "local_ordered": ord_l, "remote_ordered": ord_r,
            "max_col": max(
                max(len(r) for r in rows_l) if rows_l else 1,
                max(len(r) for r in rows_b) if rows_b else 1,
                max(len(r) for r in rows_r) if rows_r else 1,
            )
        }
        _log_merge_diagnostic(sheet_name, len(rows_l), len(rows_b), len(rows_r), len(all_keys), n_conflict_sheet)

    wb_l.close()
    wb_b.close()
    wb_r.close()
    return conflicts, sheet_data, sheet_names


def _log_merge_diagnostic(sheet_name, n_l, n_b, n_r, n_keys, n_conflict):
    """写入合并诊断到日志，便于排查 0 冲突"""
    try:
        path = _gui_log_path()
        with open(path, "a", encoding="utf-8") as f:
            from datetime import datetime
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write("%s [MERGE] Sheet=%s 行数 local=%d base=%d remote=%d all_keys=%d 冲突=%d\n" % (
                ts, sheet_name, n_l, n_b, n_r, n_keys, n_conflict))
    except Exception:
        pass


def _make_color_legend(parent, items, bg="#f0f2f5"):
    """在 parent 下生成一行颜色图例。items: [(颜色 hex 或 None, "说明文字"), ...]"""
    frame = tk.Frame(parent, bg=bg)
    for i, (color, text) in enumerate(items):
        if i > 0:
            tk.Label(frame, text="  ", bg=bg, font=("Segoe UI", 8)).pack(side=tk.LEFT)
        if color:
            patch = tk.Frame(frame, width=14, height=14, bg=color, relief=tk.SOLID, borderwidth=1)
            patch.pack(side=tk.LEFT, padx=(0, 4))
            patch.pack_propagate(False)
        tk.Label(frame, text=text, bg=bg, font=("Segoe UI", 8), fg="#1c1e21").pack(side=tk.LEFT)
    return frame


def _open_excel_file(path):
    """用系统默认程序打开 Excel 文件"""
    if not path or not os.path.isfile(path):
        return False
    try:
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
        return True
    except Exception:
        return False


def _setup_merge_styles(root):
    """为合并窗口配置简洁现代样式"""
    style = ttk.Style(root)
    try:
        style.theme_use("vista")
    except tk.TclError:
        pass
    bg = "#f0f2f5"
    fg = "#1c1e21"
    accent = "#1877f2"
    style.configure("TFrame", background=bg)
    style.configure("TLabel", background=bg, foreground=fg, font=("Segoe UI", 9))
    style.configure("TLabelframe", background=bg, font=("Segoe UI", 9))
    style.configure("TLabelframe.Label", background=bg, foreground=fg, font=("Segoe UI", 9, "bold"))
    style.configure("TButton", font=("Segoe UI", 9), padding=(12, 6))
    style.configure("Accent.TButton", font=("Segoe UI", 9, "bold"), padding=(14, 7))
    style.map("Accent.TButton", background=[("active", "#166fe5")])
    style.configure("Treeview", font=("Segoe UI", 9), rowheight=22, fieldbackground="#fff")
    style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
    root.configure(bg=bg)


class MergeWindow:
    """合并 GUI：冲突选择、打开左右 Excel、Git 信息、确认后合并"""
    def __init__(self, path_local, path_base, path_remote, path_merged):
        self.path_local = path_local
        self.path_base = path_base
        self.path_remote = path_remote
        self.path_merged = path_merged
        self.conflict_vars = []
        self.merge_done = False
        self.root = tk.Tk()
        self.root.title("Excel 三向合并")
        self.root.minsize(800, 680)
        self.root.geometry("1000x720")
        _setup_merge_styles(self.root)

        try:
            _gui_log("开始计算冲突...", None)
            self.conflicts, self.sheet_data, self.sheet_names = _compute_conflicts(
                path_local, path_base, path_remote
            )
        except Exception as e:
            import traceback
            msg = str(e) + "\n" + traceback.format_exc()
            _gui_log(msg, None, is_error=True)
            messagebox.showerror("错误", "加载失败: " + str(e))
            self.root.destroy()
            sys.exit(2)

        self.local_info, self.remote_info = _core.get_git_merge_info(path_merged)
        self._build_ui()

    def _build_ui(self):
        pad = 12
        # 底部操作区先 pack，保证始终在视口内可见（不被中间内容挤出）
        self.status_var = tk.StringVar(value="")
        bottom_bar = ttk.Frame(self.root, padding=(pad, 8))
        bottom_bar.pack(side=tk.BOTTOM, fill=tk.X)
        status_row = ttk.Frame(bottom_bar)
        status_row.pack(fill=tk.X)
        ttk.Label(status_row, textvariable=self.status_var, font=("Segoe UI", 9)).pack(anchor=tk.W)
        btn_row = ttk.Frame(bottom_bar)
        btn_row.pack(fill=tk.X, pady=(6, 0))
        self.btn_merge = ttk.Button(btn_row, text="生成合并结果", command=self._on_generate_merge, style="Accent.TButton")
        self.btn_merge.pack(side=tk.LEFT, padx=(0, 8))
        self.btn_open_merged = ttk.Button(btn_row, text="打开合并文件", command=self._on_open_merged)
        self.btn_open_merged.pack(side=tk.LEFT, padx=8)
        self.btn_confirm = ttk.Button(btn_row, text="确认无误并解决冲突", command=self._on_confirm_done)
        self.btn_confirm.pack(side=tk.LEFT, padx=8)
        self.btn_confirm.config(state=tk.DISABLED)
        ttk.Button(btn_row, text="取消", command=self._on_cancel).pack(side=tk.LEFT)
        _gui_log("已加载 %d 项差异（含仅本地有/仅线上有），请选择后点击「生成合并结果」" % len(self.conflicts), self.status_var)

        # 中间可滚动区域
        center = ttk.Frame(self.root)
        center.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0, pad))

        # 顶部：标题与目标路径
        top = ttk.Frame(center, padding=(0, 0, 0, 6))
        top.pack(fill=tk.X)
        ttk.Label(top, text="本地 (左) + 线上 (右) → 合并结果", font=("Segoe UI", 11, "bold")).pack(anchor=tk.W)
        path_short = self.path_merged if len(self.path_merged) <= 72 else "…" + self.path_merged[-68:]
        ttk.Label(top, text=path_short, font=("Segoe UI", 8)).pack(anchor=tk.W)

        # 左右版本信息卡片
        info_frame = ttk.LabelFrame(center, text="版本说明", padding=pad)
        info_frame.pack(fill=tk.X, pady=(0, 6))
        row1 = ttk.Frame(info_frame)
        row1.pack(fill=tk.X)
        left_box = ttk.LabelFrame(row1, text="本地 (Local)", padding=8)
        left_box.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self._fill_commit_info(left_box, self.local_info)
        ttk.Button(left_box, text="打开本地 Excel", command=lambda: self._open_local()).pack(anchor=tk.W, pady=(6, 0))
        right_box = ttk.LabelFrame(row1, text="线上 (Remote)", padding=8)
        right_box.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0))
        self._fill_commit_info(right_box, self.remote_info)
        ttk.Button(right_box, text="打开线上 Excel", command=lambda: self._open_remote()).pack(anchor=tk.W, pady=(6, 0))

        # 合并结果颜色说明（生成合并结果后 Excel 中的行背景色含义）
        legend_merge = _make_color_legend(center, [
            ("#CCFFCC", "绿色=新增"),
            ("#FFFF99", "黄色=修改"),
            ("#FFCCCC", "红色=冲突（需选择）"),
            (None, "无=无变化"),
        ])
        legend_merge.pack(anchor=tk.W, pady=(0, 6))

        # 冲突数与说明（含“仅本地有/仅线上有”，让两边不同的行都显示出来）
        n_need_choice = sum(1 for c in self.conflicts if not c.get("_only_local") and not c.get("_only_remote"))
        hint = "下表共 %d 项差异（其中 %d 项需选择取本地/取线上，其余为仅一方有将自动保留）" % (len(self.conflicts), n_need_choice)
        ttk.Label(center, text=hint, font=("Segoe UI", 9)).pack(anchor=tk.W, pady=(0, 2))
        ttk.Label(center, text="说明：首列相同 key 出现多行时只按一行参与合并；若需保留多行请用唯一 key。", font=("Segoe UI", 8)).pack(anchor=tk.W)
        if len(self.conflicts) == 0:
            ttk.Label(center, text="若应有冲突却显示 0：请确认三个文件首列为行关键列、且本地与线上内容确有不同；详见同目录下 MergeExcelFork.log。", font=("Segoe UI", 8)).pack(anchor=tk.W)
        ttk.Label(center, text="", font=("Segoe UI", 1)).pack(anchor=tk.W, pady=(0, 2))

        paned = ttk.PanedWindow(center, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True)

        table_frame = ttk.Frame(paned)
        paned.add(table_frame, weight=2)
        cols = ("Sheet", "Key", "选择")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=10)
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=100 if c != "Key" else 220)
        sb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for i, c in enumerate(self.conflicts):
            if c.get("_only_local"):
                choice_label = "将保留本地"
                var = tk.StringVar(value="本地")
            elif c.get("_only_remote"):
                choice_label = "将保留线上"
                var = tk.StringVar(value="线上")
            else:
                choice_label = "本地"
                var = tk.StringVar(value="本地")
            self.conflict_vars.append(var)
            self.tree.insert("", tk.END, values=(c["sheet"], c["key"], choice_label), tags=(str(i),))
        self.tree.tag_configure("conflict", background="#fff3cd")

        sel_frame = ttk.Frame(paned)
        paned.add(sel_frame, weight=0)
        ttk.Label(sel_frame, text="当前选中项：").pack(side=tk.LEFT)
        btn_frame = ttk.Frame(sel_frame)
        btn_frame.pack(side=tk.LEFT, padx=(8, 0))
        ttk.Button(btn_frame, text="取本地", command=lambda: self._set_choice("本地")).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="取线上", command=lambda: self._set_choice("线上")).pack(side=tk.LEFT, padx=2)

    def _fill_commit_info(self, parent, info):
        parts = []
        if info and isinstance(info, dict):
            if info.get("short_hash"):
                parts.append("Hash: %s" % info["short_hash"])
            if info.get("author"):
                parts.append("提交人: %s" % info["author"])
            if info.get("email"):
                parts.append("(%s)" % info["email"])
            if info.get("date"):
                parts.append(info["date"][:19] if len(info["date"]) >= 19 else info["date"])
            if info.get("message"):
                msg = info["message"]
                parts.append("事件: %s" % (msg[:50] + "…" if len(msg) > 50 else msg))
        if not parts:
            ttk.Label(parent, text="(无法获取 Git 信息)", foreground="gray", font=("", 8)).pack(anchor=tk.W)
        else:
            for p in parts:
                ttk.Label(parent, text=p, font=("", 8)).pack(anchor=tk.W)

    def _open_local(self):
        if _open_excel_file(self.path_local):
            _gui_log("已打开本地 Excel", self.status_var)
        else:
            messagebox.showwarning("提示", "文件不存在或无法打开")

    def _open_remote(self):
        if _open_excel_file(self.path_remote):
            _gui_log("已打开线上 Excel", self.status_var)
        else:
            messagebox.showwarning("提示", "文件不存在或无法打开")

    def _on_open_merged(self):
        if self.merge_done and _open_excel_file(self.path_merged):
            _gui_log("已打开合并文件", self.status_var)
        else:
            messagebox.showwarning("提示", "请先点击「生成合并结果」")

    def _on_generate_merge(self):
        for i, c in enumerate(self.conflicts):
            var = self.conflict_vars[i] if i < len(self.conflict_vars) else None
            c["_choice"] = "local" if (var and var.get() == "本地") else "remote"
        try:
            self._do_merge_with_choices()
            self.merge_done = True
            _gui_log("合并结果已生成，请确认后点击「确认无误并解决冲突」", self.status_var)
            messagebox.showinfo("完成", "合并结果已保存。\n请打开合并文件确认无误后，点击「确认无误并解决冲突」完成。")
            self.btn_merge.config(state=tk.DISABLED)
            self.btn_confirm.config(state=tk.NORMAL)
        except Exception as e:
            import traceback
            _gui_log("合并失败: " + str(e), self.status_var, is_error=True)
            messagebox.showerror("错误", str(e))

    def _on_confirm_done(self):
        if not self.merge_done:
            messagebox.showwarning("提示", "请先点击「生成合并结果」")
            return
        def _log_cb(msg, is_err=False):
            _gui_log(msg, self.status_var, is_error=is_err)
        _core.stage_merged_and_cleanup(
            self.path_merged,
            self.path_local,
            self.path_base,
            self.path_remote,
            log_callback=_log_cb,
        )
        messagebox.showinfo("完成", "冲突已解决：已 git add，已清理临时文件。Fork 将使用合并后的文件。")
        self.root.quit()
        self.root.destroy()
        sys.exit(0)

    def _set_choice(self, choice):
        sel = self.tree.selection()
        if not sel:
            return
        item = sel[0]
        idx = int(self.tree.item(item, "tags")[0])
        if idx < 0 or idx >= len(self.conflicts):
            return
        c = self.conflicts[idx]
        if c.get("_only_local"):
            choice = "本地"
        elif c.get("_only_remote"):
            choice = "线上"
        display = "将保留本地" if c.get("_only_local") else ("将保留线上" if c.get("_only_remote") else choice)
        vals = list(self.tree.item(item, "values"))
        vals[2] = display
        self.tree.item(item, values=vals)
        if idx < len(self.conflict_vars):
            self.conflict_vars[idx].set(choice)

    def _on_cancel(self):
        self.root.quit()
        self.root.destroy()
        sys.exit(1)

    def _do_merge_with_choices(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        red_font = Font(color="FF0000", bold=True)
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   # 新增
        yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 修改
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")     # 冲突
        merged_dir = os.path.dirname(os.path.abspath(self.path_merged))
        base_name = os.path.splitext(os.path.basename(self.path_merged))[0]
        backup_dir = os.path.join(merged_dir, _core._BACKUP_SUBDIR)

        choice_map = {}
        for c in self.conflicts:
            choice_map[(c["key"], c["sheet"])] = c.get("_choice", "local")

        def merge_sheet_with_choices(base_rows, local_rows, remote_rows, base_ord, local_ord, remote_ord,
                                     max_col, sheet_name):
            base_set = set(base_rows)
            local_set = set(local_rows)
            remote_set = set(remote_rows)
            all_keys = base_set | local_set | remote_set
            merged = []
            row_types = {}  # key -> "新增"|"修改"|"冲突"

            def process(key):
                br = base_rows.get(key)
                lr = local_rows.get(key)
                rr = remote_rows.get(key)
                ch = choice_map.get((key, sheet_name), "local")
                if lr is not None and rr is not None:
                    if _core._row_equal(lr, rr):
                        merged.append(list(lr))
                        if br is None:
                            row_types[key] = "新增"
                        elif not _core._row_equal(br, lr):
                            row_types[key] = "修改"
                    elif br is None or (not _core._row_equal(br, lr) and not _core._row_equal(br, rr)):
                        row_types[key] = "冲突"
                        merged.append(list(lr) if ch == "local" else list(rr))
                    elif _core._row_equal(br, lr):
                        merged.append(list(rr))
                        row_types[key] = "修改"
                    elif _core._row_equal(br, rr):
                        merged.append(list(lr))
                        row_types[key] = "修改"
                    else:
                        row_types[key] = "冲突"
                        merged.append(list(lr) if ch == "local" else list(rr))
                elif lr is not None:
                    merged.append(list(lr))
                    row_types[key] = "新增" if br is None else ("修改" if not _core._row_equal(br, lr) else None)
                elif rr is not None:
                    merged.append(list(rr))
                    row_types[key] = "新增" if br is None else ("修改" if not _core._row_equal(br, rr) else None)

            for k in base_ord:
                if k in all_keys:
                    process(k)
            for k in local_ord:
                if k not in base_set and k in all_keys:
                    process(k)
            for k in remote_ord:
                if k not in base_set and k not in local_set and k in all_keys:
                    process(k)

            for r in merged:
                while len(r) < max_col:
                    r.append("")
            return merged, row_types

        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        for sheet_name in self.sheet_names:
            sd = self.sheet_data[sheet_name]
            merged_rows, row_types = merge_sheet_with_choices(
                sd["base_rows"], sd["local_rows"], sd["remote_rows"],
                sd["base_ordered"], sd["local_ordered"], sd["remote_ordered"],
                sd["max_col"], sheet_name
            )
            if not merged_rows and not sd["local_rows"] and not sd["remote_rows"]:
                continue
            ws_out = wb_out.create_sheet(sheet_name)
            for r, row_list in enumerate(merged_rows, start=1):
                key_str = _core._cell_str(row_list[0]) if row_list else ""
                rtype = row_types.get(key_str) or ""
                fill = None
                if rtype == "新增":
                    fill = green_fill
                elif rtype == "修改":
                    fill = yellow_fill
                elif rtype == "冲突":
                    fill = red_fill
                for c, val in enumerate(row_list, start=1):
                    cell = ws_out.cell(row=r, column=c, value=val)
                    if rtype and fill:
                        cell.fill = fill
                    if rtype == "冲突":
                        cell.font = red_font

        if not wb_out.sheetnames:
            wb_out.create_sheet("Data")
        os.makedirs(os.path.dirname(self.path_merged) or ".", exist_ok=True)
        wb_out.save(self.path_merged)
        wb_out.close()

        os.makedirs(backup_dir, exist_ok=True)
        shutil.copy2(self.path_local, os.path.join(backup_dir, base_name + "_local.xlsx"))
        shutil.copy2(self.path_remote, os.path.join(backup_dir, base_name + "_remote.xlsx"))
        shutil.copy2(self.path_merged, os.path.join(backup_dir, base_name + "_merged.xlsx"))

    def run(self):
        self.root.mainloop()


class DiffWindow:
    """对比 GUI：内嵌 diff 表格，关闭时自动清理缓存"""
    def __init__(self, path_a, path_b):
        self.path_a = path_a
        self.path_b = path_b
        self.path_out = None
        self.diff_rows = []
        self.is_temp = False
        self.root = tk.Tk()
        self.root.title("Excel 二向对比")
        self.root.minsize(700, 500)
        self.root.geometry("1000x620")
        _setup_merge_styles(self.root)
        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_ui(self):
        pad = 12
        top = ttk.Frame(self.root, padding=(pad, pad, pad, 6))
        top.pack(fill=tk.X)
        ttk.Label(top, text="本地 (A) vs 线上 (B)", font=("Segoe UI", 11, "bold")).pack(anchor=tk.W)
        ttk.Label(top, text="A: %s" % (self.path_a[:70] + "…" if len(self.path_a) > 70 else self.path_a), font=("Segoe UI", 8)).pack(anchor=tk.W)
        ttk.Label(top, text="B: %s" % (self.path_b[:70] + "…" if len(self.path_b) > 70 else self.path_b), font=("Segoe UI", 8)).pack(anchor=tk.W)

        # 对比结果颜色说明（表中行背景色含义）
        legend_frame = tk.Frame(self.root, bg="#f0f2f5")
        legend_frame.pack(fill=tk.X, padx=pad, pady=(0, 6))
        _make_color_legend(legend_frame, [
            ("#CCFFCC", "绿色=B新增"),
            ("#FFCCCC", "红色=A独有"),
            ("#FFFF99", "黄色=修改"),
            (None, "无=相同"),
        ]).pack(anchor=tk.W)

        paned = ttk.PanedWindow(self.root, orient=tk.VERTICAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0, 6))

        table_frame = ttk.LabelFrame(paned, text="差异列表", padding=6)
        paned.add(table_frame, weight=2)
        cols = ("Sheet", "Key", "状态", "A(本地)", "B(线上)")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=14)
        for c, w in [("Sheet", 80), ("Key", 120), ("状态", 70), ("A(本地)", 200), ("B(线上)", 200)]:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w)
        sb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        sbh = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb.set, xscrollcommand=sbh.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        sbh.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.tag_configure("B新增", background="#ccffcc")
        self.tree.tag_configure("A独有", background="#ffcccc")
        self.tree.tag_configure("修改", background="#ffff99")

        self.status_var = tk.StringVar(value="")
        status_bar = ttk.Frame(self.root)
        status_bar.pack(fill=tk.X, padx=pad, pady=(0, 4))
        ttk.Label(status_bar, textvariable=self.status_var, font=("Segoe UI", 9)).pack(anchor=tk.W)

        try:
            out_dir = os.path.dirname(os.path.abspath(self.path_a))
            base_name = os.path.splitext(os.path.basename(self.path_a))[0]
            self.path_out = os.path.join(out_dir, base_name + _core._COMPARE_SUFFIX + ".xlsx")
            self.is_temp = "Temp" in self.path_a or "Fork" in self.path_a or "tmp" in self.path_a.lower()

            _gui_log("正在计算差异…", self.status_var)
            path_out, sheet_names, self.diff_rows = _core.get_compare_data(self.path_a, self.path_b)
            if path_out is None:
                raise RuntimeError("get_compare_data 失败")
            self.path_out = path_out

            _core._write_compare_excel(self.path_out, sheet_names, self.diff_rows, open_file=False)
            _gui_log("已生成对比文件，共 %d 行差异" % len(self.diff_rows), self.status_var)

            for sheet_name, key, status, str_a, str_b in self.diff_rows:
                sa = (str_a[:60] + "…") if len(str_a) > 60 else str_a
                sb_val = (str_b[:60] + "…") if len(str_b) > 60 else str_b
                self.tree.insert("", tk.END, values=(sheet_name, key, status, sa, sb_val), tags=(status,))
        except Exception as e:
            import traceback
            _gui_log("对比失败: " + str(e), self.status_var, is_error=True)
            messagebox.showerror("错误", str(e))

        btn_frame = ttk.Frame(self.root, padding=(pad, 8))
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="打开 Excel", command=self._open_excel, style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_frame, text="关闭", command=self._on_close).pack(side=tk.LEFT)

    def _open_excel(self):
        if self.path_out and os.path.isfile(self.path_out):
            if sys.platform == "win32":
                os.startfile(self.path_out)
            elif sys.platform == "darwin":
                subprocess.run(["open", self.path_out], check=False)
            else:
                subprocess.run(["xdg-open", self.path_out], check=False)
            _gui_log("已打开对比文件", self.status_var)
        else:
            messagebox.showwarning("提示", "对比文件不存在")

    def _on_close(self):
        if self.is_temp and self.path_out and os.path.isfile(self.path_out):
            try:
                os.remove(self.path_out)
            except Exception:
                pass
        self.root.quit()
        self.root.destroy()
        sys.exit(0)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    # 直接运行本脚本时，用 TestData 打开合并窗口便于预览界面
    _root = os.path.dirname(os.path.abspath(__file__))
    _local = os.path.join(_root, "TestData", "local.xlsx")
    _base = os.path.join(_root, "TestData", "base.xlsx")
    _remote = os.path.join(_root, "TestData", "remote.xlsx")
    _merged = os.path.join(_root, "TestData", "_output", "merged.xlsx")
    if os.path.isfile(_local) and os.path.isfile(_base) and os.path.isfile(_remote):
        os.makedirs(os.path.dirname(_merged) or ".", exist_ok=True)
        win = MergeWindow(_local, _base, _remote, _merged)
        win.run()
    else:
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "Excel 合并/对比 GUI",
            "请通过以下方式打开界面：\n\n"
            "• 合并：运行 MergeExcelFork.py 并传入 4 个参数\n"
            "  python MergeExcelFork.py <local> <base> <remote> <merged>\n\n"
            "• 对比：运行 MergeExcelFork.py 并传入 2 个参数\n"
            "  python MergeExcelFork.py <local> <remote>\n\n"
            "或在 Fork 中配置 Merge Tool / Diff Tool 后触发。\n\n"
            "若需本地预览，请确保 TestData 下有 local.xlsx、base.xlsx、remote.xlsx。"
        )
        root.destroy()
